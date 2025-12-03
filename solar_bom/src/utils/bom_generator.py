import pandas as pd
import json
import os
from typing import Dict, List, Any, Optional
from ..models.block import BlockConfig, WiringType, WiringConfig, HarnessGroup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models.device import HarnessConnection, CombinerBoxConfig
import re


class BOMGenerator:
    """Utility class for generating Bill of Materials from block configurations"""
    
    # Constants for BOM calculations
    CABLE_WASTE_FACTOR = 1.05  # 5% extra for waste/installation
    
    def __init__(self, blocks: Dict[str, BlockConfig], project=None):
        """
        Initialize BOM Generator
        
        Args:
            blocks: Dictionary of block configurations
            project: Project object (optional)
        """
        self.blocks = blocks
        self.project = project
        # Load library json files
        self.harness_library = self.load_harness_library()
        self.fuse_library = self.load_fuse_library()
        self.combiner_box_fuse_library = self.load_combiner_box_fuse_library()
        self.combiner_box_library = self.load_combiner_box_library()
        self.extender_library = self.load_extender_library()
        self.whip_library = self.load_whip_library()

    def _get_cable_size(self, harness_group: Optional[HarnessGroup], cable_type: str, block: BlockConfig) -> str:
        """
        Get cable size from harness group with fallback to block-level wiring config
        
        Args:
            harness_group: The harness group (may be None for non-harness configurations)
            cable_type: Type of cable ('string', 'extender', 'whip', or 'harness')
            block: The block configuration for fallback
            
        Returns:
            The cable size string
        """
        if harness_group:
            # For harness cable, always use the harness group's cable_size
            if cable_type == 'harness':
                return harness_group.cable_size
                
            # For other cable types, check if harness has a specific size
            # Note: We check hasattr and the value is not None and not empty string
            if cable_type == 'string':
                if hasattr(harness_group, 'string_cable_size') and harness_group.string_cable_size:
                    return harness_group.string_cable_size
            elif cable_type == 'extender':
                if hasattr(harness_group, 'extender_cable_size') and harness_group.extender_cable_size:
                    return harness_group.extender_cable_size
            elif cable_type == 'whip':
                if hasattr(harness_group, 'whip_cable_size') and harness_group.whip_cable_size:
                    return harness_group.whip_cable_size
        
        # Fall back to block-level wiring config
        if cable_type == 'string':
            return getattr(block.wiring_config, 'string_cable_size', "10 AWG")
        elif cable_type == 'extender':
            return getattr(block.wiring_config, 'extender_cable_size', "8 AWG")
        elif cable_type == 'whip':
            return getattr(block.wiring_config, 'whip_cable_size', "8 AWG")
        elif cable_type == 'harness':
            # This shouldn't happen, but provide a default
            return "8 AWG"
        
        return "8 AWG"  # Default fallback
    
    def _parse_route_harness_info(self, route_id: str) -> tuple:
        """
        Parse route_id to extract tracker and harness indices
        
        Args:
            route_id: Route identifier (e.g., "pos_extender_0_h1", "neg_whip_t2-h1_whip")
            
        Returns:
            tuple: (tracker_idx, harness_idx) or (None, None) if not found
        """
        try:
            tracker_idx = None
            harness_idx = None
            
            # Handle format like "pos_whip_t1-h1_whip"
            if '_t' in route_id and '-h' in route_id:
                # Extract the tX-hY part
                import re
                match = re.search(r't(\d+)-h(\d+)', route_id)
                if match:
                    tracker_idx = int(match.group(1)) - 1  # Convert to 0-based index
                    harness_idx = int(match.group(2)) - 1  # Convert to 0-based index
                    return tracker_idx, harness_idx
            
            # Handle old format like "pos_extender_0_h1"
            parts = route_id.split('_')
            
            # Look for tracker index (usually after cable type)
            for i, part in enumerate(parts):
                if part.isdigit():
                    tracker_idx = int(part)
                    # Check if next part has harness index
                    if i + 1 < len(parts) and parts[i + 1].startswith('h'):
                        harness_idx = int(parts[i + 1][1:])
                    break
            
            return tracker_idx, harness_idx
        except (ValueError, IndexError) as e:
            print(f"DEBUG: Parse error: {e}")
            return None, None
    
    def _get_cable_size_for_segment(self, block: BlockConfig, segment, cable_type: str) -> str:
        """
        Get cable size for a specific segment, determining which harness it belongs to
        """
        block_name = getattr(block, 'block_id', 'Unknown Block')
        
        # Handle new dict format with route information
        if isinstance(segment, dict) and 'route_id' in segment:
            route_id = segment['route_id']
            tracker_idx, harness_idx = self._parse_route_harness_info(route_id)
            
            # If we found harness info, try to get harness-specific size
            if tracker_idx is not None and harness_idx is not None:
                # Get the tracker to find string count
                if tracker_idx < len(block.tracker_positions):
                    pos = block.tracker_positions[tracker_idx]
                    string_count = len(pos.strings)
                    
                    # Look up the harness group
                    if (hasattr(block.wiring_config, 'harness_groupings') and
                        string_count in block.wiring_config.harness_groupings):
                        harness_groups = block.wiring_config.harness_groupings[string_count]
                        if harness_idx < len(harness_groups):
                            harness_group = harness_groups[harness_idx]                            
                            # Get cable size from harness group
                            if cable_type == 'string' and harness_group.string_cable_size:
                                return harness_group.string_cable_size
                            elif cable_type == 'extender' and harness_group.extender_cable_size:
                                return harness_group.extender_cable_size
                            elif cable_type == 'whip' and harness_group.whip_cable_size:
                                return harness_group.whip_cable_size
                            elif cable_type == 'harness':
                                return harness_group.cable_size
        
        # Fall back to block-level defaults
        if cable_type == 'string':
            size = block.wiring_config.string_cable_size
        elif cable_type == 'extender':
            size = block.wiring_config.extender_cable_size
        elif cable_type == 'whip':
            size = block.wiring_config.whip_cable_size
        elif cable_type == 'harness':
            size = block.wiring_config.harness_cable_size
        else:
            size = "8 AWG"
        
        return size
    
    def calculate_cable_quantities(self) -> Dict[str, Dict[str, Any]]:
        """
        Calculate cable quantities by block and type, separating positive and negative
        
        Returns:
            Dictionary with quantities by block and component type
        """
        quantities = {}
        
        for block_id, block in self.blocks.items():
            
            block_quantities = {}
            
            # Get tracker counts categorized by strings per tracker
            tracker_counts = {}
            for pos in block.tracker_positions:
                if pos.template:
                    strings_per_tracker = pos.template.strings_per_tracker
                    key = f"{strings_per_tracker}-String Tracker"
                    if key not in tracker_counts:
                        tracker_counts[key] = 0
                    tracker_counts[key] += 1
            
            # Add tracker counts to quantities
            for tracker_type, count in tracker_counts.items():
                block_quantities[tracker_type] = {
                    'description': tracker_type,
                    'quantity': count,
                    'unit': 'units',
                    'category': 'Structural'
                }
            
            # Calculate cable lengths
            cable_lengths = block.calculate_cable_lengths()
            
            if not block.wiring_config:
                quantities[block_id] = block_quantities
                continue
            
            if block.wiring_config.wiring_type == WiringType.HOMERUN:
                # Split string cable by polarity
                string_cable_size = getattr(block.wiring_config, 'string_cable_size', '10 AWG')
                
                if 'string_cable_positive' in cable_lengths:
                    string_length_feet = round(cable_lengths['string_cable_positive'] * 3.28084 * self.CABLE_WASTE_FACTOR, 1)
                    
                    block_quantities[f'Positive String Cable ({string_cable_size})'] = {
                        'description': f'DC Positive String Cable {string_cable_size}',
                        'quantity': string_length_feet,
                        'unit': 'feet',
                        'category': 'eBOS'
                    }
                
                if 'string_cable_negative' in cable_lengths:
                    string_length_feet = round(cable_lengths['string_cable_negative'] * 3.28084 * self.CABLE_WASTE_FACTOR, 1)
                    
                    block_quantities[f'Negative String Cable ({string_cable_size})'] = {
                        'description': f'DC Negative String Cable {string_cable_size}',
                        'quantity': string_length_feet,
                        'unit': 'feet',
                        'category': 'eBOS'
                    }
                    
                # Get whip cable size - check if we have harness-specific sizes
                whip_cable_sizes = {}  # Change to dict to track which harnesses use which size
                if hasattr(block.wiring_config, 'harness_groupings'):
                    for string_count, harness_groups in block.wiring_config.harness_groupings.items():
                        for hg_idx, hg in enumerate(harness_groups):
                            whip_size = self._get_cable_size(hg, 'whip', block)
                            key = f"{string_count}_{hg_idx}"
                            whip_cable_sizes[key] = whip_size

                # Check if all harnesses use the same whip size
                unique_whip_sizes = set(whip_cable_sizes.values())
                if len(unique_whip_sizes) == 1:
                    whip_cable_size = unique_whip_sizes.pop()
                elif len(unique_whip_sizes) == 0:
                    whip_cable_size = getattr(block.wiring_config, 'whip_cable_size', '8 AWG')
                else:
                    # Multiple sizes - this will be handled in segment analysis
                    whip_cable_size = None
                        
            else:  # HARNESS configuration
                # Add harnesses by number of strings they connect - split by polarity
                harness_info = self._count_harnesses_by_size(block)
                
                # Add each harness type, separately for positive and negative
                for key, info in harness_info.items():
                    string_count = info['string_count']
                    count = info['count']
                    harness_group = info['harness']
                    
                    # Get cable sizes with fallback
                    harness_cable_size = self._get_cable_size(harness_group, 'harness', block)
                    string_cable_size = self._get_cable_size(harness_group, 'string', block)
                    
                    # Calculate string spacing for harness matching
                    if block.tracker_template and block.tracker_template.module_spec:
                        module_spec = block.tracker_template.module_spec
                        modules_per_string = block.tracker_template.modules_per_string
                        module_spacing_m = block.tracker_template.module_spacing_m
                        
                        string_spacing_ft = self.calculate_string_spacing_ft(
                            modules_per_string, module_spec.width_mm, module_spacing_m
                        )
                    else:
                        string_spacing_ft = 102.0  # Default spacing
                        
                    # Get descriptions from harness library
                    pos_description = self.get_harness_description(
                        string_count, 'positive', string_spacing_ft, harness_cable_size, string_cable_size
                    )
                    neg_description = self.get_harness_description(
                        string_count, 'negative', string_spacing_ft, harness_cable_size, string_cable_size
                    )

                    block_quantities[f'Positive {string_count}-String Harness'] = {
                        'description': pos_description,
                        'quantity': count,
                        'unit': 'units',
                        'category': 'eBOS'
                    }

                    block_quantities[f'Negative {string_count}-String Harness'] = {
                        'description': neg_description,
                        'quantity': count,
                        'unit': 'units',
                        'category': 'eBOS'
                    }
                
                # Count fuses by rating
                fuse_counts = self._count_fuses_by_rating(block)
                for rating, count in fuse_counts.items():
                    block_quantities[f'DC String Fuse {rating}A'] = {
                        'description': self.get_fuse_description(rating),
                        'quantity': count,
                        'unit': 'units',
                        'category': 'eBOS'
                    }
                
                # Get whip cable size - check if we have harness-specific sizes
                whip_cable_sizes = {}  # Change to dict to track which harnesses use which size
                if hasattr(block.wiring_config, 'harness_groupings'):
                    for string_count, harness_groups in block.wiring_config.harness_groupings.items():
                        for hg_idx, hg in enumerate(harness_groups):
                            whip_size = self._get_cable_size(hg, 'whip', block)
                            key = f"{string_count}_{hg_idx}"
                            whip_cable_sizes[key] = whip_size

                # Check if all harnesses use the same whip size
                unique_whip_sizes = set(whip_cable_sizes.values())
                if len(unique_whip_sizes) == 1:
                    whip_cable_size = unique_whip_sizes.pop()
                elif len(unique_whip_sizes) == 0:
                    whip_cable_size = getattr(block.wiring_config, 'whip_cable_size', '8 AWG')
                else:
                    # Multiple sizes - this will be handled in segment analysis
                    whip_cable_size = None

                # Get extender cable size - check if we have harness-specific sizes
                extender_cable_sizes = {}  # Change to dict to track which harnesses use which size
                if hasattr(block.wiring_config, 'harness_groupings'):
                    for string_count, harness_groups in block.wiring_config.harness_groupings.items():
                        for hg_idx, hg in enumerate(harness_groups):
                            extender_size = self._get_cable_size(hg, 'extender', block)
                            key = f"{string_count}_{hg_idx}"
                            extender_cable_sizes[key] = extender_size

                # Check if all harnesses use the same extender size
                unique_extender_sizes = set(extender_cable_sizes.values())
                if len(unique_extender_sizes) == 1:
                    extender_cable_size = unique_extender_sizes.pop()
                elif len(unique_extender_sizes) == 0:
                    extender_cable_size = getattr(block.wiring_config, 'extender_cable_size', '8 AWG')
                else:
                    # Multiple sizes - this will be handled in segment analysis
                    extender_cable_size = None
            
            quantities[block_id] = block_quantities
        
        # Store the original quantities before segment analysis
        original_quantities = {}
        for block_id, block_quantities in quantities.items():
            original_quantities[block_id] = {k: v.copy() for k, v in block_quantities.items() if 'Whip Cable' in k and v['unit'] == 'feet'}
        
        quantities = self.analyze_wire_segments(quantities)

        return quantities
        
    def _count_harnesses_by_size(self, block: BlockConfig) -> Dict[str, Dict]:
        """
        Count harnesses by number of strings they connect and track cable sizes
        
        Args:
            block: Block configuration
            
        Returns:
            Dictionary mapping harness key to count and cable size info
        """
        harness_info = {}
        
        if not block.wiring_config or block.wiring_config.wiring_type != WiringType.HARNESS:
            return harness_info
        
        # Check if we have custom harness groupings
        has_custom_groupings = (hasattr(block.wiring_config, 'harness_groupings') and 
                            block.wiring_config.harness_groupings)
        
        if has_custom_groupings:
            # Process each tracker's harness configuration
            for string_count, harness_groups in block.wiring_config.harness_groupings.items():
                # Count trackers with this string count
                tracker_count = sum(1 for pos in block.tracker_positions if len(pos.strings) == string_count)
                
                # For each harness group configuration
                for harness_idx, harness in enumerate(harness_groups):
                    # The key identifies this specific harness configuration
                    actual_string_count = len(harness.string_indices)
                    
                    # Create unique key that includes cable sizes to differentiate harnesses
                    cable_key = f"{harness.cable_size}_{harness.string_cable_size or 'default'}_{harness.extender_cable_size or 'default'}_{harness.whip_cable_size or 'default'}"
                    key = f"{actual_string_count}string_{cable_key}"
                    
                    if key not in harness_info:
                        harness_info[key] = {
                            'count': 0,
                            'string_count': actual_string_count,
                            'harness': harness  # Store the actual harness object
                        }
                    harness_info[key]['count'] += tracker_count
                
                # Check for unconfigured strings
                all_configured_indices = set()
                for harness in harness_groups:
                    all_configured_indices.update(harness.string_indices)
                
                unconfigured_strings = string_count - len(all_configured_indices)
                if unconfigured_strings > 0:
                    # Create default harness for unconfigured strings
                    key = f"{unconfigured_strings}string_default"
                    if key not in harness_info:
                        harness_info[key] = {
                            'count': 0,
                            'string_count': unconfigured_strings,
                            'harness': None  # Will use defaults
                        }
                    harness_info[key]['count'] += tracker_count
        else:
            # Default: one harness per tracker
            for pos in block.tracker_positions:
                string_count = len(pos.strings)
                key = f"{string_count}string_default"
                if key not in harness_info:
                    harness_info[key] = {
                        'count': 0,
                        'string_count': string_count,
                        'harness': None
                    }
                harness_info[key]['count'] += 1
        
        return harness_info
    
    def generate_summary_data(self, quantities: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
        """
        Generate summary data for BOM with categories
        
        Args:
            quantities: Component quantities by block
            
        Returns:
            DataFrame with summary data
        """
        # Initialize dictionaries to track totals by component, description, and category
        component_totals = {}
        
        # Sum up quantities for each component type across all blocks
        for block_id, block_quantities in quantities.items():
            for component_type, details in block_quantities.items():
                description = details['description']
                quantity = details['quantity']
                unit = details['unit']
                category = details['category']
                
                # Skip tracker entries (structural category with "Tracker" in component type)
                if category == 'Structural' and 'Tracker' in component_type:
                    continue
                    
                key = (component_type, description, unit, category)
                if key not in component_totals:
                    component_totals[key] = 0
                component_totals[key] += quantity
        
        # Convert to DataFrame
        summary_data = []
        for (component_type, description, unit, category), quantity in component_totals.items():
            summary_data.append({
                'Category': category,
                'Component Type': component_type,
                'Part Number': '',
                'Description': description,
                'Quantity': round(quantity, 1) if unit == 'feet' else int(quantity),
                'Unit': unit
            })
        
        # Sort by category then component type, with numerical sorting for segments
        def sort_key(item):
            category = item['Category']
            component_type = item['Component Type']
            
            # For segment entries, extract the numeric part for proper numerical sorting
            if 'Segment' in component_type and 'ft' in component_type:
                import re
                # Extract the numeric part from patterns like "Positive Whip Cable Segment 115ft (8 AWG)"
                match = re.search(r'Segment (\d+)ft', component_type)
                if match:
                    length = int(match.group(1))
                    # Create a sort key that groups by cable type and sorts numerically by length
                    cable_type = component_type.split(' Segment')[0]  # e.g., "Positive Whip Cable"
                    return (category, cable_type, length)
            
            # For non-segment entries, sort normally
            return (category, component_type, 0)

        summary_data = sorted(summary_data, key=sort_key)

        # Add part numbers to summary data
        for item in summary_data:
            item['Part Number'] = self.get_component_part_number(item)

        # Add part numbers to summary data
        for item in summary_data:
            item['Part Number'] = self.get_component_part_number(item)

        # Create DataFrame with specific column order
        df = pd.DataFrame(summary_data)
        if not df.empty:
            # Reorder columns to put Part Number after Component Type
            columns = list(df.columns)
            if 'Part Number' in columns and 'Component Type' in columns:
                # Remove Part Number from its current position
                columns.remove('Part Number')
                # Insert it after Component Type
                ct_index = columns.index('Component Type')
                columns.insert(ct_index + 1, 'Part Number')
                df = df[columns]

        return df
    
    def get_component_part_number(self, item):
        """Get part number for a component based on its type and properties"""
        component_type = item.get('Component Type', '')
        description = item.get('Description', '')
        
                
        if 'Harness' in component_type:
            
            # Extract info from description to find harness part number
            polarity = 'positive' if 'Positive' in description else 'negative'
            
            # Extract string count - handle both numeric and word forms
            import re
            
            # First try numeric pattern (e.g., "1 String" or "2-String")
            string_match = re.search(r'(\d+)[\s-]String', description)
            if string_match:
                num_strings = int(string_match.group(1))
            else:
                # Try word forms
                if 'Two String' in description:
                    num_strings = 2
                elif 'Three String' in description:
                    num_strings = 3
                elif 'One String' in description or '1 String' in description:
                    num_strings = 1
                else:
                    return "N/A"
            
            # Get module specs from first block
            if self.blocks:
                first_block = next(iter(self.blocks.values()))
                
                # Extract trunk cable size from description (e.g., "10AWG Drops w/8AWG Trunk")
                trunk_match = re.search(r'w/(\d+)\s*AWG\s+Trunk', description)
                if trunk_match:
                    trunk_cable_size = f"{trunk_match.group(1)} AWG"
                else:
                    # Fall back to block default
                    trunk_cable_size = getattr(first_block.wiring_config, 'harness_cable_size', '8 AWG')
                
                if first_block.tracker_template and first_block.tracker_template.module_spec:
                    module_spec = first_block.tracker_template.module_spec
                    modules_per_string = first_block.tracker_template.modules_per_string
                    module_spacing_m = first_block.tracker_template.module_spacing_m
                    
                    string_spacing_ft = self.calculate_string_spacing_ft(
                        modules_per_string, module_spec.width_mm, module_spacing_m
                    )
                    
                    part_number = self.find_matching_harness_part_number(
                        num_strings, polarity, string_spacing_ft, trunk_cable_size
                    )
                    
                    # If no match found and trunk size is non-standard, return CUSTOM
                    if part_number == "N/A":
                        standard_trunk_sizes = ["8 AWG", "10 AWG"]
                        if trunk_cable_size not in standard_trunk_sizes:
                            return "CUSTOM"
                    
                    return part_number
            return "N/A"
        
        elif 'Fuse' in component_type:
            # Extract fuse rating from description
            import re
            rating_match = re.search(r'(\d+)A', description)
            if rating_match:
                rating = int(rating_match.group(1))
                return self.get_fuse_part_number_by_rating(rating)
        
        # Handle cable segments - check for "Segment" and cable type
        elif 'Segment' in component_type:
            if 'Whip Cable' in component_type:
                return self.get_whip_segment_part_number_from_item(item)
            elif 'Extender Cable' in component_type:
                return self.get_extender_segment_part_number_from_item(item)
            elif 'String Cable' in component_type:
                return self.get_extender_segment_part_number_from_item(item)
        return "N/A"
    
    def get_whip_segment_part_number_from_item(self, item):
        """Get whip cable segment part number from item data"""
        try:
            component_type = item['Component Type']
            
            # Extract polarity from component type
            polarity = 'positive' if 'Positive' in component_type else 'negative'
            
            # Extract length from component type (e.g., "Positive Whip Cable Segment 25ft (8 AWG)")
            import re
            length_match = re.search(r'(\d+)ft', component_type)
            if not length_match:
                return "N/A"
            length_ft = int(length_match.group(1))
            
            # Extract wire gauge from component type
            gauge_match = re.search(r'\(([^)]+)\)', component_type)
            if not gauge_match:
                return "N/A"
            wire_gauge = gauge_match.group(1)
            
            # Find matching whip part number
            return self.find_matching_whip_part_number(wire_gauge, polarity, length_ft)
            
        except Exception as e:
            print(f"Error getting whip segment part number: {e}")
            return "N/A"

    def get_extender_segment_part_number_from_item(self, item):
        """Get extender cable segment part number from item data"""
        try:
            component_type = item['Component Type']
            
            # Extract polarity from component type
            polarity = 'positive' if 'Positive' in component_type else 'negative'
            
            # Extract length from component type (e.g., "Positive Extender Cable Segment 30ft (8 AWG)")
            import re
            length_match = re.search(r'(\d+)ft', component_type)
            if not length_match:
                return "N/A"
            length_ft = int(length_match.group(1))
            
            # Extract wire gauge from component type
            gauge_match = re.search(r'\(([^)]+)\)', component_type)
            if not gauge_match:
                return "N/A"
            wire_gauge = gauge_match.group(1)
            
            # Find matching extender part number
            return self.find_matching_extender_part_number(wire_gauge, polarity, length_ft)
            
        except Exception as e:
            print(f"Error getting extender segment part number: {e}")
            return "N/A"
    
    def generate_detailed_data(self, quantities: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
        """
        Generate detailed data for BOM
        
        Args:
            quantities: Component quantities by block
            
        Returns:
            DataFrame with detailed data
        """
        detailed_data = []
        
        # Calculate strings per block
        strings_per_block = {}
        for block_id, block in self.blocks.items():
            total_strings = 0
            for pos in block.tracker_positions:
                total_strings += len(pos.strings)
            strings_per_block[block_id] = total_strings
        
        for block_id, block_quantities in quantities.items():
            for component_type, details in block_quantities.items():
                description = details['description']
                quantity = details['quantity']
                unit = details['unit']
                category = details['category']
                
                # Skip tracker entries (structural category with "Tracker" in component type)
                if category == 'Structural' and 'Tracker' in component_type:
                    continue
                    
                item_data = {
                    'Block': block_id,
                    'Strings': strings_per_block.get(block_id, 0),
                    'Category': category,
                    'Component Type': component_type,
                    'Description': description,
                    'Quantity': round(quantity, 1) if unit == 'feet' else int(quantity),
                    'Unit': unit
                }

                # Add part number
                item_data['Part Number'] = self.get_component_part_number(item_data)

                detailed_data.append(item_data)

        
        # Sort by block ID, category, and component type
        detailed_data = sorted(detailed_data, key=lambda x: (x['Block'], x['Category'], x['Component Type']))

        # Create DataFrame and ensure column order
        df = pd.DataFrame(detailed_data)
        if not df.empty:
            # Define the column order with Strings after Block
            column_order = ['Block', 'Strings', 'Category', 'Component Type', 'Description', 'Quantity', 'Unit', 'Part Number']
            # Only reorder columns that exist
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]

        return df

    def generate_block_allocation_data(self) -> pd.DataFrame:
        """
        Generate block allocation data showing strings per block
        
        Returns:
            DataFrame with block allocation information
        """
        allocation_data = []
        
        for block_id, block in self.blocks.items():
            # Count total strings in this block
            total_strings = 0
            tracker_count = len(block.tracker_positions)
            
            # Count strings by tracker configuration
            tracker_configs = {}
            for pos in block.tracker_positions:
                string_count = len(pos.strings)
                total_strings += string_count
                
                # Track configuration counts
                config_key = f"{string_count}-String"
                if config_key not in tracker_configs:
                    tracker_configs[config_key] = 0
                tracker_configs[config_key] += 1
            
            # Create tracker breakdown string
            tracker_breakdown = ", ".join([
                f"{count} x {config}" for config, count in sorted(tracker_configs.items())
            ])
            
            allocation_data.append({
                'Block ID': block_id,
                'Total Strings': total_strings,
                'Number of Trackers': tracker_count,
                'Tracker Configuration': tracker_breakdown,
                'Wiring Type': block.wiring_config.wiring_type.value if block.wiring_config else 'Not Configured'
            })
        
        # Sort by block ID
        allocation_data = sorted(allocation_data, key=lambda x: x['Block ID'])
        
        return pd.DataFrame(allocation_data)

    def export_bom_to_excel_with_preview_data(self, filepath: str, project_info: Optional[Dict[str, Any]] = None, 
                                          preview_data: List[Dict] = None) -> bool:
        """
        Export BOM to Excel using preview data from UI
        
        Args:
            filepath: Path to save the Excel file
            project_info: Optional dictionary with project information
            preview_data: Data from the UI preview with correct part numbers
            
        Returns:
            True if export successful, False otherwise
        """
        writer = None
        try:
            # Count combiner boxes first (moved up from later in the code)
            combiner_box_count = self.count_combiner_boxes()
            
            # Get device configurations if available
            device_configs = {}
            combiner_bom_items = []  # Store combiner BOM items once
            if combiner_box_count > 0:
                # Try to get device configurations from the UI if available
                if hasattr(self, 'parent') and hasattr(self.parent, 'device_configurator'):
                    device_configurator = self.parent.device_configurator
                    if hasattr(device_configurator, 'combiner_configs'):
                        device_configs = device_configurator.combiner_configs
                        # Generate combiner BOM items ONCE
                        combiner_bom_items = self.generate_combiner_box_bom(device_configs)
                # Try to get device configurations from the UI if available
                if hasattr(self, 'parent') and hasattr(self.parent, 'device_configurator'):
                    device_configurator = self.parent.device_configurator
                    if hasattr(device_configurator, 'combiner_configs'):
                        device_configs = device_configurator.combiner_configs
            
            # Use preview data if provided
            if preview_data:
                summary_data = pd.DataFrame(preview_data)
                
                # Add combiner box BOM items to summary if they exist
                if combiner_bom_items:  # Use the already generated items
                    # Check if combiner box items are already in the preview data
                    # If they are, don't add them again
                    existing_part_numbers = set()
                    if 'Part Number' in summary_data.columns:
                        existing_part_numbers = set(summary_data['Part Number'].dropna())
                    
                    # Filter out items that are already in the summary
                    new_items = []
                    for item in combiner_bom_items:
                        if item['Part Number'] not in existing_part_numbers:
                            new_items.append(item)
                    
                    if new_items:
                        # Convert to dataframe with proper columns
                        cb_bom_df = pd.DataFrame(new_items)
                        cb_bom_df = cb_bom_df[['Category', 'Component Type', 'Part Number', 'Description', 'Quantity', 'Unit']]
                        
                        # Append to summary data
                        summary_data = pd.concat([summary_data, cb_bom_df], ignore_index=True)
                
                # Add Category column (infer from component type)
                def get_category(component_type):
                    if 'Whip Cable Segment' in component_type:
                        return 'eBOS Segments'
                    elif 'Extender Cable Segment' in component_type:
                        return 'Extender Cable Segments'
                    elif 'String Cable' in component_type:
                        return 'eBOS'
                    elif 'Harness' in component_type:
                        return 'eBOS'
                    elif 'Fuse' in component_type:
                        return 'Electrical'
                    else:
                        return 'Other'
                
                summary_data['Category'] = summary_data['Component Type'].apply(get_category)
                
                # Reorder columns to match expected format
                column_order = ['Category', 'Component Type', 'Part Number', 'Description', 'Quantity', 'Unit']
                summary_data = summary_data[column_order]
            else:
                # Fall back to original method
                quantities = self.calculate_cable_quantities()
                summary_data = self.generate_summary_data(quantities)
            
            # Generate detailed data (this can still use the original method)
            quantities = self.calculate_cable_quantities() 
            detailed_data = self.generate_detailed_data(quantities)
            
            # Generate project info if not provided
            if project_info is None:
                project_info = self.generate_project_info()
            
            # Create Excel writer
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            
            # Write summary data
            summary_data.to_excel(writer, sheet_name='BOM Summary', index=False, startrow=15)  # Start after project info
            
            # Write detailed data  
            detailed_data.to_excel(writer, sheet_name='Block Details', index=False)
            
            # Add harness cable size information to Block Details sheet
            self._add_harness_cable_info_to_block_details(writer, detailed_data)

            # Add combiner box sheet if there are any combiner boxes
            # (combiner_box_count and device_configs already defined above)
            if combiner_box_count > 0:

                # Generate combiner box data
                combiner_data = self.generate_combiner_box_data(device_configs)
                
                # Create the Combiner Boxes sheet manually without pandas formatting
                workbook = writer.book
                if 'Combiner Boxes' in workbook.sheetnames:
                    worksheet = workbook['Combiner Boxes']
                else:
                    worksheet = workbook.create_sheet('Combiner Boxes')
                    writer.sheets['Combiner Boxes'] = worksheet
                
                # Define border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                current_row = 1
                
                # Write BOM if it exists
                if combiner_bom_items:
                    # Add BOM section header
                    cell = worksheet.cell(row=current_row, column=1, value="Combiner Box BOM")
                    cell.font = Font(bold=True, size=14)
                    current_row += 1
                    
                    # Add BOM headers with custom layout
                    # Category (A), Component Type (B), Part Number (C), Description (D-I merged), Quantity (J), Unit (K)
                    headers_config = [
                        ('Category', 1, 1),  # Column A
                        ('Component Type', 2, 1),  # Column B
                        ('Part Number', 3, 1),  # Column C
                        ('Description', 4, 6),  # Columns D-I (6 columns)
                        ('Quantity', 10, 1),  # Column J
                        ('Unit', 11, 1),  # Column K
                    ]
                    
                    for header, start_col, col_span in headers_config:
                        if col_span > 1:
                            # Merge cells for Description
                            worksheet.merge_cells(start_row=current_row, start_column=start_col, 
                                                end_row=current_row, end_column=start_col + col_span - 1)
                        cell = worksheet.cell(row=current_row, column=start_col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Add borders to all cells in the merge
                        for col in range(start_col, start_col + col_span):
                            worksheet.cell(row=current_row, column=col).border = thin_border
                    current_row += 1
                    
                    # Write BOM data with custom layout
                    for _, row_data in pd.DataFrame(combiner_bom_items).iterrows():
                        # Category (A)
                        cell = worksheet.cell(row=current_row, column=1, value=row_data.get('Category', ''))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Component Type (B)
                        cell = worksheet.cell(row=current_row, column=2, value=row_data.get('Component Type', ''))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Part Number (C)
                        cell = worksheet.cell(row=current_row, column=3, value=row_data.get('Part Number', ''))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Description (D-I merged)
                        worksheet.merge_cells(start_row=current_row, start_column=4, 
                                            end_row=current_row, end_column=9)
                        cell = worksheet.cell(row=current_row, column=4, value=row_data.get('Description', ''))
                        cell.alignment = Alignment(horizontal='left', vertical='center')  # Left align description
                        # Add borders to all cells in the merge
                        for col in range(4, 10):
                            worksheet.cell(row=current_row, column=col).border = thin_border
                        
                        # Quantity (J)
                        cell = worksheet.cell(row=current_row, column=10, value=row_data.get('Quantity', ''))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Unit (K)
                        cell = worksheet.cell(row=current_row, column=11, value=row_data.get('Unit', ''))
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        current_row += 1
                    
                    # Add spacing
                    current_row += 2
                    
                    # Add configuration details header
                    cell = worksheet.cell(row=current_row, column=1, value="Combiner Box Configuration Details")
                    cell.font = Font(bold=True, size=14)
                    current_row += 1
                
                # Write configuration details
                if not combiner_data.empty:
                    # Add headers
                    for col, header in enumerate(combiner_data.columns, 1):
                        cell = worksheet.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1
                    
                    # Write data
                    for _, row_data in combiner_data.iterrows():
                        for col, value in enumerate(row_data, 1):
                            cell = worksheet.cell(row=current_row, column=col, value=value)
                            if value is not None and value != '':
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Generate and add Block Allocation sheet
            block_allocation_data = self.generate_block_allocation_data()
            block_allocation_data.to_excel(writer, sheet_name='Block Allocation', index=False)
            
            # Format the sheets (same as original method)
            workbook = writer.book
            summary_sheet = writer.sheets['BOM Summary']
            
            # Add project info to summary sheet
            row = 1
            summary_sheet.merge_cells(f'A{row}:F{row}')
            project_info_cell = summary_sheet.cell(row=row, column=1, value="Project Information")
            project_info_cell.font = Font(bold=True, size=14)
            project_info_cell.alignment = Alignment(horizontal='center')
            
            if project_info:
                # First set of info in columns A and B
                main_info = ['Project Name', 'Customer', 'Location', 'System Size (kW DC)', 
                            'Number of Modules', 'Module Manufacturer', 'Module Model', 
                            'Inverter Manufacturer', 'Inverter Model', 'DC Collection']
                
                # Additional info for columns C and D
                additional_info = {
                    'String Size': project_info.get('String Size', 'Unknown'),
                    'Number of Strings': project_info.get('Number of Strings', 0),
                    'Module Wiring': project_info.get('Module Wiring', 'Unknown'),
                    'Module Dimensions': project_info.get('Module Dimensions', 'Unknown'),
                    'Number of Combiner Boxes': project_info.get('Number of Combiner Boxes', 0)
                }
                
                # Write main info with units
                row = 2
                for key in main_info:
                    if key in project_info:
                        value = project_info[key]
                        
                        # Add units to specific fields
                        if key == 'System Size (kW DC)' and isinstance(value, (int, float)):
                            value = f"{round(value, 2)} kW"
                        elif key == 'Number of Modules' and isinstance(value, int):
                            value = f"{value} modules"
                        
                        summary_sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
                        summary_sheet.cell(row=row, column=2, value=value)
                        row += 1
                
                # Write additional info in columns C and D
                row = 2
                for key, value in additional_info.items():
                    # Add units to specific fields
                    if key == 'String Size' and value != 'Unknown':
                        value = f"{value} modules per string"
                    elif key == 'Number of Strings' and isinstance(value, int):
                        value = f"{value} strings"
                    elif key == 'Number of Combiner Boxes' and isinstance(value, int):
                        if value == 1:
                            value = f"{value} combiner box"
                        else:
                            value = f"{value} combiner boxes"
                    
                    summary_sheet.cell(row=row, column=3, value=key).font = Font(bold=True)
                    summary_sheet.cell(row=row, column=4, value=value)
                    row += 1

            # Add disclaimer note
            row = 12
            summary_sheet.merge_cells(f'A{row}:F{row+1}')  # Merge cells A12:F13
            disclaimer_cell = summary_sheet.cell(row=row, column=1, 
                value="Preliminary cable sizes - to be reviewed by Electrical Engineer of Record before ordering")
            disclaimer_cell.font = Font(bold=True, color="FF0000", size=11)  # Red bold text
            disclaimer_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add border around merged cells
            for row_num in range(12, 14):  # Rows 12 and 13
                for col_num in range(1, 7):  # Columns A through F
                    cell = summary_sheet.cell(row=row_num, column=col_num)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Add section header for BOM
            row = 14
            summary_sheet.merge_cells(f'A{row}:E{row}')
            summary_sheet.cell(row=row, column=1, value="Bill of Materials").font = Font(bold=True, size=14)
            
            # Format sheets
            self._format_excel_sheet(workbook['BOM Summary'], summary_data, start_row=15)
            self._format_excel_sheet(workbook['Block Details'], detailed_data)

            # Format block allocation sheet
            if 'Block Allocation' in workbook.sheetnames:
                self._format_block_allocation_sheet(workbook['Block Allocation'], block_allocation_data)
            
            # Add filter
            summary_sheet.auto_filter.ref = f"A15:F{15 + len(summary_data)}"
            
            # Save and open
            writer.close()
            writer = None
            
            try:
                os.startfile(filepath)
            except Exception as e:
                print(f"File was saved but could not be opened automatically: {str(e)}")

            return True
            
        except Exception as e:
            print(f"Error exporting BOM: {str(e)}")
            if isinstance(e, PermissionError):
                raise
            return False
        finally:
            if writer is not None:
                try:
                    writer.close()
                except:
                    pass

    def _add_harness_cable_info_to_block_details(self, writer, detailed_data):
        """Add harness-specific cable size information to Block Details sheet"""
        try:
            worksheet = writer.sheets['Block Details']
            
            # Find the last row with data
            last_data_row = len(detailed_data) + 1  # +1 for header
            
            # Add some spacing
            info_start_row = last_data_row + 3
            
            # Check if any blocks have custom harness cable sizes
            has_custom_sizes = False
            harness_info = []
            
            for block_id, block in self.blocks.items():
                if (hasattr(block, 'wiring_config') and 
                    block.wiring_config and 
                    block.wiring_config.wiring_type == WiringType.HARNESS and
                    hasattr(block.wiring_config, 'harness_groupings')):
                    
                    for string_count, harness_list in block.wiring_config.harness_groupings.items():
                        for harness_idx, harness in enumerate(harness_list):
                            # Check if this harness has custom cable sizes
                            if (hasattr(harness, 'string_cable_size') and harness.string_cable_size) or \
                               (hasattr(harness, 'extender_cable_size') and harness.extender_cable_size) or \
                               (hasattr(harness, 'whip_cable_size') and harness.whip_cable_size):
                                has_custom_sizes = True
                                
                                # Build info string
                                actual_string_count = len(harness.string_indices)
                                info = {
                                    'Block': block_id,
                                    'Harness Type': f"{actual_string_count}-string harness",
                                    'String Cable': harness.string_cable_size if hasattr(harness, 'string_cable_size') and harness.string_cable_size else f"Default ({block.wiring_config.string_cable_size})",
                                    'Harness Cable': harness.cable_size,
                                    'Extender Cable': harness.extender_cable_size if hasattr(harness, 'extender_cable_size') and harness.extender_cable_size else f"Default ({block.wiring_config.extender_cable_size})",
                                    'Whip Cable': harness.whip_cable_size if hasattr(harness, 'whip_cable_size') and harness.whip_cable_size else f"Default ({block.wiring_config.whip_cable_size})"
                                }
                                harness_info.append(info)
            
            if has_custom_sizes and harness_info:
                # Add header
                worksheet.cell(row=info_start_row, column=1, value="Harness-Specific Cable Sizes:").font = Font(bold=True, size=12)
                
                # Add column headers
                headers = ['Block', 'Harness Type', 'String Cable', 'Harness Cable', 'Extender Cable', 'Whip Cable']
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=info_start_row + 1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Add data rows
                for row_idx, info in enumerate(harness_info, info_start_row + 2):
                    for col_idx, (key, value) in enumerate(info.items(), 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Highlight custom sizes
                        if key in ['String Cable', 'Extender Cable', 'Whip Cable'] and not value.startswith('Default'):
                            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                
                # Auto-fit columns
                for col_idx in range(1, 7):
                    column_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[column_letter].auto_size = True
                    
        except Exception as e:
            print(f"Error adding harness cable info to Excel: {e}")

    def filter_data_by_checked_components(self, data_df, checked_components, is_detailed=False):
        """Filter DataFrame based on checked components"""
        if not checked_components:
            return data_df
        
        # Create set of checked component descriptions for fast lookup
        checked_descriptions = {comp['description'] for comp in checked_components}
        
        # Filter DataFrame
        if is_detailed:
            # For detailed data, filter by Description column
            filtered_df = data_df[data_df['Description'].isin(checked_descriptions)]
        else:
            # For summary data, filter by Description column
            filtered_df = data_df[data_df['Description'].isin(checked_descriptions)]
        
        return filtered_df.reset_index(drop=True)                       

    def _format_excel_sheet(self, worksheet, data: pd.DataFrame, start_row: int = 1):
        """
        Format Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet
            data: DataFrame with data
            start_row: Row to start formatting from (default=1)
        """
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Warning style
        warning_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        
        # Format headers
        for col_num, column_title in enumerate(data.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border

        # Apply center alignment to Part Number column data
        if 'Part Number' in data.columns:
            part_number_col = list(data.columns).index('Part Number') + 1
            for row_num in range(start_row + 1, start_row + len(data) + 2):  # +2 to include last row
                cell = worksheet.cell(row=row_num, column=part_number_col)
                cell.alignment = centered_alignment
        
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row+1, 
                                                    max_row=start_row+len(data)+1,  # +1 to ensure last row is included
                                                    min_col=1, 
                                                    max_col=len(data.columns)), 1):
            # Get the row index
            row_index = start_row + i
            
            # Check if this is a warning row
            is_warning = False
            if 'Unit' in data.columns:
                unit_col_idx = list(data.columns).index('Unit')
                unit_value = worksheet.cell(row=row_index, column=unit_col_idx + 1).value
                if unit_value and 'warning' in str(unit_value).lower():
                    is_warning = True
            
            # Apply warning formatting if this is a warning row
            if is_warning:
                for cell in row:
                    cell.fill = warning_fill
            
            # Add borders to all cells
            for cell in row:
                cell.border = border
        
        # Auto-adjust column width with maximum constraints
        for column in worksheet.columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            
            # Calculate the maximum content length in this column
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            # Apply width with constraints
            adjusted_width = max_length + 2
            
            # Set maximum widths for specific columns
            if column_name == 'A':  # Category column
                adjusted_width = min(adjusted_width, 25)  # Max width of 25 for column A
            elif column_name == 'B':  # Component Type column  
                adjusted_width = min(adjusted_width, 35)  # Max width of 35 for column B
            elif column_name == 'D':  # Description column
                adjusted_width = min(adjusted_width, 50)  # Max width of 50 for description
            else:
                adjusted_width = min(adjusted_width, 30)  # General max width for other columns
            
            # Also set a minimum width
            adjusted_width = max(adjusted_width, 10)  # Minimum width of 10
            
            worksheet.column_dimensions[column_name].width = adjusted_width

    def generate_project_info(self) -> Dict[str, Any]:
        """
        Generate general project information for BOM header
        
        Returns:
            Dictionary with project info
        """
        # Initialize with default values
        info = {
            'System Size (kW DC)': 0,
            'Number of Modules': 0,
            'Module Manufacturer': 'Unknown',
            'Module Model': 'Unknown',
            'DC Collection': 'Unknown',
            'Inverter Manufacturer': 'Unknown',
            'Inverter Model': 'Unknown',
            'String Size': 0,
            'Number of Strings': 0,
            'Module Wiring': 'Unknown',
            'Module Dimensions': 'Unknown',
            'Number of Combiner Boxes': 0
        }
        
        # Calculate total system size and module count
        total_modules = 0
        module_manufacturer = set()
        module_model = set()
        inverter_manufacturer = set()
        inverter_model = set()
        dc_collection_types = set()
        
        # New counters for additional info
        total_strings = 0
        string_sizes = set()
        module_dimensions = set()
        
        for block_id, block in self.blocks.items():
            # Count modules
            if block.tracker_template and block.tracker_template.module_spec:
                module_spec = block.tracker_template.module_spec
                block_modules = 0
                
                # Count modules directly from strings
                for pos in block.tracker_positions:
                    for string in pos.strings:
                        block_modules += string.num_modules
                
                total_modules += block_modules
                
                # Add module info
                module_manufacturer.add(module_spec.manufacturer)
                module_model.add(module_spec.model)
                
                # Calculate system size
                info['System Size (kW DC)'] += (block_modules * module_spec.wattage) / 1000
                
                # Get module dimensions
                dim_str = f"{int(module_spec.length_mm)} mm x {int(module_spec.width_mm)} mm"
                module_dimensions.add(dim_str)

            # Count combiner boxes
            combiner_box_count = self.count_combiner_boxes()
            info['Number of Combiner Boxes'] = combiner_box_count
            
            # Count strings and get string size
            if block.tracker_positions:
                for pos in block.tracker_positions:
                    total_strings += len(pos.strings)
            
            # Get string size from tracker template
            if block.tracker_template:
                string_sizes.add(block.tracker_template.modules_per_string)
            
            # Add inverter info
            if block.inverter:
                inverter_manufacturer.add(block.inverter.manufacturer)
                inverter_model.add(block.inverter.model)
            
            # Add DC collection type
            if block.wiring_config:
                dc_collection_types.add(block.wiring_config.wiring_type.value)
        
        # Set the collected values
        info['Number of Modules'] = total_modules
        info['Module Manufacturer'] = ', '.join(module_manufacturer) if module_manufacturer else 'Unknown'
        info['Module Model'] = ', '.join(module_model) if module_model else 'Unknown'
        info['Inverter Manufacturer'] = ', '.join(inverter_manufacturer) if inverter_manufacturer else 'Unknown'
        info['Inverter Model'] = ', '.join(inverter_model) if inverter_model else 'Unknown'
        info['DC Collection'] = ', '.join(dc_collection_types) if dc_collection_types else 'Unknown'
        
        # Set new fields
        info['String Size'] = ', '.join(str(s) for s in sorted(string_sizes)) if string_sizes else 'Unknown'
        info['Number of Strings'] = total_strings
        info['Module Dimensions'] = ', '.join(module_dimensions) if module_dimensions else 'Unknown'
        info['Number of Combiner Boxes'] = len(self.blocks)
        
        # Get module wiring from project
        if hasattr(self.project, 'wiring_mode'):
            wiring_mode = "Leapfrog" if self.project.wiring_mode == 'leapfrog' else "Daisy Chain"
            info['Module Wiring'] = wiring_mode
        
        return info

    def analyze_wire_segments(self, quantities: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """Analyze wire segments and add segment counts to quantities"""
        for block_id, block in self.blocks.items():
            block_quantities = quantities.get(block_id, {})
            
            # Skip blocks without wiring config
            if not block.wiring_config:
                continue
            
            # Get realistic cable routes if available, otherwise use regular routes
            cable_routes = getattr(block.wiring_config, 'realistic_cable_routes', {})
            if not cable_routes:
                cable_routes = block.wiring_config.cable_routes
            
            # Group routes by type and analyze segments
            string_segments_pos = []
            string_segments_neg = []
            whip_segments_pos = []
            whip_segments_neg = []
            extender_segments_pos = []
            extender_segments_neg = []
            
            for route_id, route in cable_routes.items():
                # Skip routes with less than 2 points
                if len(route) < 2:
                    continue
                    
                # Calculate segment length
                segment_length = 0
                for i in range(len(route) - 1):
                    dx = route[i+1][0] - route[i][0]
                    dy = route[i+1][1] - route[i][1]
                    segment_length += (dx**2 + dy**2)**0.5
                
                # Convert to feet
                segment_length_feet = segment_length * 3.28084
                
                # Categorize by route type and polarity
                if "pos_src" in route_id or "pos_node" in route_id or "pos_string" in route_id:
                    string_segments_pos.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })
                elif "neg_src" in route_id or "neg_node" in route_id or "neg_string" in route_id:
                    string_segments_neg.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })
                elif "pos_dev" in route_id or "pos_main" in route_id or "whip_pos" in route_id or "pos_whip" in route_id:
                    # Add underground routing component if enabled
                    if hasattr(block, 'underground_routing') and block.underground_routing:
                        underground_addition_m = 2 * (block.pile_reveal_m + block.trench_depth_m)
                        underground_addition_ft = underground_addition_m * 3.28084
                        segment_length_feet += underground_addition_ft
                    whip_segments_pos.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })
                elif "neg_dev" in route_id or "neg_main" in route_id or "whip_neg" in route_id or "neg_whip" in route_id:
                    # Add underground routing component if enabled
                    if hasattr(block, 'underground_routing') and block.underground_routing:
                        underground_addition_m = 2 * (block.pile_reveal_m + block.trench_depth_m)
                        underground_addition_ft = underground_addition_m * 3.28084
                        segment_length_feet += underground_addition_ft
                    whip_segments_neg.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })
            
            # Process string segments only for HOMERUN wiring
            if block.wiring_config.wiring_type == WiringType.HOMERUN:
                string_size = block.wiring_config.string_cable_size
                self._add_segment_analysis(block_quantities, string_segments_pos, 
                                        string_size, "Positive String Cable", 5)
                self._add_segment_analysis(block_quantities, string_segments_neg, 
                                        string_size, "Negative String Cable", 5)
                
                # Calculate and add total string entries from segments
                self.calculate_totals_from_segments(block_quantities, string_size, "Positive String Cable")
                self.calculate_totals_from_segments(block_quantities, string_size, "Negative String Cable")
            
            # Process whip segments for all wiring types
            # For harness configurations, we need to group segments by cable size
            if block.wiring_config.wiring_type == WiringType.HARNESS and hasattr(block.wiring_config, 'harness_groupings'):
                # Group whip segments by cable size
                whip_segments_by_size_pos = {}
                whip_segments_by_size_neg = {}
                
                # Analyze each segment to determine its cable size
                for segment in whip_segments_pos:
                    # Determine which harness this segment belongs to based on its route
                    cable_size = self._get_cable_size_for_segment(block, segment, 'whip')
                    
                    if cable_size not in whip_segments_by_size_pos:
                        whip_segments_by_size_pos[cable_size] = []
                    whip_segments_by_size_pos[cable_size].append(segment)
                
                for segment in whip_segments_neg:
                    cable_size = self._get_cable_size_for_segment(block, segment, 'whip')
                    
                    if cable_size not in whip_segments_by_size_neg:
                        whip_segments_by_size_neg[cable_size] = []
                    whip_segments_by_size_neg[cable_size].append(segment)
                
                # Process each cable size group separately
                for cable_size, segments in whip_segments_by_size_pos.items():
                    self._add_segment_analysis(block_quantities, segments, 
                                            cable_size, f"Positive Whip Cable ({cable_size})", 1)
                    self.calculate_totals_from_segments(block_quantities, cable_size, f"Positive Whip Cable ({cable_size})")

                for cable_size, segments in whip_segments_by_size_neg.items():
                    self._add_segment_analysis(block_quantities, segments, 
                                            cable_size, f"Negative Whip Cable ({cable_size})", 1)
                    self.calculate_totals_from_segments(block_quantities, cable_size, f"Negative Whip Cable ({cable_size})")
            else:
                # Non-harness or single cable size - process normally
                whip_size = getattr(block.wiring_config, 'whip_cable_size', "8 AWG")
                self._add_segment_analysis(block_quantities, whip_segments_pos, 
                                        whip_size, "Positive Whip Cable", 1)
                self._add_segment_analysis(block_quantities, whip_segments_neg, 
                                        whip_size, "Negative Whip Cable", 1)
                
                self.calculate_totals_from_segments(block_quantities, whip_size, "Positive Whip Cable")
                self.calculate_totals_from_segments(block_quantities, whip_size, "Negative Whip Cable")

            # Process extender segments
            extender_segments_pos = []
            extender_segments_neg = []

            for route_id, route in cable_routes.items():
                # Skip routes with less than 2 points
                if len(route) < 2:
                    continue
                    
                # Calculate segment length
                segment_length = 0
                for i in range(len(route) - 1):
                    dx = route[i+1][0] - route[i][0]
                    dy = route[i+1][1] - route[i][1]
                    segment_length += (dx**2 + dy**2)**0.5
                
                # Convert to feet
                segment_length_feet = segment_length * 3.28084
                
                # Categorize extender routes
                if "pos_extender" in route_id:
                    extender_segments_pos.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })
                elif "neg_extender" in route_id:
                    extender_segments_neg.append({
                        'length': segment_length_feet,
                        'route_id': route_id
                    })

            # Process extender segments - handle harness-specific cable sizes
            if block.wiring_config.wiring_type == WiringType.HARNESS and hasattr(block.wiring_config, 'harness_groupings'):
                # Group extender segments by cable size
                extender_segments_by_size_pos = {}
                extender_segments_by_size_neg = {}
                
                # Analyze each segment to determine its cable size
                for segment in extender_segments_pos:
                    # Determine which harness this segment belongs to based on its route
                    cable_size = self._get_cable_size_for_segment(block, segment, 'extender')
                    
                    if cable_size not in extender_segments_by_size_pos:
                        extender_segments_by_size_pos[cable_size] = []
                    extender_segments_by_size_pos[cable_size].append(segment)
                
                for segment in extender_segments_neg:
                    cable_size = self._get_cable_size_for_segment(block, segment, 'extender')
                    
                    if cable_size not in extender_segments_by_size_neg:
                        extender_segments_by_size_neg[cable_size] = []
                    extender_segments_by_size_neg[cable_size].append(segment)
                
                # Process each cable size group separately
                for cable_size, segments in extender_segments_by_size_pos.items():
                    self._add_segment_analysis(block_quantities, segments, 
                                            cable_size, f"Positive Extender Cable ({cable_size})", 5)
                    self.calculate_totals_from_segments(block_quantities, cable_size, f"Positive Extender Cable ({cable_size})")

                for cable_size, segments in extender_segments_by_size_neg.items():
                    self._add_segment_analysis(block_quantities, segments, 
                                            cable_size, f"Negative Extender Cable ({cable_size})", 5)
                    self.calculate_totals_from_segments(block_quantities, cable_size, f"Negative Extender Cable ({cable_size})")
            else:
                # Non-harness or single cable size - process normally
                extender_size = getattr(block.wiring_config, 'extender_cable_size', "8 AWG")
                self._add_segment_analysis(block_quantities, extender_segments_pos, 
                                        extender_size, "Positive Extender Cable", 5)
                self._add_segment_analysis(block_quantities, extender_segments_neg, 
                                        extender_size, "Negative Extender Cable", 5)

                # Calculate and add total extender entries from segments
                self.calculate_totals_from_segments(block_quantities, extender_size, "Positive Extender Cable")
                self.calculate_totals_from_segments(block_quantities, extender_size, "Negative Extender Cable")
                    
            # Update quantities
            quantities[block_id] = block_quantities
        
        return quantities
    
    def load_harness_library(self):
        """Load harness library from JSON file"""
        try:
            # Get the path relative to this file (src/utils/bom_generator.py)
            current_dir = os.path.dirname(os.path.abspath(__file__))  # src/utils/
            project_root = os.path.dirname(os.path.dirname(current_dir))  # Go up two levels to get to project root
            library_path = os.path.join(project_root, 'data', 'harness_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading harness library: {e}")
            return {}
        
    def load_extender_library(self):
        """Load the extender library from JSON file"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            library_path = os.path.join(project_root, 'data', 'extender_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading extender library: {e}")
            return {}

    def load_whip_library(self):
        """Load the whip library from JSON file"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            library_path = os.path.join(project_root, 'data', 'whip_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading whip library: {e}")
            return {}
        
    def get_whip_description_format(self, wire_gauge):
        """Get whip description format from library"""
        # Find any whip with matching wire gauge to get description format
        for spec in self.whip_library.values():
            if spec.get('wire_gauge') == wire_gauge:
                # Extract format without the length
                desc = spec.get('description', '')
                # Replace the length part with placeholder
                import re
                return re.sub(r'LENGTH \(FT\): \d+', 'LENGTH (FT): {length}', desc)
        # Default format if not found
        return f"1500 VDC {wire_gauge} WHIP WITH MC4 CONNECTOR AND BLUNT CUT END, LENGTH (FT): {{length}}"

    def get_extender_description_format(self, wire_gauge):
        """Get extender description format from library"""
        # Find any extender with matching wire gauge to get description format
        for spec in self.extender_library.values():
            if spec.get('wire_gauge') == wire_gauge:
                # Extract format without the length
                desc = spec.get('description', '')
                # Replace the length part with placeholder
                import re
                return re.sub(r'LENGTH \(FT\): \d+', 'LENGTH (FT): {length}', desc)
        # Default format if not found
        return f"1500 VDC {wire_gauge} EXTENDER WITH MC4 CONNECTORS, LENGTH (FT): {{length}}"
    
    def get_harness_description(self, num_strings, polarity, string_spacing_ft, trunk_cable_size, string_cable_size):
        """Get harness description from library based on matching part number"""      
        # First find the matching part number
        part_number = self.find_matching_harness_part_number(
            num_strings, polarity, string_spacing_ft, trunk_cable_size
        )
        
        if part_number == "N/A" or " or " in part_number:
            # If no match or multiple matches, return a generic description
            pol_text = "Positive" if polarity == 'positive' else "Negative"
            string_text = "String" if num_strings == 1 else "String"
            fuse_text = ""
            
            # Try to determine if this should be fused based on polarity and string count
            if polarity == 'positive' and num_strings > 1:
                # Make an educated guess about fuse rating based on string count
                if num_strings <= 2:
                    fuse_text = "20A Fuses, " if string_spacing_ft > 110 else "Unfused, "
                else:
                    fuse_text = "30A Fuses, " if string_spacing_ft > 120 else "20A Fuses, "
            
            return f"{num_strings} {string_text}, {pol_text}, {fuse_text}{string_cable_size} Drops w/{trunk_cable_size} Trunk, {int(string_spacing_ft)}' string length, MC4 connectors"
        
        # If single part number found, get its description
        if part_number in self.harness_library:
            return self.harness_library[part_number].get('description', f"Harness {part_number}")
        
        # Handle multiple matches by using first one's description format
        if " or " in part_number:
            first_part = part_number.split(" or ")[0]
            if first_part in self.harness_library:
                return self.harness_library[first_part].get('description', f"Harness {first_part}")
        
        return f"Harness {part_number}"
    
    def get_fuse_description(self, fuse_rating_amps):
        """Get fuse description from library based on rating"""
        # Find the fuse part number by rating
        part_number = self.get_fuse_part_number_by_rating(fuse_rating_amps)
        
        # If we have a valid part number, get its description
        if part_number and part_number != "N/A" and not part_number.startswith("FUSE-"):
            if part_number in self.fuse_library:
                return self.fuse_library[part_number].get('description', f'DC String Fuse {fuse_rating_amps}A')
        
        # Default description if not found
        return f'DC String Fuse {fuse_rating_amps}A'

    def calculate_string_spacing_ft(self, modules_per_string, module_width_mm, module_spacing_m):
        """Calculate string spacing in feet based on module specs"""
        try:
            # Calculate total string length: (modules * width) + ((modules-1) * spacing between modules)
            total_length_mm = (modules_per_string * module_width_mm + 
                            (modules_per_string - 1) * module_spacing_m * 1000)
            
            # Convert to feet
            total_length_ft = total_length_mm / 1000 * 3.28084
            
            return round(total_length_ft, 1)
        except Exception as e:
            print(f"Error calculating string spacing: {e}")
            return 0

    def find_matching_harness_part_number(self, num_strings, polarity, calculated_spacing_ft, trunk_cable_size=None):
        """Find matching harness part number from library"""
        try:
            matches = []
            
            # Available harness spacing options based on your library
            # First Solar uses 26', Standard uses 102, 113, 122, 133
            available_spacings = [26.0, 102.0, 113.0, 122.0, 133.0]
            
            # Find the next largest available spacing
            target_spacing = None
            for spacing in sorted(available_spacings):
                if spacing >= calculated_spacing_ft:
                    target_spacing = spacing
                    break
            
            if target_spacing is None:
                target_spacing = max(available_spacings)  # Use largest if calculated is bigger than all
            
            # Search for matching harnesses
            for part_number, spec in self.harness_library.items():
                # Skip comment entries
                if part_number.startswith('_comment_'):
                    continue
                    
                # Check basic criteria
                if (spec.get('num_strings') == num_strings and 
                    spec.get('polarity') == polarity and
                    abs(spec.get('string_spacing_ft', 0) - target_spacing) < 0.1):
                                        
                    # If trunk cable size is specified, filter by it
                    if trunk_cable_size:
                        spec_trunk_size = spec.get('trunk_cable_size', spec.get('trunk_wire_gauge'))
                        if spec_trunk_size != trunk_cable_size:
                            continue
                            
                    matches.append(part_number)
            
            if matches:
                if len(matches) == 1:
                    return matches[0]
                else:
                    result = " or ".join(sorted(matches))
                    return result
            else:
                return "N/A"
                
        except Exception as e:
            print(f"Error finding harness match: {e}")
            return "N/A"
        
    def find_matching_extender_part_number(self, wire_gauge, polarity, required_length_ft):
        """Find matching extender part number from library"""
        try:
            # Round up to next 5ft increment for extender lengths
            target_length = ((required_length_ft - 1) // 5 + 1) * 5
            target_length = max(10, target_length)  # Remove the max 300 clamp
            
            # If over 300ft, return custom part number with wire gauge and polarity
            if target_length > 300:
                polarity_code = 'P' if polarity == 'positive' else 'N'
                awg_size = wire_gauge.split()[0]  # Extract just the number from "8 AWG"
                return f"EXT-{awg_size}-{polarity_code}-{target_length}-CUSTOM"
            
            # Search for matching extender
            for part_number, spec in self.extender_library.items():
                # Skip comment entries
                if part_number.startswith('_comment_'):
                    continue
                if (spec.get('wire_gauge') == wire_gauge and 
                    spec.get('polarity') == polarity and
                    spec.get('length_ft') == target_length):
                    return part_number
            
            return "N/A"
            
        except Exception as e:
            print(f"Error finding extender part number: {e}")
            return "N/A"

    def find_matching_whip_part_number(self, wire_gauge, polarity, required_length_ft):
        """Find matching whip part number from library"""
        try:
            # Round up to next 5ft increment for whip lengths
            target_length = ((required_length_ft - 1) // 5 + 1) * 5
            target_length = max(10, target_length)  # Remove the max 300 clamp

            # If over 300ft, return custom part number with wire gauge and polarity
            if target_length > 300:
                polarity_code = 'P' if polarity == 'positive' else 'N'
                awg_size = wire_gauge.split()[0]  # Extract just the number from "8 AWG"
                return f"WHI-{awg_size}-{polarity_code}-{target_length}-CUSTOM"
            
            # Search for matching whip
            for part_number, spec in self.whip_library.items():
                # Skip comment entries
                if part_number.startswith('_comment_'):
                    continue
                if (spec.get('wire_gauge') == wire_gauge and 
                    spec.get('polarity') == polarity and
                    spec.get('length_ft') == target_length):
                    return part_number

            return "N/A"
            
        except Exception as e:
            print(f"Error finding whip part number: {e}")
            return "N/A"
        
    def get_fuse_part_number_by_rating(self, fuse_rating_amps):
        """Get fuse part number by rating from fuse library"""
        try:
            # Find exact match first
            for part_number, spec in self.fuse_library.items():
                if spec.get('fuse_rating_amps') == fuse_rating_amps:
                    return part_number
            
            # If no exact match, find the next higher rating
            available_ratings = []
            for spec in self.fuse_library.values():
                rating = spec.get('fuse_rating_amps')
                if rating and rating >= fuse_rating_amps:
                    available_ratings.append((rating, spec.get('part_number')))
            
            if available_ratings:
                # Sort by rating and return the lowest that meets the requirement
                available_ratings.sort(key=lambda x: x[0])
                return available_ratings[0][1]
            
            # If no suitable fuse found, return placeholder
            return f"FUSE-{fuse_rating_amps}A-NOT-FOUND"
            
        except Exception as e:
            print(f"Error getting fuse part number: {e}")
            return f"FUSE-{fuse_rating_amps}A-ERROR"

    def _add_segment_analysis(self, block_quantities, segments, cable_size, cable_type_name, round_to=5):
        """Add individual cable segments to BOM"""
        
        if not segments:
            return
        
        # Convert segments to lengths only for processing
        segment_lengths = []
        for segment in segments:
            if isinstance(segment, dict) and 'length' in segment:
                segment_lengths.append(segment['length'])
            else:
                # Backward compatibility - segment is already a float
                segment_lengths.append(segment)
            
        # Add waste factor
        segment_lengths = [s * self.CABLE_WASTE_FACTOR for s in segment_lengths]
        
        # Always use 5ft increment for whip cables, otherwise use the passed-in increment
        actual_increment = 5 if "Whip Cable" in cable_type_name else round_to
        
        # Round up to nearest increment, with minimum of 10ft
        rounded_segments = []
        for s in segment_lengths:
            rounded_length = actual_increment * ((s + actual_increment - 0.1) // actual_increment + 1)
            # Ensure minimum length of 10ft
            rounded_length = max(10, rounded_length)
            rounded_segments.append(rounded_length)
        
        # Count segments by length
        segment_counts = {}
        for length in rounded_segments:
            if length not in segment_counts:
                segment_counts[length] = 0
            segment_counts[length] += 1
            
        # Add segment counts to quantities
        for length, count in segment_counts.items():
            segment_key = f"{cable_type_name} Segment {int(length)}ft ({cable_size})"
            
            # Determine category based on cable type
            if "Extender Cable" in cable_type_name:
                category = 'Extender Cable Segments'
            else:
                category = 'eBOS Segments'
            
            # Generate proper description based on cable type
            if "Whip Cable" in cable_type_name:
                base_desc = self.get_whip_description_format(cable_size)
                description = base_desc.format(length=int(length))
            elif "Extender Cable" in cable_type_name:
                base_desc = self.get_extender_description_format(cable_size)
                description = base_desc.format(length=int(length))
            else:
                description = f"{int(length)}ft {cable_type_name} Segment ({cable_size})"

            block_quantities[segment_key] = {
                'description': description,
                'quantity': count,
                'unit': 'segments',
                'category': category
            }

    def calculate_totals_from_segments(self, block_quantities, cable_size, prefix):
        """Calculate and add total cable entry from segment entries"""
        total_length = 0
        segment_entries = {}
        
        # Find all segment entries for this cable type
        for key, details in block_quantities.items():
            if prefix in key and "Segment" in key and details['unit'] == 'segments':
                # Extract length from key (format: "{prefix} Segment {length}ft ({cable_size})")
                try:
                    segment_str = key.split("Segment ")[1].split("ft")[0]
                    length = float(segment_str)
                    count = details['quantity']
                    total_length += length * count
                    segment_entries[key] = details
                except (ValueError, IndexError):
                    continue
        
        # Only add total if we found segments
        if total_length > 0:
            # For whip cables, ensure the total is an integer (no decimal places)
            # This maintains the 5ft increment pattern
            if "Whip Cable" in prefix:
                total_length = int(total_length)
            
            total_key = f"{prefix} ({cable_size})"

            # Determine category based on cable type
            if "Extender Cable" in prefix:
                category = 'Extender Cables'
            else:
                category = 'eBOS'

            # Generate proper description for totals
            if "Whip Cable" in prefix:
                description = self.get_whip_description_format(cable_size).format(length="TOTAL")
            elif "Extender Cable" in prefix:
                description = self.get_extender_description_format(cable_size).format(length="TOTAL")
            else:
                description = f'DC {prefix} {cable_size}'

            block_quantities[total_key] = {
                'description': description,
                'quantity': total_length if "Whip Cable" in prefix or "Extender Cable" in prefix else round(total_length, 1),
                'unit': 'feet',
                'category': category
            }
            
        return total_length > 0
    
    def _count_fuses_by_rating(self, block: BlockConfig) -> Dict[int, int]:
        """
        Count fuses by rating
        
        Args:
            block: Block configuration
            
        Returns:
            Dictionary mapping fuse rating to count
        """
        fuse_counts = {}
        
        if not block.wiring_config or block.wiring_config.wiring_type != WiringType.HARNESS:
            return fuse_counts
            
        # Check all string counts
        for string_count, harness_groups in getattr(block.wiring_config, 'harness_groupings', {}).items():
            # Count trackers with this string count - this is the key fix
            tracker_count = sum(1 for pos in block.tracker_positions if len(pos.strings) == string_count)
            
            for harness in harness_groups:
                # Only count fuses if use_fuse is True and there's more than one string
                use_fuse = getattr(harness, 'use_fuse', len(harness.string_indices) > 1)
                if use_fuse and len(harness.string_indices) > 1:
                    rating = getattr(harness, 'fuse_rating_amps', 15)
                    
                    # Count one fuse per string in the harness, multiplied by tracker count
                    num_strings = len(harness.string_indices)
                    fuse_counts[rating] = fuse_counts.get(rating, 0) + (num_strings * tracker_count)
        
        # Check for trackers without custom harness configuration
        if not hasattr(block.wiring_config, 'harness_groupings') or not block.wiring_config.harness_groupings:
            # Calculate default fuse rating
            default_rating = 15
            if block.tracker_template and block.tracker_template.module_spec:
                isc = block.tracker_template.module_spec.isc
                nec_min = isc * 1.25
                for rating in [5, 10, 15, 20, 25, 30, 35, 40, 45]:
                    if rating >= nec_min:
                        default_rating = rating
                        break
            
            # Count trackers with multiple strings
            for pos in block.tracker_positions:
                string_count = len(pos.strings)
                if string_count > 1:  # Only count for 2+ strings
                    # Add one fuse per string (positive side only)
                    fuse_counts[default_rating] = fuse_counts.get(default_rating, 0) + string_count
        
        return fuse_counts
    
    def load_fuse_library(self):
        """Load fuse library from JSON file"""
        try:
            # Get the path relative to this file
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            library_path = os.path.join(project_root, 'data', 'fuse_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading fuse library: {e}")
            return {}
        
    def load_combiner_box_fuse_library(self):
        """Load combiner box fuse library from JSON file"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            library_path = os.path.join(project_root, 'data', 'combiner_box_fuse_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading combiner box fuse library: {e}")
            return {}
    
    def load_combiner_box_library(self):
        """Load combiner box library from JSON file"""
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            library_path = os.path.join(project_root, 'data', 'combiner_box_library.json')
            
            with open(library_path, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading combiner box library: {e}")
            return {}
        
    def get_combiner_box_part_number(self, num_inputs: int, max_fuse_size: int, breaker_size: int, use_whips: bool) -> str:
        """
        Generate combiner box part number based on specifications
        
        Returns part number or 'CUSTOM' if not in standard library
        """
        # Determine input capacity (14 for High Amp, 30 for Standard Amp)
        if num_inputs <= 14:
            input_indicator = 14
        elif num_inputs <= 30:
            input_indicator = 30
        else:
            return "CUSTOM"  # Over 30 inputs
        
        # Determine fuse holder rating
        if max_fuse_size <= 20:
            fuse_indicator = 20
        elif max_fuse_size <= 32:
            fuse_indicator = 25
        else:
            fuse_indicator = 32  # 32 and above
            
        # Check if fuse size exceeds 65A (max in our library)
        if max_fuse_size > 65:
            return "CUSTOM"
        
        # Determine breaker size (round up to standard size)
        standard_breakers = [400, 500, 600]
        breaker_indicator = None
        for std_breaker in standard_breakers:
            if breaker_size <= std_breaker:
                breaker_indicator = std_breaker
                break
        
        if breaker_indicator is None:
            return "CUSTOM"  # Breaker over 600A
        
        # Determine whips
        whip_indicator = 2 if use_whips else 0
        
        # Generate part number
        part_number = f"CB-{input_indicator}-{fuse_indicator}-{breaker_indicator}-{whip_indicator}"
        
        # Check if it exists in library
        if part_number not in self.combiner_box_library:
            return "CUSTOM"
        
        return part_number
    
    def get_combiner_box_fuse_part_number(self, fuse_size: int) -> str:
        """Get combiner box fuse part number for given amperage"""
        part_number = f"CB-F-{fuse_size}"
        
        # Check if it exists in library
        if part_number not in self.combiner_box_fuse_library:
            return "CUSTOM"
        
        return part_number 

    def generate_combiner_box_bom(self, device_configs: Dict[str, 'CombinerBoxConfig']) -> List[Dict]:
        """
        Generate BOM data for combiner boxes and their fuses
        
        Returns list of BOM items with part numbers and quantities
        """
        bom_items = []
        
        if not device_configs:
            return bom_items
        
        # Simple dictionaries to track totals by part number
        combiner_totals = {}  # part_number -> quantity
        fuse_totals = {}      # part_number -> quantity
        
        for combiner_id, config in device_configs.items():
            if not config.connections:
                continue
            
            # Get max fuse size (uniform across all connections due to uniformity rule)
            max_fuse = max(conn.get_display_fuse_size() for conn in config.connections)
            
            # Get number of inputs
            num_inputs = len(config.connections)
            
            # Get breaker size
            breaker_size = config.get_display_breaker_size()
            
            # Get whips setting (default to True if not set)
            use_whips = getattr(config, 'use_whips', True)
            
            # Generate part numbers
            cb_part_number = self.get_combiner_box_part_number(
                num_inputs, max_fuse, breaker_size, use_whips
            )
            fuse_part_number = self.get_combiner_box_fuse_part_number(max_fuse)
                        
            # Simply add to totals
            combiner_totals[cb_part_number] = combiner_totals.get(cb_part_number, 0) + 1
            fuse_totals[fuse_part_number] = fuse_totals.get(fuse_part_number, 0) + num_inputs
        
        # Create BOM items for combiner boxes
        for cb_part, quantity in combiner_totals.items():
            cb_description = "Combiner Box"
            if cb_part in self.combiner_box_library:
                cb_description = self.combiner_box_library[cb_part].get('description', cb_description)
            elif cb_part == "CUSTOM":
                cb_description = "CUSTOM Combiner Box"
            
            bom_items.append({
                'Category': 'Electrical',
                'Component Type': 'Combiner Box',
                'Part Number': cb_part,
                'Description': cb_description,
                'Quantity': quantity,
                'Unit': 'each',
                'Custom': cb_part == "CUSTOM"
            })
        
        # Create BOM items for fuses
        for fuse_part, quantity in fuse_totals.items():
            fuse_description = "Combiner Box Fuse"
            if fuse_part in self.combiner_box_fuse_library:
                fuse_description = self.combiner_box_fuse_library[fuse_part].get('description', fuse_description)
            elif fuse_part == "CUSTOM":
                fuse_description = "CUSTOM Combiner Box Fuse"
            
            bom_items.append({
                'Category': 'Electrical',
                'Component Type': 'Combiner Box Fuse',
                'Part Number': fuse_part,
                'Description': fuse_description,
                'Quantity': quantity,
                'Unit': 'each',
                'Custom': fuse_part == "CUSTOM"
            })
        
        return bom_items
        
    def generate_combiner_box_data(self, device_configs: Dict[str, 'CombinerBoxConfig']) -> pd.DataFrame:
        """
        Generate combiner box configuration data for Excel export
        
        Args:
            device_configs: Dictionary of combiner box configurations
            
        Returns:
            DataFrame with combiner box data
        """
        combiner_data = []
        
        for combiner_id in sorted(device_configs.keys()):
            config = device_configs[combiner_id]
            
            # Add a header row for each combiner box
            for i, conn in enumerate(config.connections):
                row_data = {
                    'Combiner': combiner_id,
                    'Tracker': conn.tracker_id,
                    'Harness': conn.harness_id,
                    '# Strings': conn.num_strings,
                    'Module Isc': f"{conn.module_isc:.2f}",
                    'NEC Safety Factor': conn.nec_factor,
                    'Harness Current': f"{conn.harness_current:.2f}",
                    'Fuse Size': conn.get_display_fuse_size(),
                    'Cable Size': conn.get_display_cable_size(),
                    'Total Current': '',  # Only on first row
                    'Breaker Size': ''    # Only on first row
                }
                
                # Add total current and breaker size only to first row
                if i == 0:
                    row_data['Total Current'] = f"{config.total_input_current:.2f}"
                    row_data['Breaker Size'] = config.get_display_breaker_size()
                
                combiner_data.append(row_data)
            
            # Add empty row between combiner boxes (except after last one)
            if combiner_id != sorted(device_configs.keys())[-1]:
                combiner_data.append({})
        
        return pd.DataFrame(combiner_data)
    
    def count_combiner_boxes(self) -> int:
        """Count the number of combiner boxes in the project"""
        count = 0
        for block in self.blocks.values():
            if hasattr(block, 'device_type'):
                from ..models.block import DeviceType
                if block.device_type == DeviceType.COMBINER_BOX:
                    count += 1
        return count
    
    def _format_combiner_sheet(self, worksheet, data: pd.DataFrame):
        """
        Format the combiner box sheet with special handling for cable mismatches
        
        Args:
            worksheet: openpyxl worksheet
            data: DataFrame with combiner data
        """
        # Use standard formatting first
        self._format_excel_sheet(worksheet, data)
        
        # Additional formatting
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        red_font = Font(color="FF0000")
        centered_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply center alignment to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:  # Only format cells with content
                    cell.alignment = centered_alignment
        
        # Track if we found any mismatches
        has_mismatches = False
        
        # Check for cable size mismatches and apply red formatting
        if hasattr(self, 'parent') and hasattr(self.parent, 'device_configurator'):
            device_configurator = self.parent.device_configurator
            if hasattr(device_configurator, 'combiner_configs'):
                # Get the actual combiner configs to check for mismatches
                for row_idx in range(2, worksheet.max_row + 1):  # Skip header row
                    combiner_id = worksheet.cell(row=row_idx, column=1).value
                    if not combiner_id or combiner_id == '':  # Skip empty rows
                        continue
                        
                    # Get the config for this combiner
                    config = device_configurator.combiner_configs.get(combiner_id)
                    if config:
                        # Find which connection this row represents
                        tracker_id = worksheet.cell(row=row_idx, column=2).value
                        harness_id = worksheet.cell(row=row_idx, column=3).value
                        
                        # Find the matching connection
                        for conn in config.connections:
                            if conn.tracker_id == tracker_id and conn.harness_id == harness_id:
                                if conn.is_cable_size_mismatch():
                                    has_mismatches = True
                                    # Apply red background to entire row
                                    for col in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=row_idx, column=col)
                                        cell.fill = red_fill
                                break
        
        # Add legend/note to the right of the data if there are mismatches
        if has_mismatches:
            # Find the rightmost column with data
            last_data_col = worksheet.max_column
            note_col = last_data_col + 2  # Leave one column gap
            
            # Add header for the note
            note_header_cell = worksheet.cell(row=1, column=note_col, value="Notes")
            note_header_cell.font = Font(bold=True)
            note_header_cell.alignment = Alignment(horizontal='left')
            
            # Add the note about red highlighting
            note_cell = worksheet.cell(row=2, column=note_col, 
                                    value="Red rows indicate cable size mismatch with wiring configuration")
            note_cell.font = Font(color="FF0000", italic=True)
            note_cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            # Set column width for note
            worksheet.column_dimensions[get_column_letter(note_col)].width = 40

            # Highlight CUSTOM items in yellow
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Check for BOM section and highlight CUSTOM items
            for row in worksheet.iter_rows(min_row=2):  # Skip header
                part_number_cell = row[2] if len(row) > 2 else None  # Part Number is column C (index 2)
                if part_number_cell and part_number_cell.value == "CUSTOM":
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = yellow_fill

    def _format_block_allocation_sheet(self, worksheet, data: pd.DataFrame):
        """
        Format the Block Allocation sheet with proper styling
        
        Args:
            worksheet: openpyxl worksheet
            data: DataFrame with block allocation data
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_idx in range(1, len(data.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border
        
        # Format data rows
        for row_idx in range(2, len(data) + 2):
            for col_idx in range(1, len(data.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                
                # Center align numeric columns
                if col_idx in [2, 3]:  # Total Strings and Number of Trackers columns
                    cell.alignment = centered_alignment
        
        # Add summary row
        summary_row = len(data) + 3
        worksheet.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        
        # Calculate totals
        total_strings = data['Total Strings'].sum() if 'Total Strings' in data.columns else 0
        total_trackers = data['Number of Trackers'].sum() if 'Number of Trackers' in data.columns else 0
        
        worksheet.cell(row=summary_row, column=2, value=total_strings).font = Font(bold=True)
        worksheet.cell(row=summary_row, column=2).alignment = centered_alignment
        worksheet.cell(row=summary_row, column=3, value=total_trackers).font = Font(bold=True)
        worksheet.cell(row=summary_row, column=3).alignment = centered_alignment
        
        # Apply borders to summary row
        for col_idx in range(1, len(data.columns) + 1):
            cell = worksheet.cell(row=summary_row, column=col_idx)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='thin')
            )
        
        # Auto-fit columns
        for col_idx in range(1, len(data.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header
            column = worksheet[column_letter]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width