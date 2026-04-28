import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from typing import Optional, Dict, List, Any
from pathlib import Path
import json
import uuid
import math
import copy
from datetime import datetime
from collections import defaultdict
from src.utils.string_allocation import allocate_strings, allocate_strings_sequential, allocate_strings_spatial
from .site_preview import SitePreviewWindow, QuickEstimateDialog


class QuickEstimate(ttk.Frame):
    """Quick estimation tool for early-stage project sizing with hierarchical structure"""
    
    def __init__(self, parent, current_project=None, on_save=None):
        super().__init__(parent)
        self.current_project = current_project
        self.estimate_id = None
        self.on_save = on_save
        self.pricing_data = self.load_pricing_data()
        self._estimate_id_map = {}  # Map display name to ID
        self.available_modules = self.load_module_library()  # {display_name: ModuleSpec}
        self.selected_module = None  # Currently selected ModuleSpec
        self.available_inverters = self.load_inverter_library()  # {display_name: InverterSpec}
        self.selected_inverter = None  # Currently selected InverterSpec
        
        self.groups = []
        self.pads = []  # Collection points (inverter pads)
        self.measurements = []  # Saved measurement polylines [[wx,wy],...] per measurement
        self.device_names = {}  # {device_idx: "custom_name"} for CB/SI renaming
        self.device_feeder_sizes = {}  # {device_idx: "cable_size"} per-device feeder/homerun size
        self.device_feeder_parallel_counts = {}  # {device_idx: int} per-device parallel sets per pole
        self.last_combiner_assignments = []  # Structured CB data for Device Configurator
        self.last_si_assignments = []        # Structured SI data for Device Configurator
        self._harness_combos = []  # Track harness combo widgets for LV collection disabling
        self.selected_group_idx = None
        self._updating_listbox = False
        self.enabled_templates = self.load_enabled_templates()

        # Global settings defaults
        self.module_width_default = 1134
        self.modules_per_string_default = 28
        
        # Track currently selected item
        self.checked_items = set()  # Items checked for export
        self._results_stale = True
        self._calc_btn = None  # Reference to calculate button
        self._autosave_after_id = None
        
        # Allocation lock state
        self.allocation_locked = False
        self.locked_allocation_result = None
        self.manually_edited = False  # True once Edit Devices autosaves a change

        # Track the open Site Preview window (None when closed)
        self._site_preview_window = None

        # Guard to suppress re-entrant harness trace during on_template_change
        self._suppress_harness_trace = False

        self.setup_ui()
        
        # Load most recent estimate or show empty state
        self._refresh_estimate_dropdown(auto_select=True)

    def load_pricing_data(self):
        """Load pricing data from JSON file"""
        try:
            pricing_path = Path('data/pricing_data.json')
            if pricing_path.exists():
                with open(pricing_path, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading pricing data: {e}")
        return {}

    def load_module_library(self):
        """Load modules from the module templates JSON file"""
        modules = {}
        try:
            module_path = Path('data/module_templates.json')
            if module_path.exists():
                with open(module_path, 'r') as f:
                    data = json.load(f)
                
                from ..models.module import ModuleSpec, ModuleType, ModuleOrientation
                
                for manufacturer, models in data.items():
                    if not isinstance(models, dict):
                        continue
                    for model_name, spec_data in models.items():
                        try:
                            # Handle enum conversions
                            mod_type = ModuleType(spec_data.get('type', 'Mono PERC'))
                            orientation = ModuleOrientation(spec_data.get('default_orientation', 'Portrait'))
                            
                            module = ModuleSpec(
                                manufacturer=spec_data.get('manufacturer', manufacturer),
                                model=spec_data.get('model', model_name),
                                type=mod_type,
                                length_mm=float(spec_data.get('length_mm', 0)),
                                width_mm=float(spec_data.get('width_mm', 0)),
                                depth_mm=float(spec_data.get('depth_mm', 40)),
                                weight_kg=float(spec_data.get('weight_kg', 25)),
                                wattage=float(spec_data.get('wattage', 0)),
                                vmp=float(spec_data.get('vmp', 0)),
                                imp=float(spec_data.get('imp', 0)),
                                voc=float(spec_data.get('voc', 0)),
                                isc=float(spec_data.get('isc', 0)),
                                max_system_voltage=float(spec_data.get('max_system_voltage', 1500)),
                                temperature_coefficient_pmax=spec_data.get('temperature_coefficient_pmax'),
                                temperature_coefficient_voc=spec_data.get('temperature_coefficient_voc'),
                                temperature_coefficient_isc=spec_data.get('temperature_coefficient_isc'),
                                default_orientation=orientation,
                                cells_per_module=int(spec_data.get('cells_per_module', 72))
                            )
                            display_name = f"{module.manufacturer} {module.model} ({module.wattage}W)"
                            modules[display_name] = module
                        except (ValueError, TypeError) as e:
                            print(f"Error loading module {manufacturer}/{model_name}: {e}")
        except Exception as e:
            print(f"Error loading module library: {e}")
        return modules
    
    def load_inverter_library(self):
        """Load inverters from the inverters JSON file"""
        inverters = {}
        try:
            inverter_path = Path('data/inverters.json')
            if inverter_path.exists():
                with open(inverter_path, 'r') as f:
                    data = json.load(f)
                
                from ..models.inverter import InverterSpec, MPPTChannel, MPPTConfig, InverterType
                
                for name, specs in data.items():
                    try:
                        rated_power = specs.get('rated_power_kw', specs.get('rated_power', 10.0))
                        max_dc_power = specs.get('max_dc_power_kw', float(rated_power) * 1.5)
                        inverter_type_str = specs.get('inverter_type', 'String')
                        
                        channels = []
                        for ch in specs.get('mppt_channels', []):
                            channels.append(MPPTChannel(**ch))
                        
                        inverter = InverterSpec(
                            manufacturer=specs.get('manufacturer', 'Unknown'),
                            model=specs.get('model', 'Unknown'),
                            inverter_type=InverterType(inverter_type_str),
                            rated_power_kw=float(rated_power),
                            max_dc_power_kw=float(max_dc_power),
                            max_efficiency=float(specs.get('max_efficiency', 98.0)),
                            mppt_channels=channels,
                            mppt_configuration=MPPTConfig(specs.get('mppt_configuration', 'Independent')),
                            max_dc_voltage=float(specs.get('max_dc_voltage', 1500)),
                            startup_voltage=float(specs.get('startup_voltage', 150)),
                            nominal_ac_voltage=float(specs.get('nominal_ac_voltage', 400.0)),
                            max_ac_current=float(specs.get('max_ac_current', 40.0)),
                            power_factor=float(specs.get('power_factor', 0.99)),
                            dimensions_mm=tuple(specs.get('dimensions_mm', (1000, 600, 300))),
                            weight_kg=float(specs.get('weight_kg', 75.0)),
                            ip_rating=specs.get('ip_rating', 'IP65'),
                            max_short_circuit_current=specs.get('max_short_circuit_current')
                        )
                        display_name = f"{inverter.manufacturer} {inverter.model}"
                        inverters[display_name] = inverter
                    except Exception as e:
                        print(f"Warning: Failed to load inverter '{name}': {e}")
        except Exception as e:
            print(f"Error loading inverter library: {e}")
        return inverters

    def load_enabled_templates(self):
        """Load tracker templates that are enabled for the current project.
        Returns {template_key: template_data_dict} for enabled templates only."""
        templates = {}
        try:
            template_path = Path('data/tracker_templates.json')
            if not template_path.exists():
                return templates
            
            with open(template_path, 'r') as f:
                data = json.load(f)
            
            if not data:
                return templates
            
            # Parse hierarchical format
            all_templates = {}
            first_value = next(iter(data.values()))
            if isinstance(first_value, dict) and not any(key in first_value for key in ['module_orientation', 'modules_per_string']):
                for manufacturer, template_group in data.items():
                    for template_name, template_data in template_group.items():
                        unique_name = f"{manufacturer} - {template_name}"
                        all_templates[unique_name] = template_data
            else:
                all_templates = data
            
            # Filter to enabled only
            if self.current_project and hasattr(self.current_project, 'enabled_templates'):
                for key in self.current_project.enabled_templates:
                    if key in all_templates:
                        templates[key] = all_templates[key]
            else:
                templates = all_templates
                
        except Exception as e:
            print(f"Error loading enabled templates: {e}")
        
        return templates

    def get_template_display_name(self, template_key):
        """Build a short display name for a template key.
        E.g. 'JA Solar - JAM72S30-550/MR 2x28 3S' -> 'JAM72S30-550/MR 2x28 3S (3-string)'
        """
        template_data = self.enabled_templates.get(template_key)
        if not template_data:
            return template_key
        
        spt = template_data.get('strings_per_tracker', '?')
        # Strip manufacturer prefix for shorter display
        if ' - ' in template_key:
            short_name = template_key.split(' - ', 1)[1]
        else:
            short_name = template_key
        return f"{short_name} ({spt}S)"

    def get_template_module_id(self, template_key):
        """Get a module identity string from a template for consistency checking.
        Returns 'Manufacturer Model' or None."""
        template_data = self.enabled_templates.get(template_key)
        if not template_data:
            return None
        module_spec = template_data.get('module_spec', {})
        manufacturer = module_spec.get('manufacturer', '')
        model = module_spec.get('model', '')
        if manufacturer and model:
            return f"{manufacturer} {model}"
        return None

    def get_group_module_id(self, group):
        """Get the module identity for a group based on its first segment's template.
        Returns module_id string or None if no segments have templates."""
        for seg in group.get('segments', []):
            ref = seg.get('template_ref')
            if ref:
                return self.get_template_module_id(ref)
        return None

    def get_compatible_templates(self, group):
        """Get template keys compatible with the group's existing module.
        If group has no template-linked segments, all enabled templates are compatible."""
        group_module = self.get_group_module_id(group)
        
        if group_module is None:
            # No module constraint yet — all enabled templates are valid
            return list(self.enabled_templates.keys())
        
        # Filter to templates with the same module
        compatible = []
        for key in self.enabled_templates:
            if self.get_template_module_id(key) == group_module:
                compatible.append(key)
        return compatible
    
    def refresh_templates(self):
        """Reload enabled templates from disk and rebuild segment dropdowns."""
        self.enabled_templates = self.load_enabled_templates()
        if self.selected_group_idx is not None:
            try:
                self._rebuild_group_details(self.selected_group_idx)
            except Exception:
                pass

    def _derive_module_from_templates(self):
        """Derive the active module and modules_per_string from linked templates.
        Sets self.selected_module and updates modules_per_string_var.
        Returns True if a module was found."""
        from ..models.module import ModuleSpec, ModuleType
        
        # Scan all groups for the first template-linked segment
        for group in self.groups:
            for seg in group.get('segments', []):
                ref = seg.get('template_ref')
                if ref and ref in self.enabled_templates:
                    tdata = self.enabled_templates[ref]
                    module_data = tdata.get('module_spec', {})
                    mps = tdata.get('modules_per_string', 28)
                    
                    try:
                        self.selected_module = ModuleSpec(
                            manufacturer=module_data.get('manufacturer', 'Unknown'),
                            model=module_data.get('model', 'Unknown'),
                            type=ModuleType(module_data.get('type', 'Mono PERC')),
                            length_mm=float(module_data.get('length_mm', 2000)),
                            width_mm=float(module_data.get('width_mm', 1000)),
                            depth_mm=float(module_data.get('depth_mm', 40)),
                            weight_kg=float(module_data.get('weight_kg', 25)),
                            wattage=float(module_data.get('wattage', 400)),
                            vmp=float(module_data.get('vmp', 40)),
                            imp=float(module_data.get('imp', 10)),
                            voc=float(module_data.get('voc', 48)),
                            isc=float(module_data.get('isc', 10.5)),
                            max_system_voltage=float(module_data.get('max_system_voltage', 1500))
                        )
                    except Exception as e:
                        print(f"Error creating ModuleSpec from template: {e}")
                        return False
                    
                    # Update modules_per_string from template
                    if hasattr(self, 'modules_per_string_var'):
                        self.modules_per_string_var.set(str(mps))
                    
                    # Update module info label
                    if hasattr(self, 'module_info_label'):
                        self.module_info_label.config(
                            text=f"{self.selected_module.manufacturer} {self.selected_module.model} ({self.selected_module.wattage}W)  |  Isc: {self.selected_module.isc}A  |  Width: {self.selected_module.width_mm}mm",
                            foreground='black'
                        )
                    
                    self._update_strings_per_inverter()
                    return True
        
        # No templates linked — try fallback from saved estimate data
        if self.current_project and self.estimate_id:
            estimate_data = self.current_project.quick_estimates.get(self.estimate_id, {})
            saved_module_name = estimate_data.get('module_name', '')
            
            if saved_module_name and saved_module_name in self.available_modules:
                self.selected_module = self.available_modules[saved_module_name]
                saved_mps = estimate_data.get('modules_per_string', 28)
                if hasattr(self, 'modules_per_string_var'):
                    self.modules_per_string_var.set(str(saved_mps))
                if hasattr(self, 'module_info_label'):
                    self.module_info_label.config(
                        text=f"(Legacy) {self.selected_module.manufacturer} {self.selected_module.model} ({self.selected_module.wattage}W)  —  link templates to update",
                        foreground='orange'
                    )
                self._update_strings_per_inverter()
                return True
            
            # Try matching by partial name (old format might not match exactly)
            if saved_module_name:
                for display_name, mod in self.available_modules.items():
                    if (mod.manufacturer in saved_module_name and 
                        mod.model in saved_module_name):
                        self.selected_module = mod
                        saved_mps = estimate_data.get('modules_per_string', 28)
                        if hasattr(self, 'modules_per_string_var'):
                            self.modules_per_string_var.set(str(saved_mps))
                        if hasattr(self, 'module_info_label'):
                            self.module_info_label.config(
                                text=f"(Legacy) {mod.manufacturer} {mod.model} ({mod.wattage}W)  —  link templates to update",
                                foreground='orange'
                            )
                        self._update_strings_per_inverter()
                        return True
        
        # Truly no module available
        self.selected_module = None
        if hasattr(self, 'module_info_label'):
            self.module_info_label.config(
                text="No templates linked — link a tracker template to a segment",
                foreground='gray'
            )
        return False

    def _get_group_module_length_m(self, group: dict) -> float:
        """Return the N-S module dimension (m) for the first templated segment in this group.
        Used for GCR <-> row spacing conversion. Returns None if no template is found."""
        for seg in group.get('segments', []):
            ref = seg.get('template_ref')
            if ref and ref in self.enabled_templates:
                tdata = self.enabled_templates[ref]
                ms = tdata.get('module_spec', {})
                orientation = tdata.get('module_orientation', 'Portrait')
                if orientation == 'Portrait':
                    length_mm = ms.get('length_mm', 0)
                else:
                    length_mm = ms.get('width_mm', 0)
                if length_mm > 0:
                    return length_mm / 1000.0
        return None

    def generate_id(self, prefix: str) -> str:
        """Generate a unique ID with a prefix"""
        return f"{prefix}_{uuid.uuid4().hex[:8]}"
            
    def disable_combobox_scroll(self, combobox):
        """Prevent combobox from responding to mouse wheel"""
        def _ignore_scroll(event):
            return "break"
        
        combobox.bind("<MouseWheel>", _ignore_scroll)
        # For Linux compatibility
        combobox.bind("<Button-4>", _ignore_scroll)
        combobox.bind("<Button-5>", _ignore_scroll)

    def update_string_count(self):
        """Legacy method — now handled by _update_group_string_count"""
        pass

    # ==================== Data Management ====================

    def add_group(self) -> int:
        """Add a new group and return its index"""
        group_num = len(self.groups) + 1
        
        # Start unconstrained — let the user pick a template
        default_ref = None
        default_spt = 3
        
        # Default row spacing: inherit from last group, or fall back to 20 ft
        if self.groups:
            default_row_spacing = self.groups[-1].get('row_spacing_ft', 20.0)
        else:
            default_row_spacing = 20.0

        # Inherit strings_per_inv from last group, or use None (means: use global)
        default_spi = self.groups[-1].get('strings_per_inv', None) if self.groups else None

        group = {
            'name': f"Group {group_num}",
            'device_position': 'middle',
            'driveline_angle': 0.0,
            'azimuth': 180,
            'tracker_alignment': 'motor',
            'row_spacing_ft': default_row_spacing,
            'strings_per_inv': default_spi,
            'segments': [
                {'quantity': 1, 'strings_per_tracker': default_spt, 'harness_config': str(default_spt), 'template_ref': default_ref}
            ]
        }
        self.groups.append(group)
        self._refresh_group_listbox()
        
        # Select the new group
        idx = len(self.groups) - 1
        self.group_listbox.selection_clear(0, tk.END)
        self.group_listbox.selection_set(idx)
        self.group_listbox.see(idx)
        self.on_group_select(None)
        
        self._mark_dirty()
        return idx

    def copy_selected_group(self):
        """Copy the currently selected group"""
        sel = self.group_listbox.curselection()
        if not sel:
            return
        
        source = self.groups[sel[0]]
        new_group = copy.deepcopy(source)
        new_group['name'] = f"{source['name']} (Copy)"
        
        # Insert after the selected group
        insert_idx = sel[0] + 1
        self.groups.insert(insert_idx, new_group)
        self._refresh_group_listbox()
        
        self.group_listbox.selection_clear(0, tk.END)
        self.group_listbox.selection_set(insert_idx)
        self.group_listbox.see(insert_idx)
        self.on_group_select(None)
        
        self._mark_dirty()

    def delete_selected_group(self):
        """Delete the currently selected group"""
        sel = self.group_listbox.curselection()
        if not sel:
            return
        
        idx = sel[0]
        del self.groups[idx]
        self._reconcile_locked_allocation()
        self.selected_group_idx = None
        self._refresh_group_listbox(preserve_selection=False)
        
        # Select nearest group
        if self.groups:
            new_idx = min(idx, len(self.groups) - 1)
            self.group_listbox.selection_set(new_idx)
            self.on_group_select(None)
        else:
            self.clear_details_panel()
        
        self._mark_dirty()
    
    def move_group_up(self):
        """Move selected group up"""
        sel = self.group_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        
        idx = sel[0]
        self.groups[idx], self.groups[idx - 1] = self.groups[idx - 1], self.groups[idx]
        self._refresh_group_listbox()
        self.group_listbox.selection_set(idx - 1)
        self.on_group_select(None)
        
        self._mark_dirty()
    
    def move_group_down(self):
        """Move selected group down"""
        sel = self.group_listbox.curselection()
        if not sel or sel[0] >= len(self.groups) - 1:
            return
        
        idx = sel[0]
        self.groups[idx], self.groups[idx + 1] = self.groups[idx + 1], self.groups[idx]
        self._refresh_group_listbox()
        self.group_listbox.selection_set(idx + 1)
        self.on_group_select(None)
        
        self._mark_dirty()
    
    def _refresh_group_listbox(self, preserve_selection=True):
        """Refresh the listbox display from self.groups"""
        sel = self.group_listbox.curselection()
        old_idx = sel[0] if sel else None
        
        self._updating_listbox = True
        self.group_listbox.delete(0, tk.END)
        for group in self.groups:
            total_trackers = sum(seg['quantity'] for seg in group['segments'])
            total_strings = sum(int(seg['quantity'] * seg['strings_per_tracker']) for seg in group['segments'])
            display = f"{group['name']}  ({total_trackers}T / {total_strings}S)"
            self.group_listbox.insert(tk.END, display)
        
        if preserve_selection and old_idx is not None and old_idx < len(self.groups):
            self.group_listbox.selection_set(old_idx)
        self._updating_listbox = False

    def round_whip_length(self, raw_length_ft):
        """Apply 5% waste factor and round up to nearest 10ft increment (min 10ft)"""
        import math
        WASTE_FACTOR = 1.05
        INCREMENT = 10
        length_with_waste = raw_length_ft * WASTE_FACTOR
        rounded = INCREMENT * math.ceil(length_with_waste / INCREMENT)
        return max(10, int(rounded))
    
    def load_combiner_library(self):
        """Load combiner box library from JSON file"""
        try:
            lib_path = Path('data/combiner_box_library.json')
            if lib_path.exists():
                with open(lib_path, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading combiner box library: {e}")
        return {}

    def get_fuse_holder_category(self, fuse_current_amps):
        """Determine fuse holder rating category from required fuse current"""
        if fuse_current_amps <= 20:
            return "20A and Below"
        elif fuse_current_amps <= 32:
            return "25-32A"
        else:
            return "32A and Above"

    def find_combiner_box(self, strings_per_cb, breaker_size, fuse_holder_rating):
        """Find a matching combiner box from the library.
        
        Returns the part_number and description, or None if no match.
        Prefers smallest max_inputs that fits, with whips (whip_length > 0).
        """
        combiner_library = self.load_combiner_library()
        
        candidates = []
        for part_num, cb_data in combiner_library.items():
            if (cb_data.get('max_inputs', 0) >= strings_per_cb and
                cb_data.get('breaker_size', 0) == breaker_size and
                cb_data.get('fuse_holder_rating', '') == fuse_holder_rating):
                candidates.append(cb_data)
        
        if not candidates:
            return None
        
        # Sort by max_inputs (prefer smallest that fits), then prefer with whips
        candidates.sort(key=lambda c: (c['max_inputs'], -c.get('whip_length_ft', 0)))
        return candidates[0]

    def calculate_cb_whip_distances(self, total_rows, num_combiners, row_spacing_ft):
        """Calculate whip distances for each row based on evenly-spaced CB placement.
        
        Returns a list of (distance_ft, cb_index) for each row.
        
        CB placement logic:
        - Rows are divided into groups, first CB(s) get extra rows if uneven
        - Each CB is placed at the center of its group
        - Each row connects to its nearest CB
        """
        if total_rows <= 0 or num_combiners <= 0:
            return []
        
        # Divide rows into groups for each CB
        base_group_size = total_rows // num_combiners
        extra_rows = total_rows % num_combiners
        
        # Build groups - first CBs get the extra rows
        groups = []
        row_start = 1
        for cb_idx in range(num_combiners):
            group_size = base_group_size + (1 if cb_idx < extra_rows else 0)
            row_end = row_start + group_size - 1
            # CB position is center of its group
            cb_position = (row_start + row_end) / 2.0
            groups.append({
                'cb_idx': cb_idx,
                'row_start': row_start,
                'row_end': row_end,
                'cb_position': cb_position
            })
            row_start = row_end + 1
        
        # Calculate distance from each row to its assigned CB
        whip_distances = []
        for group_num in range(1, total_rows + 1):
            # Find which group this row belongs to
            for group in groups:
                if group['row_start'] <= group_num <= group['row_end']:
                    distance_ft = abs(group_num - group['cb_position']) * row_spacing_ft
                    whip_distances.append((distance_ft, group['cb_idx']))
                    break
        
        return whip_distances

    def calculate_whip_distances_from_positions(self, allocation_result, topology, num_devices, row_spacing_ft=None):
        """Calculate whip distances using device positions derived from allocation.
        
        For Distributed String and Centralized String: uses allocation result
        to map each inverter's trackers to a device, computing real E-W + N-S distance.
        
        For Central Inverter: falls back to the abstract even-spacing method.
        
        Returns a flat list of distance_ft values, one per tracker.
        """
        if not allocation_result:
            total_trackers = sum(
                sum(seg['quantity'] for seg in group['segments'])
                for group in self.groups
            )
            fallback_spacing = self.groups[0].get('row_spacing_ft', 20.0) if self.groups else 20.0
            old_distances = self.calculate_cb_whip_distances(
                total_trackers, num_devices, fallback_spacing
            )
            # Build flat spt list matching tracker order
            spt_list = []
            for group in self.groups:
                for seg in group['segments']:
                    for _ in range(seg['quantity']):
                        spt_list.append(seg['strings_per_tracker'])
            return [(d[0], spt_list[i] if i < len(spt_list) else 0, i, -1) for i, d in enumerate(old_distances)]
        
        # Build world X for every global tracker index
        # Uses saved group positions or auto-layout, same as site preview
        tracker_world_x = []
        tracker_world_y = []
        tracker_group = []
        auto_x_cursor = 0.0
        
        for grp_idx, group in enumerate(self.groups):
            saved_x = group.get('position_x')
            saved_y = group.get('position_y')
            if saved_x is not None:
                group_x = saved_x
            else:
                group_x = auto_x_cursor
            group_y = saved_y if saved_y is not None else 0.0
            
            driveline_angle_deg = group.get('driveline_angle', 0.0)
            driveline_tan = math.tan(math.radians(driveline_angle_deg)) if driveline_angle_deg != 0 else 0.0
            
            grp_row_spacing = group.get('row_spacing_ft', 20.0)
            local_idx = 0
            for seg in group['segments']:
                for _ in range(seg['quantity']):
                    local_x_offset = local_idx * grp_row_spacing
                    tracker_world_x.append(group_x + local_x_offset)
                    tracker_world_y.append(group_y + local_x_offset * driveline_tan)
                    tracker_group.append(grp_idx)
                    local_idx += 1
            
            # Advance auto cursor
            group_tracker_count = sum(seg['quantity'] for seg in group['segments'])
            auto_x_cursor += group_tracker_count * grp_row_spacing + grp_row_spacing * 2
        
        # Compute device world X and metadata for each inverter
        inverters = allocation_result.get('inverters', [])
        device_info = []  # list of (world_x, device_position) per inverter
        
        for inv_idx, inv in enumerate(inverters):
            harness_map = inv.get('harness_map', [])
            if not harness_map:
                device_info.append(None)
                continue
            
            # Find majority group
            group_counts = {}
            for entry in harness_map:
                tidx = entry['tracker_idx']
                if tidx < len(tracker_group):
                    grp = tracker_group[tidx]
                    group_counts[grp] = group_counts.get(grp, 0) + 1
            
            if not group_counts:
                device_info.append(None)
                continue
            
            primary_grp = max(group_counts, key=group_counts.get)
            group_source = self.groups[primary_grp] if primary_grp < len(self.groups) else {}
            device_position = group_source.get('device_position', 'middle')
            
            # Device X = average of its trackers' world X positions in the primary group
            inv_tracker_xs = []
            for entry in harness_map:
                tidx = entry['tracker_idx']
                if tidx < len(tracker_world_x) and tracker_group[tidx] == primary_grp:
                    inv_tracker_xs.append(tracker_world_x[tidx])
            
            if inv_tracker_xs:
                device_x = (min(inv_tracker_xs) + max(inv_tracker_xs)) / 2.0
            else:
                device_x = 0

            # For 'middle' placement: snap device_x to the nearest row gap.
            # The midpoint formula lands in the gap for even tracker counts but on a tracker
            # left-edge for odd counts (1, 3, 5...) — in the odd case bias half a pitch east.
            if device_position == 'middle' and inv_tracker_xs:
                grp_pitch = group_source.get('row_spacing_ft', 20.0)
                if grp_pitch > 0:
                    center_offset = (max(inv_tracker_xs) - min(inv_tracker_xs)) / 2.0
                    frac = (center_offset / grp_pitch) % 1
                    if frac < 0.01:  # integer multiple = on a tracker, not in a gap
                        device_x += grp_pitch / 2

            # Compute device Y from average of its trackers' Y positions + device_position offset
            inv_tracker_ys = []
            for entry in harness_map:
                tidx = entry['tracker_idx']
                if tidx < len(tracker_world_y) and tracker_group[tidx] == primary_grp:
                    inv_tracker_ys.append(tracker_world_y[tidx])
            device_y = (min(inv_tracker_ys) + max(inv_tracker_ys)) / 2.0 if inv_tracker_ys else 0.0
            
            device_info.append((device_x, device_y, device_position))
        
        def get_ns_base_offset(device_position):
            """Fixed N-S offset from device_position setting (north/south/middle)."""
            if device_position == 'middle':
                return 0.0
            return 6.5  # 5ft offset + half device height
        
        # Compute distance for each tracker to its assigned device
        # Deduplicate: each physical tracker generates whips once, even if split across inverters
        whip_distances = []
        seen_trackers = set()
        
        for inv_idx, inv in enumerate(inverters):
            if inv_idx >= len(device_info) or device_info[inv_idx] is None:
                continue
            
            dev_x, dev_y, dev_position = device_info[inv_idx]
            ns_base = get_ns_base_offset(dev_position)
            
            for entry in inv['harness_map']:
                tidx = entry['tracker_idx']
                is_split = entry.get('is_split', False)
                
                # Allow split tracker portions through (they appear in multiple inverters)
                if tidx in seen_trackers and not is_split:
                    continue
                seen_trackers.add(tidx)
                
                if tidx >= len(tracker_world_x):
                    continue
                
                ew_distance = abs(tracker_world_x[tidx] - dev_x)
                # N-S inter-row distance (tracker to device) — stored for extender adjustment
                ns_inter_row = abs(tracker_world_y[tidx] - dev_y)
                
                # Whip = E-W only; N-S distance goes to extender
                total_distance = ew_distance
                
                # Store SIGNED N-S offset: positive = CB is south of tracker
                if not hasattr(self, '_tracker_ns_to_device'):
                    self._tracker_ns_to_device = {}
                signed_ns = dev_y - tracker_world_y[tidx]
                self._tracker_ns_to_device[(tidx, inv_idx)] = signed_ns
                
                spt = entry.get('strings_per_tracker', 0)

                whip_distances.append((total_distance, spt, tidx, inv_idx))
        
        return whip_distances
    
    def calculate_routed_feeder_distances(self, allocation_result, topology, row_spacing_ft):
        """Calculate routed feeder distances from each device to its assigned pad.
        
        Returns a dict with:
            'feeder_distances': list of (device_label, distance_ft) tuples
            'feeder_total_ft': total feeder cable
            'feeder_count': number of feeder runs
            'homerun_distances': list of (pad_label, distance_ft) for AC homeruns (stub)
        """
        result = {
            'feeder_distances': [],
            'feeder_total_ft': 0,
            'feeder_count': 0,
        }
        
        if not self.pads or not allocation_result:
            return result
        
        inverters = allocation_result.get('inverters', [])
        if not inverters:
            return result
        
        # Build device world X positions (same as whip calc)
        device_world_x = []
        device_group = []
        auto_x_cursor = 0.0
        
        for grp_idx, group in enumerate(self.groups):
            saved_x = group.get('position_x')
            group_x = saved_x if saved_x is not None else auto_x_cursor
            
            local_idx = 0
            tracker_info_local = []
            for seg in group['segments']:
                for _ in range(seg['quantity']):
                    tracker_info_local.append((grp_idx, local_idx))
                    local_idx += 1
            
            group_tracker_count = sum(seg['quantity'] for seg in group['segments'])
            auto_x_cursor += group_tracker_count * row_spacing_ft + row_spacing_ft * 2
        
        # Recompute device centers (same logic as whip calc and site preview)
        tracker_world_x = []
        tracker_grp = []
        group_x_map = {}
        auto_x_cursor = 0.0

        for grp_idx, group in enumerate(self.groups):
            saved_x = group.get('position_x')
            group_x = saved_x if saved_x is not None else auto_x_cursor
            group_x_map[grp_idx] = group_x

            local_idx = 0
            for seg in group['segments']:
                for _ in range(seg['quantity']):
                    tracker_world_x.append(group_x + local_idx * row_spacing_ft)
                    tracker_grp.append(grp_idx)
                    local_idx += 1

            group_tracker_count = sum(seg['quantity'] for seg in group['segments'])
            auto_x_cursor += group_tracker_count * row_spacing_ft + row_spacing_ft * 2
        
        # Compute device positions (center X, Y based on device_position)
        device_positions = []
        for inv_idx, inv in enumerate(inverters):
            harness_map = inv.get('harness_map', [])
            if not harness_map:
                device_positions.append(None)
                continue
            
            group_counts = {}
            for entry in harness_map:
                tidx = entry['tracker_idx']
                if tidx < len(tracker_grp):
                    g = tracker_grp[tidx]
                    group_counts[g] = group_counts.get(g, 0) + 1
            
            if not group_counts:
                device_positions.append(None)
                continue
            
            primary_grp = max(group_counts, key=group_counts.get)
            group_source = self.groups[primary_grp] if primary_grp < len(self.groups) else {}
            device_position = group_source.get('device_position', 'middle')
            
            inv_xs = [tracker_world_x[e['tracker_idx']] for e in harness_map 
                      if e['tracker_idx'] < len(tracker_world_x) and tracker_grp[e['tracker_idx']] == primary_grp]
            
            dev_x = (min(inv_xs) + max(inv_xs)) / 2.0 if inv_xs else 0

            # For 'middle' placement: snap dev_x to the nearest row gap (same logic as whip calc).
            if device_position == 'middle' and inv_xs:
                grp_pitch = group_source.get('row_spacing_ft', row_spacing_ft)
                if grp_pitch > 0:
                    center_offset = (max(inv_xs) - min(inv_xs)) / 2.0
                    frac = (center_offset / grp_pitch) % 1
                    if frac < 0.01:
                        dev_x += grp_pitch / 2

            # Approximate device Y from group position
            saved_y = group_source.get('position_y', 0)
            group_y = saved_y if saved_y is not None else 0
            
            # Get tracker length for Y offset calculation
            first_ref = None
            for seg in group_source.get('segments', []):
                ref = seg.get('template_ref')
                if ref and ref in self.enabled_templates:
                    first_ref = ref
                    break
            
            tracker_length_ft = 180.0  # fallback
            if first_ref:
                dims = self._get_estimate_tracker_dims_ft(first_ref)
                if dims:
                    tracker_length_ft = dims[1]
            
            if device_position == 'north':
                dev_y = group_y - 5.0
            elif device_position == 'south':
                dev_y = group_y + tracker_length_ft + 5.0
            else:
                dev_y = group_y + tracker_length_ft / 2.0
            
            device_positions.append((dev_x, dev_y, primary_grp))
        
        # Build device -> pad lookup
        device_to_pad = {}
        for pad_idx, pad in enumerate(self.pads):
            for dev_idx in pad.get('assigned_devices', []):
                device_to_pad[dev_idx] = pad_idx
        
        # Compute routed feeder distance (row-direction projection) from each device to its pad
        for dev_idx, dev_pos in enumerate(device_positions):
            if dev_pos is None:
                continue

            pad_idx = device_to_pad.get(dev_idx)
            if pad_idx is None or pad_idx >= len(self.pads):
                continue

            pad = self.pads[pad_idx]
            pad_cx = pad['x'] + pad.get('width_ft', 10.0) / 2
            pad_cy = pad['y'] + pad.get('height_ft', 8.0) / 2

            dev_x, dev_y, primary_grp = dev_pos
            group_source = self.groups[primary_grp] if primary_grp < len(self.groups) else {}

            azimuth = group_source.get('azimuth', 180)
            rotation_deg = azimuth - 180
            driveline_angle_deg = group_source.get('driveline_angle', 0.0)

            # Rotation center: horizontal midpoint of group, vertical midpoint of tracker length
            gx = group_x_map.get(primary_grp, 0)
            gy = group_source.get('position_y', 0) or 0
            grp_count = sum(seg['quantity'] for seg in group_source.get('segments', []))
            first_ref_r = None
            for seg in group_source.get('segments', []):
                ref = seg.get('template_ref')
                if ref and ref in self.enabled_templates:
                    first_ref_r = ref
                    break
            tlen_r = 180.0
            twid_r = 0.0
            if first_ref_r:
                dims_r = self._get_estimate_tracker_dims_ft(first_ref_r)
                if dims_r:
                    twid_r, tlen_r = dims_r[0], dims_r[1]
            rot_cx = gx + (twid_r + max(0, grp_count - 1) * row_spacing_ft) / 2
            rot_cy = gy + tlen_r / 2

            # Rotate device center by group azimuth
            if rotation_deg != 0:
                rad = math.radians(rotation_deg)
                cos_r = math.cos(rad)
                sin_r = math.sin(rad)
                dx = dev_x - rot_cx
                dy = dev_y - rot_cy
                dev_x = rot_cx + dx * cos_r - dy * sin_r
                dev_y = rot_cy + dx * sin_r + dy * cos_r

            # Row direction vector (driveline + azimuth rotation)
            driveline_tan = math.tan(math.radians(driveline_angle_deg)) if driveline_angle_deg != 0 else 0.0
            mag = math.sqrt(1.0 + driveline_tan ** 2)
            rdx_u, rdy_u = 1.0 / mag, driveline_tan / mag
            if rotation_deg != 0:
                cos_r = math.cos(math.radians(rotation_deg))
                sin_r = math.sin(math.radians(rotation_deg))
                row_dx = rdx_u * cos_r - rdy_u * sin_r
                row_dy = rdx_u * sin_r + rdy_u * cos_r
            else:
                row_dx, row_dy = rdx_u, rdy_u

            # Corner via projection onto row direction line through device
            t = (pad_cx - dev_x) * row_dx + (pad_cy - dev_y) * row_dy
            corner_x = dev_x + t * row_dx
            corner_y = dev_y + t * row_dy
            leg2 = math.sqrt((pad_cx - corner_x) ** 2 + (pad_cy - corner_y) ** 2)
            routed = abs(t) + leg2

            label = f"Dev-{dev_idx+1:02d}"
            result['feeder_distances'].append((dev_idx, label, routed))
            result['feeder_total_ft'] += routed
            result['feeder_count'] += 1
        
        return result
    
    def _build_wire_sizing_frame(self, parent):
        """Build the Wire Sizing widget to the right of Global Settings."""
        from src.utils.cable_sizing import (
            CABLE_SIZE_ORDER, CABLE_SIZE_ORDER_EXTENDED, 
            get_available_sizes
        )
        
        ws_frame = ttk.LabelFrame(parent, text="Wire Sizing", padding="5")
        ws_frame.pack(side='left', fill='y', padx=(10, 0))
        self._ws_frame = ws_frame
        
        # Row 0: Temp rating, Material toggle, Reset button
        controls_row = ttk.Frame(ws_frame)
        controls_row.pack(fill='x', pady=(0, 5))
        
        ttk.Label(controls_row, text="Temp:").pack(side='left', padx=(0, 2))
        self._ws_temp_var = tk.StringVar(value=self.wire_sizing.get('temp_rating', '90C'))
        temp_combo = ttk.Combobox(
            controls_row, textvariable=self._ws_temp_var,
            values=['60C', '75C', '90C'], state='readonly', width=4
        )
        temp_combo.pack(side='left', padx=(0, 8))
        self.disable_combobox_scroll(temp_combo)
        
        ttk.Label(controls_row, text="Feeder:").pack(side='left', padx=(0, 2))
        self._ws_material_var = tk.StringVar(value=self.wire_sizing.get('feeder_material', 'aluminum'))
        material_combo = ttk.Combobox(
            controls_row, textvariable=self._ws_material_var,
            values=['aluminum', 'copper'], state='readonly', width=9
        )
        material_combo.pack(side='left', padx=(0, 8))
        self.disable_combobox_scroll(material_combo)
        
        ttk.Button(controls_row, text="Reset", width=5,
                    command=self._reset_wire_sizing_to_recommended).pack(side='left')
        
        # Column headers
        header_row = ttk.Frame(ws_frame)
        header_row.pack(fill='x', pady=(0, 2))
        
        header_font = ('TkDefaultFont', 8, 'bold')
        ttk.Label(header_row, text="Strings", font=header_font, width=7).pack(side='left')
        ttk.Label(header_row, text="Harness", font=header_font, width=10).pack(side='left', padx=2)
        ttk.Label(header_row, text="Extender", font=header_font, width=10).pack(side='left', padx=2)
        ttk.Label(header_row, text="Whip", font=header_font, width=10).pack(side='left', padx=2)
        
        # Dynamic rows container — cleared and rebuilt by refresh_wire_sizing_table
        self._ws_rows_frame = ttk.Frame(ws_frame)
        self._ws_rows_frame.pack(fill='x')
        
        # Storage for combo vars and override tracking
        self._ws_lv_combos = {}    # {(string_count, cable_type): tk.StringVar}
        self._ws_feeder_var = tk.StringVar(value=self.wire_sizing.get('dc_feeder', ''))
        self._ws_homerun_var = tk.StringVar(value=self.wire_sizing.get('ac_homerun', ''))
        self._ws_dc_feeder_blanket_var = tk.BooleanVar(value=self.wire_sizing.get('dc_feeder_blanket_enabled', False))
        self._ws_ac_homerun_blanket_var = tk.BooleanVar(value=self.wire_sizing.get('ac_homerun_blanket_enabled', False))
        
        # Available LV cable sizes (copper AWG only — pre-made assemblies)
        self._ws_lv_sizes = CABLE_SIZE_ORDER  # ['10 AWG' through '4/0 AWG']
        
        # Traces for temp and material changes
        self._ws_temp_var.trace_add('write', lambda *a: self._on_wire_sizing_setting_changed())
        self._ws_material_var.trace_add('write', lambda *a: self._on_wire_sizing_setting_changed())
        
    def _on_wire_sizing_setting_changed(self):
        """Handle change to temp rating or feeder material — recalc recommendations."""
        self.wire_sizing['temp_rating'] = self._ws_temp_var.get()
        self.wire_sizing['feeder_material'] = self._ws_material_var.get()
        self._reset_wire_sizing_to_recommended()
        self._mark_dirty()

    def refresh_wire_sizing_table(self):
        """Refresh the wire sizing table rows based on current harness configs."""
        if not hasattr(self, '_ws_rows_frame'):
            return
        
        from src.utils.cable_sizing import CABLE_SIZE_ORDER, get_available_sizes
        
        # Clear existing rows
        for widget in self._ws_rows_frame.winfo_children():
            widget.destroy()
        self._ws_lv_combos.clear()
        
        active_counts = self._collect_active_string_counts()
        topology = self.topology_var.get() if hasattr(self, 'topology_var') else 'Distributed String'
        by_sc = self.wire_sizing.get('by_string_count', {})
        
        # Build LV rows (harness/extender/whip per string count)
        for sc in active_counts:
            row_frame = ttk.Frame(self._ws_rows_frame)
            row_frame.pack(fill='x', pady=1)
            
            ttk.Label(row_frame, text=f"{sc}-str", width=7).pack(side='left')
            
            # Get current sizes from wire_sizing dict (int or str key)
            entry = by_sc.get(sc) or by_sc.get(str(sc)) or {}
            
            for cable_type in ('harness', 'extender', 'whip'):
                current_val = entry.get(cable_type, '10 AWG')
                var = tk.StringVar(value=current_val)
                combo = ttk.Combobox(
                    row_frame, textvariable=var,
                    values=CABLE_SIZE_ORDER, state='readonly', width=8
                )
                combo.pack(side='left', padx=2)
                self.disable_combobox_scroll(combo)
                
                # Store reference
                self._ws_lv_combos[(sc, cable_type)] = var
                
                # Bind change handler with closure
                var.trace_add('write', lambda *a, s=sc, ct=cable_type: self._on_lv_size_changed(s, ct))
        
        # Separator before feeder rows
        if active_counts and (topology != 'Distributed String' or True):
            sep = ttk.Separator(self._ws_rows_frame, orient='horizontal')
            sep.pack(fill='x', pady=3)
        
        # DC Feeder row (only for Centralized String / Central Inverter)
        if topology in ('Centralized String', 'Central Inverter'):
            feeder_row = ttk.Frame(self._ws_rows_frame)
            feeder_row.pack(fill='x', pady=1)

            ttk.Label(feeder_row, text="DC Fdr", width=7, foreground='gray').pack(side='left')
            material = self.wire_sizing.get('feeder_material', 'aluminum')
            feeder_sizes = get_available_sizes(material)
            current_feeder = self.wire_sizing.get('dc_feeder', '')
            self._ws_feeder_var.set(current_feeder)
            dc_blanket_on = self.wire_sizing.get('dc_feeder_blanket_enabled', False)
            self._ws_dc_feeder_blanket_var.set(dc_blanket_on)
            feeder_combo = ttk.Combobox(
                feeder_row, textvariable=self._ws_feeder_var,
                values=feeder_sizes, state='readonly' if dc_blanket_on else 'disabled', width=10
            )
            feeder_combo.pack(side='left', padx=2)
            self.disable_combobox_scroll(feeder_combo)

            if dc_blanket_on:
                self._ws_feeder_var.trace_add('write', lambda *a: self._on_feeder_size_changed('dc_feeder'))
                ttk.Label(feeder_row, text="(applies to all)", foreground='gray',
                          font=('TkDefaultFont', 7)).pack(side='left', padx=(2, 0))
            else:
                ttk.Label(feeder_row, text="← per device", foreground='gray',
                          font=('TkDefaultFont', 7)).pack(side='left', padx=(2, 0))
            ttk.Checkbutton(feeder_row, text="Apply to all",
                            variable=self._ws_dc_feeder_blanket_var,
                            command=lambda: self._on_feeder_blanket_toggled('dc_feeder')
                            ).pack(side='left', padx=(6, 0))
        
        # AC Homerun row (all topologies)
        homerun_row = ttk.Frame(self._ws_rows_frame)
        homerun_row.pack(fill='x', pady=1)

        is_distributed = (topology == 'Distributed String')
        label_color = 'gray' if is_distributed else 'black'

        ttk.Label(homerun_row, text="AC HR", width=7, foreground=label_color).pack(side='left')
        material = self.wire_sizing.get('feeder_material', 'aluminum')
        homerun_sizes = get_available_sizes(material)
        current_homerun = self.wire_sizing.get('ac_homerun', '')
        self._ws_homerun_var.set(current_homerun)

        if is_distributed:
            ac_blanket_on = self.wire_sizing.get('ac_homerun_blanket_enabled', False)
            self._ws_ac_homerun_blanket_var.set(ac_blanket_on)
            homerun_combo = ttk.Combobox(
                homerun_row, textvariable=self._ws_homerun_var,
                values=homerun_sizes, state='readonly' if ac_blanket_on else 'disabled', width=10
            )
            homerun_combo.pack(side='left', padx=2)
            self.disable_combobox_scroll(homerun_combo)
            if ac_blanket_on:
                self._ws_homerun_var.trace_add('write', lambda *a: self._on_feeder_size_changed('ac_homerun'))
                ttk.Label(homerun_row, text="(applies to all)", foreground='gray',
                          font=('TkDefaultFont', 7)).pack(side='left', padx=(2, 0))
            else:
                ttk.Label(homerun_row, text="← per device", foreground='gray',
                          font=('TkDefaultFont', 7)).pack(side='left', padx=(2, 0))
            ttk.Checkbutton(homerun_row, text="Apply to all",
                            variable=self._ws_ac_homerun_blanket_var,
                            command=lambda: self._on_feeder_blanket_toggled('ac_homerun')
                            ).pack(side='left', padx=(6, 0))
        else:
            homerun_combo = ttk.Combobox(
                homerun_row, textvariable=self._ws_homerun_var,
                values=homerun_sizes, state='readonly', width=10
            )
            homerun_combo.pack(side='left', padx=2)
            self.disable_combobox_scroll(homerun_combo)
            self._ws_homerun_var.trace_add('write', lambda *a: self._on_feeder_size_changed('ac_homerun'))
    
    def _reset_wire_sizing_to_recommended(self):
        """Reset all wire sizes to calculated recommendations."""
        from src.utils.cable_sizing import (
            recommend_lv_cable_sizes, recommend_dc_feeder_size, 
            recommend_ac_homerun_size
        )
        
        temp = self._ws_temp_var.get() if hasattr(self, '_ws_temp_var') else '90C'
        material = self._ws_material_var.get() if hasattr(self, '_ws_material_var') else 'aluminum'
        
        self.wire_sizing['temp_rating'] = temp
        self.wire_sizing['feeder_material'] = material
        self.wire_sizing['user_overrides'] = {}  # Clear all overrides
        
        # Get module Isc
        module_isc = self.selected_module.isc if self.selected_module else 13.0
        
        # Recommend LV sizes for each active string count
        active_counts = self._collect_active_string_counts()
        by_sc = {}
        for sc in active_counts:
            sizes = recommend_lv_cable_sizes(sc, module_isc, nec_factor=1.56, temp_rating=temp)
            by_sc[sc] = sizes
        self.wire_sizing['by_string_count'] = by_sc
        
        # Recommend DC feeder
        topology = self.topology_var.get() if hasattr(self, 'topology_var') else 'Distributed String'
        if topology in ('Centralized String', 'Central Inverter'):
            breaker = self._get_float_var(self.breaker_size_var, 400.0)
            self.wire_sizing['dc_feeder'] = recommend_dc_feeder_size(breaker, material, temp)
        else:
            self.wire_sizing['dc_feeder'] = ''
        
        # Recommend AC homerun
        if self.selected_inverter and hasattr(self.selected_inverter, 'max_ac_current'):
            max_ac = self.selected_inverter.max_ac_current
            self.wire_sizing['ac_homerun'] = recommend_ac_homerun_size(max_ac, material, temp)
        else:
            self.wire_sizing['ac_homerun'] = ''
        
        # Rebuild the table UI with new values
        self.refresh_wire_sizing_table()
        self._mark_dirty()

    def _on_lv_size_changed(self, string_count, cable_type):
        """Handle user changing a harness/extender/whip size in the wire sizing table."""
        from src.utils.cable_sizing import get_cable_size_index
        
        key = (string_count, cable_type)
        if key not in self._ws_lv_combos:
            return
        
        new_size = self._ws_lv_combos[key].get()
        
        # Ensure by_string_count dict has this entry
        by_sc = self.wire_sizing.setdefault('by_string_count', {})
        entry = by_sc.get(string_count) or by_sc.get(str(string_count))
        if entry is None:
            entry = {'harness': '10 AWG', 'extender': '10 AWG', 'whip': '10 AWG'}
            by_sc[string_count] = entry
        
        entry[cable_type] = new_size
        
        # Enforce floor constraint: extender and whip >= harness
        harness_idx = get_cable_size_index(entry['harness'])
        
        if cable_type == 'harness':
            # If harness got bigger, bump extender and whip up if they're smaller
            for dep_type in ('extender', 'whip'):
                dep_idx = get_cable_size_index(entry[dep_type])
                if dep_idx < harness_idx:
                    entry[dep_type] = entry['harness']
                    dep_key = (string_count, dep_type)
                    if dep_key in self._ws_lv_combos:
                        self._ws_lv_combos[dep_key].set(entry['harness'])
        else:
            # If extender or whip was set smaller than harness, snap it back
            my_idx = get_cable_size_index(new_size)
            if my_idx < harness_idx:
                entry[cable_type] = entry['harness']
                self._ws_lv_combos[key].set(entry['harness'])
        
        # Track as user override
        overrides = self.wire_sizing.setdefault('user_overrides', {})
        overrides[f"{string_count}_{cable_type}"] = True

        self._mark_dirty()
        self._propagate_wire_sizing_to_devices()

    def _on_feeder_size_changed(self, feeder_type):
        """Handle user changing DC feeder or AC homerun size."""
        if feeder_type == 'dc_feeder':
            self.wire_sizing['dc_feeder'] = self._ws_feeder_var.get()
            if self.wire_sizing.get('dc_feeder_blanket_enabled', False):
                self.device_feeder_sizes.clear()
        elif feeder_type == 'ac_homerun':
            self.wire_sizing['ac_homerun'] = self._ws_homerun_var.get()
            if self.wire_sizing.get('ac_homerun_blanket_enabled', False):
                self.device_feeder_sizes.clear()

        overrides = self.wire_sizing.setdefault('user_overrides', {})
        overrides[feeder_type] = True

        self._mark_dirty()
        self._propagate_wire_sizing_to_devices()

    def _on_feeder_blanket_toggled(self, feeder_type):
        """Handle 'Apply to all' toggle for DC feeder or AC homerun."""
        if feeder_type == 'dc_feeder':
            blanket_on = self._ws_dc_feeder_blanket_var.get()
            self.wire_sizing['dc_feeder_blanket_enabled'] = blanket_on
        else:  # ac_homerun
            blanket_on = self._ws_ac_homerun_blanket_var.get()
            self.wire_sizing['ac_homerun_blanket_enabled'] = blanket_on

        if blanket_on:
            self.device_feeder_sizes.clear()

        if self._site_preview_window is not None:
            try:
                if self._site_preview_window.winfo_exists():
                    self._site_preview_window.draw()
            except Exception:
                pass

        self._mark_dirty()
        self._schedule_autosave()
        self.refresh_wire_sizing_table()

    def _propagate_wire_sizing_to_devices(self):
        """Propagate updated wire sizes from the Wire Sizing table to downstream surfaces.

        Surfaces updated:
          - last_combiner_assignments: wire_gauge per connection (in-place)
          - device_configurator: actual_cable_size for non-manually-set connections
          - SitePreviewWindow: canvas redrawn so the device info panel reflects new gauges
        """
        if self.allocation_locked:
            return
        if not getattr(self, 'last_combiner_assignments', None):
            return

        # Update wire_gauge in last_combiner_assignments based on current wire_sizing
        for cb in self.last_combiner_assignments:
            for conn in cb.get('connections', []):
                conn['wire_gauge'] = self.get_wire_size_for('whip', conn['num_strings'])

        # Push to Device Configurator (QE-mode only, skips manual overrides)
        if hasattr(self, 'main_app') and hasattr(self.main_app, 'device_configurator'):
            dc = self.main_app.device_configurator
            if dc.data_source == 'quick_estimate' and dc.combiner_configs:
                dc.update_cable_sizes_from_qe(self.last_combiner_assignments)

        # Refresh Site Preview canvas if it is currently open
        from .site_preview import SitePreviewWindow
        for child in self.winfo_children():
            if isinstance(child, SitePreviewWindow):
                try:
                    child.refresh_wire_gauges()
                except Exception:
                    pass

    def _refresh_wire_sizing_for_segments(self):
        """Refresh wire sizing when segments/harness configs change.
        Preserves user overrides, adds new string counts with recommendations,
        removes string counts no longer in use."""
        from src.utils.cable_sizing import recommend_lv_cable_sizes
        
        if not hasattr(self, '_ws_rows_frame'):
            return
        
        temp = self.wire_sizing.get('temp_rating', '90C')
        module_isc = self.selected_module.isc if self.selected_module else 13.0
        overrides = self.wire_sizing.get('user_overrides', {})
        
        active_counts = self._collect_active_string_counts()
        by_sc = self.wire_sizing.get('by_string_count', {})

        # Add recommendations for any new string counts
        _table_changed = False
        for sc in active_counts:
            if sc not in by_sc and str(sc) not in by_sc:
                sizes = recommend_lv_cable_sizes(sc, module_isc, nec_factor=1.56, temp_rating=temp)
                by_sc[sc] = sizes
                _table_changed = True

        # Remove string counts no longer in use, clearing any overrides for them
        keys_to_remove = []
        for sc_key in list(by_sc.keys()):
            sc_int = int(sc_key) if isinstance(sc_key, str) else sc_key
            if sc_int not in active_counts:
                keys_to_remove.append((sc_key, sc_int))

        if keys_to_remove:
            _table_changed = True
        for key, sc_int in keys_to_remove:
            del by_sc[key]
            for ct in ('harness', 'extender', 'whip'):
                overrides.pop(f"{sc_int}_{ct}", None)

        self.wire_sizing['by_string_count'] = by_sc
        if _table_changed:
            self.refresh_wire_sizing_table()
    
    def _on_topology_changed_wire_sizing(self):
        """Handle topology change — show/hide DC feeder row, recalc if needed."""
        from src.utils.cable_sizing import recommend_dc_feeder_size
        
        topology = self.topology_var.get()
        material = self.wire_sizing.get('feeder_material', 'aluminum')
        temp = self.wire_sizing.get('temp_rating', '90C')
        overrides = self.wire_sizing.get('user_overrides', {})
        
        if topology in ('Centralized String', 'Central Inverter'):
            # Generate DC feeder recommendation if not already overridden
            if not overrides.get('dc_feeder') and not self.wire_sizing.get('dc_feeder'):
                try:
                    breaker = float(self.breaker_size_var.get())
                except (ValueError, AttributeError):
                    breaker = 400.0
                self.wire_sizing['dc_feeder'] = recommend_dc_feeder_size(breaker, material, temp)
        else:
            self.wire_sizing['dc_feeder'] = ''
        
        self.refresh_wire_sizing_table()
    
    def _on_breaker_changed_wire_sizing(self):
        """Handle breaker size change — update DC feeder recommendation."""
        from src.utils.cable_sizing import recommend_dc_feeder_size
        
        overrides = self.wire_sizing.get('user_overrides', {})
        if overrides.get('dc_feeder'):
            return  # User has overridden, don't auto-update
        
        topology = self.topology_var.get()
        if topology not in ('Centralized String', 'Central Inverter'):
            return
        
        material = self.wire_sizing.get('feeder_material', 'aluminum')
        temp = self.wire_sizing.get('temp_rating', '90C')
        
        breaker = self._get_float_var(self.breaker_size_var, 400.0)
        
        self.wire_sizing['dc_feeder'] = recommend_dc_feeder_size(breaker, material, temp)
        self.refresh_wire_sizing_table()
    
    def _on_inverter_changed_wire_sizing(self):
        """Handle inverter change — update AC homerun recommendation."""
        from src.utils.cable_sizing import recommend_ac_homerun_size
        
        overrides = self.wire_sizing.get('user_overrides', {})
        if overrides.get('ac_homerun'):
            return  # User has overridden, don't auto-update
        
        material = self.wire_sizing.get('feeder_material', 'aluminum')
        temp = self.wire_sizing.get('temp_rating', '90C')
        
        if self.selected_inverter and hasattr(self.selected_inverter, 'max_ac_current'):
            max_ac = self.selected_inverter.max_ac_current
            self.wire_sizing['ac_homerun'] = recommend_ac_homerun_size(max_ac, material, temp)
        else:
            self.wire_sizing['ac_homerun'] = ''
        
        self.refresh_wire_sizing_table()
    
    def _collect_active_string_counts(self):
        """Scan segments and return sorted list of harness string counts in use."""
        string_counts = set()
        lv_method = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
        
        for group in self.groups:
            for segment in group.get('segments', []):
                if lv_method in ('String HR', 'Trunk Bus'):
                    string_counts.add(1)
                else:
                    harness_config = segment.get('harness_config', '')
                    if harness_config:
                        for part in harness_config.split('+'):
                            try:
                                sc = int(part.strip())
                                if sc > 0:
                                    string_counts.add(sc)
                            except ValueError:
                                pass
                    else:
                        spt = segment.get('strings_per_tracker', 1)
                        if spt > 0:
                            string_counts.add(spt)
        
        # Also include string counts from split tracker harness portions
        split_details = getattr(self, '_split_tracker_details', {})
        tracker_seg_map = getattr(self, '_tracker_to_segment', [])
        for tidx, details in split_details.items():
            for portion in details.get('portions', []):
                for h_size in portion.get('harnesses', []):
                    if h_size > 0:
                        string_counts.add(h_size)

        result = sorted(string_counts)

        return result

    def get_wire_size_for(self, cable_type, num_strings=None):
        """
        Look up wire size from the wire_sizing dict.
        
        Args:
            cable_type: 'harness', 'extender', 'whip', 'dc_feeder', 'ac_homerun'
            num_strings: Required for harness/extender/whip — the harness string count
            
        Returns:
            str: Cable size string (e.g., '10 AWG', '500 kcmil')
        """
        if cable_type in ('dc_feeder', 'ac_homerun'):
            size = self.wire_sizing.get(cable_type, '')
            return size if size else '4/0 AWG'
        
        if num_strings is None:
            return '10 AWG'
        
        by_sc = self.wire_sizing.get('by_string_count', {})
        # Try int key first, then string key (JSON round-trip converts int keys to strings)
        entry = by_sc.get(num_strings) or by_sc.get(str(num_strings))
        if entry:
            return entry.get(cable_type, '10 AWG')

        # No exact match — find the nearest configured string count.
        # Split trackers (e.g. 1-string) inherit from the closest configured count.
        if by_sc and num_strings is not None:
            int_keys = []
            for k in by_sc.keys():
                try:
                    int_keys.append(int(k))
                except (ValueError, TypeError):
                    pass
            if int_keys:
                # Prefer the closest count <= num_strings; otherwise use the minimum
                floor_keys = [k for k in int_keys if k <= num_strings]
                closest = max(floor_keys) if floor_keys else min(int_keys)
                fallback_entry = by_sc.get(closest) or by_sc.get(str(closest))
                if fallback_entry:
                    return fallback_entry.get(cable_type, '10 AWG')

        return '10 AWG'
    
    def _gauge_to_string_count(self, cable_type, gauge):
        """Reverse-lookup: find a string count that maps to this gauge for a given cable type.
        Used to pass num_strings to lookup_part_and_price for correct part matching."""
        by_sc = self.wire_sizing.get('by_string_count', {})
        for sc_key, entry in by_sc.items():
            if entry.get(cable_type) == gauge:
                return int(sc_key) if isinstance(sc_key, str) else sc_key
        return None

    
    def _get_effective_wire_size(self, cable_type):
        """
        Get the effective wire size for a cable type across all active string counts.
        If all string counts agree, returns that size. Otherwise returns the largest.
        Used for whips/extenders which are bucketed by length without string count context.
        
        Args:
            cable_type: 'harness', 'extender', or 'whip'
        Returns:
            str: Cable size string
        """
        from src.utils.cable_sizing import get_cable_size_index
        
        by_sc = self.wire_sizing.get('by_string_count', {})
        if not by_sc:
            return '10 AWG'
        
        sizes = set()
        for sc_key, entry in by_sc.items():
            size = entry.get(cable_type, '10 AWG')
            sizes.add(size)
        
        if len(sizes) == 1:
            return sizes.pop()
        
        # Multiple sizes — return the largest
        best = None
        best_idx = -1
        for s in sizes:
            idx = get_cable_size_index(s)
            if idx > best_idx:
                best_idx = idx
                best = s
        
        result = best or '10 AWG'
        return result

    def lookup_part_and_price(self, item_type, **kwargs):
        """Look up part number and unit price for a BOM item.
        item_type: 'harness', 'whip', 'extender'
        Returns (part_number, unit_price_str, ext_price_str)
        """
        try:
            import os
            import json
            from src.utils.pricing_lookup import PricingLookup

            # Wire gauge is now looked up per cable type from wire_sizing dict
            wire_gauge = None  # Will be set per item_type below

            pricing = PricingLookup()
            part_number = 'N/A'

            if item_type == 'harness':
                num_strings = kwargs.get('num_strings', 1)
                polarity = kwargs.get('polarity', 'positive')
                qty = kwargs.get('qty', 1)

                # Calculate string spacing from module width and modules per string
                mps = self._get_int_var(self.modules_per_string_var, 28)
                module_width_mm = self.selected_module.width_mm if self.selected_module else 1134
                # Use 0.02m (20mm) as default module gap spacing
                string_spacing_ft = (mps * module_width_mm + (mps - 1) * 20) / 1000 * 3.28084

                # Load harness library
                current_dir = os.path.dirname(os.path.abspath(__file__))
                project_root = os.path.dirname(os.path.dirname(current_dir))
                lib_path = os.path.join(project_root, 'data', 'harness_library.json')
                with open(lib_path, 'r') as f:
                    harness_library = json.load(f)

                available_spacings = [26.0, 102.0, 113.0, 122.0, 133.0]
                target_spacing = max(available_spacings)
                for sp in sorted(available_spacings):
                    if sp >= string_spacing_ft:
                        target_spacing = sp
                        break

                matches = []
                for pn, spec in harness_library.items():
                    if pn.startswith('_comment_'):
                        continue
                    if (spec.get('num_strings') == num_strings and
                            spec.get('polarity') == polarity and
                            abs(spec.get('string_spacing_ft', 0) - target_spacing) < 0.1):
                        spec_trunk = spec.get('trunk_cable_size', spec.get('trunk_wire_gauge', ''))
                        harness_gauge = self.get_wire_size_for('harness', num_strings)
                        if spec_trunk == harness_gauge:
                            matches.append(pn)
                part_number = matches[0] if len(matches) == 1 else ('N/A' if not matches else ' or '.join(sorted(matches)))

            elif item_type in ('whip', 'extender'):
                polarity = kwargs.get('polarity', 'positive')
                length_ft = kwargs.get('length_ft', 10)
                qty = kwargs.get('qty', 1)

                target_length = ((length_ft - 1) // 5 + 1) * 5
                target_length = max(10, target_length)

                lib_name = 'whip_library.json' if item_type == 'whip' else 'extender_library.json'
                current_dir = os.path.dirname(os.path.abspath(__file__))
                project_root = os.path.dirname(os.path.dirname(current_dir))
                lib_path = os.path.join(project_root, 'data', lib_name)
                with open(lib_path, 'r') as f:
                    library = json.load(f)

                # Look up wire gauge for this cable type
                # If num_strings is provided, use per-string-count size; otherwise use effective size
                item_num_strings = kwargs.get('num_strings', None)
                if item_num_strings is not None:
                    item_wire_gauge = self.get_wire_size_for(item_type, item_num_strings)
                else:
                    item_wire_gauge = self._get_effective_wire_size(item_type)

                for pn, spec in library.items():
                    if pn.startswith('_comment_'):
                        continue
                    if (spec.get('wire_gauge') == item_wire_gauge and
                            spec.get('polarity') == polarity and
                            spec.get('length_ft') == target_length):
                        part_number = pn
                        break
            else:
                qty = kwargs.get('qty', 1)

            # Look up price
            unit_price = pricing.get_price(part_number) if part_number != 'N/A' else None
            if unit_price is not None:
                qty_val = kwargs.get('qty', 1)
                return part_number, f"${unit_price:,.2f}", f"${unit_price * qty_val:,.2f}"
            else:
                return part_number, '', ''

        except Exception as e:
            print(f"lookup_part_and_price error ({item_type}): {e}")
            return 'N/A', '', ''

    def _lookup_description(self, part_number):
        """Return the library description for part_number, or 'N/A' if not found."""
        if not part_number or part_number in ('N/A', ''):
            return 'N/A'
        import os as _os, json as _json
        cur = _os.path.dirname(_os.path.abspath(__file__))
        root = _os.path.dirname(_os.path.dirname(cur))
        for lib in ('harness_library.json', 'whip_library.json', 'extender_library.json',
                    'fuse_library.json', 'combiner_box_library.json', 'combiner_box_fuse_library.json'):
            try:
                with open(_os.path.join(root, 'data', lib)) as _f:
                    data = _json.load(_f)
                if part_number in data:
                    return data[part_number].get('description', 'N/A')
            except Exception:
                pass
        return 'N/A'

    def _on_results_tree_click(self, event):
        """Handle click on results tree - toggle checkbox if include column clicked"""
        item = self.results_tree.identify_row(event.y)
        column = self.results_tree.identify_column(event.x)
        if item and column == '#1':  # include column
            self._toggle_result_item(item)

    def _toggle_result_item(self, item):
        """Toggle the checked state of a results tree item"""
        values = list(self.results_tree.item(item, 'values'))
        # Don't allow toggling section headers
        if str(values[1]).startswith('---'):
            return
        if values[0] == '☐':
            values[0] = '☑'
            self.checked_items.add(item)
            self.results_tree.item(item, values=values, tags=('checked',))
        else:
            values[0] = '☐'
            self.checked_items.discard(item)
            self.results_tree.item(item, values=values, tags=('unchecked',))

    def _results_select_all(self):
        """Check all non-section rows in results tree"""
        for item in self.results_tree.get_children():
            values = list(self.results_tree.item(item, 'values'))
            if not str(values[1]).startswith('---'):
                values[0] = '☑'
                self.checked_items.add(item)
                self.results_tree.item(item, values=values, tags=('checked',))

    def _results_select_none(self):
        """Uncheck all rows in results tree"""
        for item in self.results_tree.get_children():
            values = list(self.results_tree.item(item, 'values'))
            if not str(values[1]).startswith('---'):
                values[0] = '☐'
                self.checked_items.discard(item)
                self.results_tree.item(item, values=values, tags=('unchecked',))
        self.checked_items.clear()

    def get_default_combiner_price(self):
        """Get default combiner box price from pricing data"""
        try:
            combiner_data = self.pricing_data.get('combiner_boxes', {})
            for key, value in combiner_data.items():
                if isinstance(value, dict) and 'price' in value:
                    return float(value['price'])
                elif isinstance(value, (int, float)):
                    return float(value)
            return 0.0
        except:
            return 0.0

    def _get_default_harness_config_from_template(self, template_ref):
        """Derive the default harness config string from a tracker template's motor position.
        
        Convention: first number = north side of motor, second = south side.
        If no motor, returns single harness (full SPT as a string).
        
        Returns a config string like "7+6" or "13".
        """
        if not template_ref or template_ref not in self.enabled_templates:
            return None
        
        tdata = self.enabled_templates[template_ref]
        spt = tdata.get('strings_per_tracker', 3)
        has_motor = tdata.get('has_motor', True)
        
        if not has_motor:
            return str(int(spt))
        
        motor_placement = tdata.get('motor_placement_type', 'between_strings')
        
        if motor_placement == 'between_strings':
            motor_pos_after = tdata.get('motor_position_after_string', None)
            if motor_pos_after is not None:
                north = int(motor_pos_after)
                south = int(spt) - north
                if north > 0 and south > 0:
                    return f"{north}+{south}"
        elif motor_placement == 'middle_of_string':
            motor_string_idx = tdata.get('motor_string_index', None)
            if motor_string_idx is not None:
                # Motor is in the middle of a string, so the split is at that string boundary
                # motor_string_index is 0-based; strings 0..idx are north, idx+1..end are south
                north = int(motor_string_idx) + 1
                south = int(spt) - north
                if north > 0 and south > 0:
                    return f"{north}+{south}"
        
        # Fallback: single harness
        return str(int(spt))

    def get_harness_options(self, num_strings):
        """Generate harness configuration options for a given string count.
        
        Returns the full-size option plus all 2-part splits (largest first).
        Max 2 harnesses per tracker.
        """
        num_strings = int(num_strings)
        if num_strings < 1:
            return ["1"]
        
        options = [str(num_strings)]
        
        # 2-part splits: both orderings (north+south convention)
        for i in range(1, num_strings):
            if i == num_strings - i:
                # Equal split — only one entry needed
                options.append(f"{i}+{i}")
            else:
                options.append(f"{num_strings - i}+{i}")
        
        return options

    def parse_harness_config(self, config_str):
        """Parse harness config string like '2+1' into list of integers [2, 1]"""
        try:
            return [int(x) for x in config_str.split('+')]
        except:
            return []
        
    def _get_effective_harness_config(self, seg):
        """Get the effective harness config for a segment, considering LV Collection Method.
        
        When 'String HR' is selected, override to all 1-string harnesses (1+1+...+1).
        Otherwise, return the segment's stored harness_config.
        """
        lv_method = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
        if lv_method in ('String HR', 'Trunk Bus'):
            spt = seg.get('strings_per_tracker', 1)
            return '+'.join(['1'] * int(spt))
        return seg.get('harness_config', str(seg.get('strings_per_tracker', 1)))
        
    def calculate_extender_lengths_per_segment(self, seg, device_position, string_offset=0, target_y_offset=0, harness_sizes_override=None):
        """Calculate per-harness positive and negative extender lengths for a segment.
        
        Returns a list of (pos_length_ft, neg_length_ft) tuples, one per harness in the config.
        Multiply each by seg['quantity'] for total counts.
        """
        template_ref = seg.get('template_ref')
        if harness_sizes_override is not None:
            harness_sizes = harness_sizes_override
        else:
            harness_config = self._get_effective_harness_config(seg)
            harness_sizes = self.parse_harness_config(harness_config)
        spt = seg['strings_per_tracker']
        
        if not harness_sizes:
            return []
        
        m_to_ft = 3.28084
        
        # Get template geometry
        tracker_length_ft = None
        string_length_ft = None
        motor_y_ft = None
        motor_gap_ft = 0
        has_motor = False
        
        if template_ref and template_ref in self.enabled_templates:
            tdata = self.enabled_templates[template_ref]
            mod_spec = tdata.get('module_spec', {})
            orientation = tdata.get('module_orientation', 'Portrait')
            mps = tdata.get('modules_per_string', 28)
            spacing_m = tdata.get('module_spacing_m', 0.02)
            has_motor = tdata.get('has_motor', True)
            motor_gap_m = tdata.get('motor_gap_m', 1.0) if has_motor else 0
            
            if orientation == 'Portrait':
                mod_along_m = mod_spec.get('width_mm', 1000) / 1000
            else:
                mod_along_m = mod_spec.get('length_mm', 2000) / 1000
            
            string_length_ft = (mps * mod_along_m + (mps - 1) * spacing_m) * m_to_ft
            motor_gap_ft = motor_gap_m * m_to_ft
            
            # Total tracker length — use full strings + partial modules
            full_spt = int(spt)
            partial_mods = round((spt - full_spt) * mps) if spt != full_spt else 0
            total_modules = full_spt * mps + partial_mods
            tracker_length_m = (total_modules * mod_along_m + 
                               (total_modules - 1) * spacing_m + 
                               (motor_gap_m if has_motor else 0))
            tracker_length_ft = tracker_length_m * m_to_ft
            
            # Motor Y position from north end
            if has_motor:
                motor_placement = tdata.get('motor_placement_type', 'between_strings')
                motor_pos_after = tdata.get('motor_position_after_string', None)
                motor_string_idx = tdata.get('motor_string_index', None)
                motor_split_north = tdata.get('motor_split_north', mps // 2)
                
                # Partial string on north adds height before everything
                partial_north_mods = 0
                spt_val = tdata.get('strings_per_tracker', 1)
                if spt_val != int(spt_val) and tdata.get('partial_string_side', 'north') == 'north':
                    partial_north_mods = round((spt_val - int(spt_val)) * mps)
                partial_north_m = partial_north_mods * (mod_along_m + spacing_m) if partial_north_mods > 0 else 0
                
                # Partial string on north shifts motor further south
                partial_north_mods_ext = 0
                if spt != int(spt):
                    tdata_ext = self.enabled_templates.get(template_ref, {})
                    if tdata_ext.get('partial_string_side', 'north') == 'north':
                        partial_north_mods_ext = round((spt - int(spt)) * mps)
                partial_north_m_ext = partial_north_mods_ext * (mod_along_m + spacing_m)
                
                if motor_placement == 'between_strings':
                    if motor_pos_after is not None:
                        strings_before = motor_pos_after
                    else:
                        strings_before = 1  # default: after first string
                    modules_before = strings_before * mps
                    motor_y_m = partial_north_m + (modules_before * mod_along_m + 
                                (modules_before - 1) * spacing_m + spacing_m)
                    motor_y_ft = motor_y_m * m_to_ft
                elif motor_placement == 'middle_of_string':
                    if motor_string_idx is not None:
                        strings_before = motor_string_idx - 1  # 1-based
                    else:
                        strings_before = 0
                    modules_before = strings_before * mps + motor_split_north
                    motor_y_m = partial_north_m + (modules_before * mod_along_m + 
                                max(modules_before - 1, 0) * spacing_m + spacing_m)
                    motor_y_ft = motor_y_m * m_to_ft
                else:
                    motor_y_ft = tracker_length_ft / 2
            else:
                motor_y_ft = tracker_length_ft / 2
        
        # Fallback if template not found
        if tracker_length_ft is None:
            module_width_ft = (self.selected_module.width_mm / 304.8) if self.selected_module else 3.3
            try:
                mps = int(self.modules_per_string_var.get())
            except ValueError:
                mps = 28
            string_length_ft = module_width_ft * mps
            motor_gap_ft = 3.28
            tracker_length_ft = string_length_ft * spt + motor_gap_ft
            motor_y_ft = tracker_length_ft / 2
            has_motor = True
        
        # Determine extender target Y based on device position
        device_offset_ft = 5.0
        if device_position == 'north':
            target_y = -device_offset_ft
        elif device_position == 'south':
            target_y = tracker_length_ft + device_offset_ft
        else:  # middle
            target_y = motor_y_ft

        # Shift target for far-away trackers (signed: positive = CB south)
        target_y += target_y_offset
        
        # Resolve polarity convention
        polarity = self.polarity_convention_var.get()
        
        # Determine which absolute string index the motor gap follows
        motor_after_string = None  # 0-based absolute string index
        if has_motor and template_ref and template_ref in self.enabled_templates:
            tdata_motor = self.enabled_templates[template_ref]
            motor_placement = tdata_motor.get('motor_placement_type', 'between_strings')
            if motor_placement == 'between_strings':
                pos_after = tdata_motor.get('motor_position_after_string', None)
                if pos_after is not None:
                    motor_after_string = pos_after - 1 if pos_after > 0 else 0
                else:
                    motor_after_string = 0
            elif motor_placement == 'middle_of_string':
                # Motor is inside a string — gap is effectively after (string_index - 1)
                # but for extender purposes, treat it as between string boundaries
                idx = tdata_motor.get('motor_string_index', 1)
                motor_after_string = idx - 1 if idx > 0 else 0
        elif has_motor:
            motor_after_string = 0  # default fallback
        
        # Build ALL string boundary positions for the full tracker N→S
        inter_string_gap = (spacing_m if (template_ref and template_ref in self.enabled_templates) else 0.02) * m_to_ft
        full_string_count = int(spt)
        all_string_positions = []  # (north_edge, south_edge) for every string on tracker
        y_cursor = 0.0
        for abs_idx in range(full_string_count):
            north_edge = y_cursor
            south_edge = y_cursor + string_length_ft
            all_string_positions.append((north_edge, south_edge))
            y_cursor = south_edge
            if has_motor and motor_after_string is not None and abs_idx == motor_after_string:
                y_cursor += motor_gap_ft
            elif abs_idx < full_string_count - 1:
                y_cursor += inter_string_gap

        # Assign harness indices starting from string_offset
        # (For non-split trackers offset=0; for split portions, offset = physical start position)
        string_positions = []
        str_pos = string_offset
        for h_idx, h_size in enumerate(harness_sizes):
            for s in range(h_size):
                if str_pos < len(all_string_positions):
                    n_edge, s_edge = all_string_positions[str_pos]
                    string_positions.append((n_edge, s_edge, h_idx))
                    str_pos += 1
        
        # Calculate per-harness extender lengths
        result = []
        for h_idx, h_size in enumerate(harness_sizes):
            harness_strings = [(n, s) for n, s, hi in string_positions if hi == h_idx]
            
            if not harness_strings:
                result.append((10.0, 10.0))
                continue
            
            # Find the device-side string (closest to target_y)
            # The extender only runs from the harness's device-side terminal
            # to the whip point — the harness cable covers the internal span.
            if device_position == 'north':
                device_side = harness_strings[0]    # northernmost string
            elif device_position == 'south':
                device_side = harness_strings[-1]   # southernmost string
            else:  # middle
                harness_center = (harness_strings[0][0] + harness_strings[-1][1]) / 2
                if harness_center <= target_y:
                    device_side = harness_strings[-1]   # harness north of device
                else:
                    device_side = harness_strings[0]    # harness south of device
            
            ds_north = device_side[0]  # north edge of device-side string
            ds_south = device_side[1]  # south edge of device-side string
            
            # Determine which edge is positive and which is negative
            if polarity == 'Negative Always South':
                pos_y = ds_north
                neg_y = ds_south
            elif polarity == 'Negative Always North':
                pos_y = ds_south
                neg_y = ds_north
            elif polarity == 'Negative Toward Device':
                if device_position == 'north':
                    neg_y = ds_north
                    pos_y = ds_south
                elif device_position == 'south':
                    neg_y = ds_south
                    pos_y = ds_north
                else:
                    if ds_north < motor_y_ft:
                        neg_y = ds_south
                        pos_y = ds_north
                    else:
                        neg_y = ds_north
                        pos_y = ds_south
            elif polarity == 'Positive Toward Device':
                if device_position == 'north':
                    pos_y = ds_north
                    neg_y = ds_south
                elif device_position == 'south':
                    pos_y = ds_south
                    neg_y = ds_north
                else:
                    if ds_north < motor_y_ft:
                        pos_y = ds_south
                        neg_y = ds_north
                    else:
                        pos_y = ds_north
                        neg_y = ds_south
            else:
                pos_y = ds_north
                neg_y = ds_south
            
            pos_extender = max(abs(pos_y - target_y), 5.0)
            neg_extender = max(abs(neg_y - target_y), 5.0)
            
            result.append((pos_extender, neg_extender))

        return result
        
    def load_estimate(self):
        """Load estimate data from the project"""
        if not self.current_project or not self.estimate_id:
            return
        
        estimate_data = self.current_project.quick_estimates.get(self.estimate_id)
        if not estimate_data:
            return
        
        self._loading = True
        
        # Module is now derived from templates — skip old module dropdown restore
        # (module will be set by _derive_module_from_templates after groups load)
        
        # Restore numeric fields
        # modules_per_string: store saved value as fallback; templates will override if linked
        self._saved_modules_per_string = estimate_data.get('modules_per_string', 28)
        if hasattr(self, 'modules_per_string_var'):
            self.modules_per_string_var.set(str(self._saved_modules_per_string))
        # Load wire sizing (new format) or migrate from old wire_gauge
        saved_wire_sizing = estimate_data.get('wire_sizing')
        if saved_wire_sizing:
            self.wire_sizing = copy.deepcopy(saved_wire_sizing)
        else:
            # Backward compat: old estimates had a single wire_gauge field
            self.wire_sizing = {
                'temp_rating': '90C',
                'feeder_material': 'aluminum',
                'by_string_count': {},
                'dc_feeder': '',
                'ac_homerun': '',
                'user_overrides': {}
            }
        
        # Restore inverter selection
        saved_inverter_name = estimate_data.get('inverter_name', '')
        if saved_inverter_name and hasattr(self, 'inverter_combo'):
            if saved_inverter_name in self.available_inverters:
                self.inverter_select_var.set(saved_inverter_name)
                self._on_inverter_selected()
        
        # Restore topology and DC:AC ratio
        if hasattr(self, 'topology_var'):
            self.topology_var.set(estimate_data.get('topology', 'Distributed String'))
        if hasattr(self, 'central_inv_count_var'):
            self.central_inv_count_var.set(str(estimate_data.get('central_inverter_count', '1')))
        if hasattr(self, 'dc_ac_ratio_var'):
            self.dc_ac_ratio_var.set(str(estimate_data.get('dc_ac_ratio', 1.25)))
        if hasattr(self, 'breaker_size_var'):
            self.breaker_size_var.set(estimate_data.get('breaker_size', '400'))
        if hasattr(self, 'dc_feeder_distance_var'):
            self.dc_feeder_distance_var.set(str(estimate_data.get('dc_feeder_distance', 500)))
        if hasattr(self, 'polarity_convention_var'):
            self.polarity_convention_var.set(estimate_data.get('polarity_convention', 'Negative Always South'))
        if hasattr(self, 'lv_collection_var'):
            self.lv_collection_var.set(estimate_data.get('lv_collection_method', 'Wire Harness'))
        if hasattr(self, 'ac_homerun_distance_var'):
            self.ac_homerun_distance_var.set(str(estimate_data.get('ac_homerun_distance', 50)))
        
        # Load groups (new format) or convert from old subarrays format
        saved_subarrays = estimate_data.get('subarrays', {})
        saved_groups = estimate_data.get('groups', estimate_data.get('rows', []))
        
        self.groups.clear()
        self._refresh_group_listbox()
        
        if saved_groups:
            # Ensure all segments have template_ref field (backward compat)
            global_row_spacing_fallback = estimate_data.get('row_spacing_ft', 20.0)
            for group in saved_groups:
                for seg in group.get('segments', []):
                    if 'template_ref' not in seg:
                        seg['template_ref'] = None
                # Backward compat: older saves have no per-group row spacing
                if 'row_spacing_ft' not in group:
                    group['row_spacing_ft'] = global_row_spacing_fallback
                # Backward compat: older saves have no per-group strings_per_inv
                if 'strings_per_inv' not in group:
                    group['strings_per_inv'] = None
            self.groups = copy.deepcopy(saved_groups)
        elif saved_subarrays:
            # Backward compat: convert old subarray/block format to groups
            group_num = 0
            for subarray_id, subarray_data in saved_subarrays.items():
                for block_id, block_data in subarray_data.get('blocks', {}).items():
                    group_num += 1
                    segments = []
                    for tracker in block_data.get('trackers', []):
                        segments.append({
                            'quantity': tracker.get('quantity', 1),
                            'strings_per_tracker': tracker.get('strings', 3),
                            'harness_config': tracker.get('harness_config', '3'),
                            'template_ref': None
                        })
                    if not segments:
                        segments = [{'quantity': 1, 'strings_per_tracker': 3, 'harness_config': '3', 'template_ref': None}]
                    self.groups.append({
                        'name': block_data.get('name', f"Group {group_num}"),
                        'segments': segments
                    })
        else:
            self._loading = False
            self.add_group()
            return
        
        self._refresh_group_listbox()
        
        # Select first group
        if self.groups:
            self.group_listbox.selection_set(0)
            self.on_group_select(None)
        
        self._last_inspect_mode = estimate_data.get('inspect_mode', False)
        if hasattr(self, 'use_routed_var'):
                    self.use_routed_var.set(estimate_data.get('use_routed_distances', False))
        self.pads = copy.deepcopy(estimate_data.get('pads', []))
        self.measurements = copy.deepcopy(estimate_data.get('measurements', []))
        
        # Load device names (convert str keys back to int)
        saved_names = estimate_data.get('device_names', {})
        self.device_names = {int(k): v for k, v in saved_names.items()}
        
        # Load per-device feeder sizes (convert str keys back to int)
        saved_feeder_sizes = estimate_data.get('device_feeder_sizes', {})
        self.device_feeder_sizes = {int(k): v for k, v in saved_feeder_sizes.items()}
        
        # Load per-device parallel counts (convert str keys back to int)
        saved_parallel_counts = estimate_data.get('device_feeder_parallel_counts', {})
        self.device_feeder_parallel_counts = {int(k): int(v) for k, v in saved_parallel_counts.items()}
        
        # Load allocation lock state
        self.allocation_locked = estimate_data.get('allocation_locked', False)
        self.locked_allocation_result = estimate_data.get('locked_allocation_result', None)
        
        # Restore combiner assignments (includes start_string_pos for physical ordering)
        saved_cb_assignments = estimate_data.get('combiner_assignments', None)
        if saved_cb_assignments:
            self.last_combiner_assignments = copy.deepcopy(saved_cb_assignments)

        # Restore SI assignments
        saved_si_assignments = estimate_data.get('si_assignments', None)
        if saved_si_assignments:
            self.last_si_assignments = copy.deepcopy(saved_si_assignments)

        # Derive module from templates
        self._derive_module_from_templates()

        # Refresh wire sizing table — always reconcile saved data with current segments
        # This adds missing string counts and removes stale ones
        self._refresh_wire_sizing_for_segments()

        # Re-enable autosave now that loading is complete
        self._loading = False

        # Auto-calculate on load (silent — skip warning if no module yet)
        self.after(100, lambda: self.calculate_estimate(silent=True))

    def save_estimate(self):
        """Save estimate data to the project"""
        if not self.current_project or not self.estimate_id:
            return
        
        # Get current estimate or create new one
        if self.estimate_id not in self.current_project.quick_estimates:
            self.current_project.quick_estimates[self.estimate_id] = {
                'name': 'Unnamed Estimate',
                'created_date': datetime.now().isoformat(),
            }
        
        estimate_data = self.current_project.quick_estimates[self.estimate_id]
        
        # Update global settings
        if self.selected_module:
            estimate_data['module_name'] = f"{self.selected_module.manufacturer} {self.selected_module.model} ({self.selected_module.wattage}W)"
            estimate_data['module_width_mm'] = self.selected_module.width_mm
            estimate_data['module_isc'] = self.selected_module.isc
        
        estimate_data['modules_per_string'] = self._get_int_var(self.modules_per_string_var, 28)
        estimate_data['wire_sizing'] = copy.deepcopy(self.wire_sizing)
        
        # Save inverter selection
        if self.selected_inverter:
            estimate_data['inverter_name'] = self.inverter_select_var.get()
        
        # Save topology and DC:AC ratio
        estimate_data['topology'] = self.topology_var.get()
        estimate_data['central_inverter_count'] = self.central_inv_count_var.get() if hasattr(self, 'central_inv_count_var') else '1'
        estimate_data['inspect_mode'] = getattr(self, '_last_inspect_mode', False)
        estimate_data['use_routed_distances'] = self.use_routed_var.get()
        estimate_data['breaker_size'] = self.breaker_size_var.get()
        estimate_data['polarity_convention'] = self.polarity_convention_var.get()
        estimate_data['lv_collection_method'] = self.lv_collection_var.get()
        estimate_data['dc_feeder_distance'] = self._get_float_var(self.dc_feeder_distance_var, estimate_data.get('dc_feeder_distance', 500.0))
        estimate_data['ac_homerun_distance'] = self._get_float_var(self.ac_homerun_distance_var, estimate_data.get('ac_homerun_distance', 500.0))
        estimate_data['dc_ac_ratio'] = self._get_float_var(self.dc_ac_ratio_var, 1.25)
        
        # Update modified date
        estimate_data['modified_date'] = datetime.now().isoformat()
        
        # Save groups (new format) — deep copy to avoid reference aliasing
        estimate_data['groups'] = copy.deepcopy(self.groups)
        estimate_data['subarrays'] = {}
        
        # Save pads and measurements
        estimate_data['pads'] = copy.deepcopy(self.pads)
        estimate_data['measurements'] = copy.deepcopy(self.measurements)
        
        # Save device names (convert int keys to str for JSON)
        estimate_data['device_names'] = {str(k): v for k, v in self.device_names.items()}
        
        # Save per-device feeder sizes (convert int keys to str for JSON)
        estimate_data['device_feeder_sizes'] = {str(k): v for k, v in self.device_feeder_sizes.items()}
        
        # Save per-device parallel counts (convert int keys to str for JSON)
        estimate_data['device_feeder_parallel_counts'] = {str(k): v for k, v in self.device_feeder_parallel_counts.items()}
        
        # Save allocation lock state
        estimate_data['allocation_locked'] = self.allocation_locked
        if self.allocation_locked and self.locked_allocation_result is not None:
            estimate_data['locked_allocation_result'] = copy.deepcopy(self.locked_allocation_result)
        else:
            estimate_data['locked_allocation_result'] = None
        
        # Save combiner assignments (includes start_string_pos for physical ordering)
        if self.last_combiner_assignments:
            estimate_data['combiner_assignments'] = copy.deepcopy(self.last_combiner_assignments)
        else:
            estimate_data['combiner_assignments'] = None

        # Save SI assignments (inverter_spec is not serialisable — strip it before saving)
        if self.last_si_assignments:
            si_serialisable = []
            for entry in self.last_si_assignments:
                si_serialisable.append({k: v for k, v in entry.items() if k != 'inverter_spec'})
            estimate_data['si_assignments'] = si_serialisable
        else:
            estimate_data['si_assignments'] = None

        # Notify callback
        if self.on_save:
            self.on_save()

    # ==================== Estimate Management ====================

    def _refresh_estimate_dropdown(self, auto_select=False):
        """Refresh the estimate dropdown with saved estimates"""
        if not self.current_project:
            return
        
        estimates = self.current_project.quick_estimates
        
        # Build list of estimate names for dropdown
        estimate_names = []
        self._estimate_id_map = {}
        
        for est_id, est_data in estimates.items():
            name = est_data.get('name', 'Unnamed Estimate')
            estimate_names.append(name)
            self._estimate_id_map[name] = est_id
        
        if hasattr(self, 'estimate_combo'):
            self.estimate_combo['values'] = estimate_names
        
        if auto_select and estimates:
            # Find most recently modified
            most_recent_id = None
            most_recent_date = None
            for est_id, est_data in estimates.items():
                mod_date = est_data.get('modified_date')
                if mod_date:
                    if most_recent_date is None or mod_date > most_recent_date:
                        most_recent_date = mod_date
                        most_recent_id = est_id
            
            if most_recent_id is None:
                most_recent_id = list(estimates.keys())[0]
            
            self.estimate_id = most_recent_id
            if hasattr(self, 'estimate_var'):
                self.estimate_var.set(estimates[most_recent_id].get('name', ''))
            self.load_estimate()
        elif auto_select and not estimates:
            # No estimates exist — create a default one
            self.new_estimate()

    def _on_estimate_selected(self, event=None):
        """Handle selection of an estimate from the dropdown"""
        # Save current estimate before switching
        if self.estimate_id:
            self.save_estimate()
        
        selected_name = self.estimate_var.get()
        if selected_name in self._estimate_id_map:
            self.estimate_id = self._estimate_id_map[selected_name]
            self.load_estimate()

    def _on_estimate_name_changed(self, event=None):
        """Handle renaming of the current estimate via the combobox"""
        if not self.estimate_id or not self.current_project:
            return
        
        new_name = self.estimate_var.get().strip()
        if not new_name:
            return
        
        if self.estimate_id in self.current_project.quick_estimates:
            self.current_project.quick_estimates[self.estimate_id]['name'] = new_name
            self._refresh_estimate_dropdown()
            self.estimate_var.set(new_name)

    def new_estimate(self):
        """Create a new quick estimate"""
        if not self.current_project:
            return
        
        # Save current estimate first
        if self.estimate_id:
            self.save_estimate()
        
        # Generate new ID and name
        estimate_id = f"estimate_{uuid.uuid4().hex[:8]}"
        existing_names = {e.get('name', '') for e in self.current_project.quick_estimates.values()}
        rev_num = 0
        while f"rev{rev_num}" in existing_names:
            rev_num += 1
        estimate_name = f"rev{rev_num}"
        
        new_estimate = {
            'name': estimate_name,
            'created_date': datetime.now().isoformat(),
            'modified_date': datetime.now().isoformat(),
            'row_spacing_ft': 20.0,
            'topology': 'Centralized String',
            'dc_ac_ratio': 1.25,
            'subarrays': {},
            'groups': [],
            'pads': []
        }
        
        self.current_project.quick_estimates[estimate_id] = new_estimate
        
        self.estimate_id = estimate_id
        self._refresh_estimate_dropdown()
        self.estimate_var.set(estimate_name)
        
        # Clear and set up fresh
        self._clear_estimate_ui()
        self.add_group()
        self._derive_module_from_templates()
        
        if self.on_save:
            self.on_save()

    def copy_estimate(self):
        """Duplicate the currently selected estimate"""
        if not self.current_project or not self.estimate_id:
            messagebox.showinfo("No Estimate", "No estimate selected to copy.")
            return
        
        # Save current estimate first so we capture latest state
        self.save_estimate()
        
        # Get current estimate data
        source_data = self.current_project.quick_estimates.get(self.estimate_id)
        if not source_data:
            return
        
        # Deep copy the estimate data
        new_data = copy.deepcopy(source_data)
        
        # Generate new ID and name
        new_id = f"estimate_{uuid.uuid4().hex[:8]}"
        source_name = source_data.get('name', 'Unnamed')
        new_data['name'] = f"{source_name} (Copy)"
        new_data['created_date'] = datetime.now().isoformat()
        new_data['modified_date'] = datetime.now().isoformat()
        
        # Store in project
        self.current_project.quick_estimates[new_id] = new_data
        
        # Switch to the new copy
        self.estimate_id = new_id
        self._refresh_estimate_dropdown()
        self.estimate_var.set(new_data['name'])
        self.load_estimate()
        
        if self.on_save:
            self.on_save()

    def delete_estimate(self):
        """Delete the currently selected estimate"""
        if not self.current_project or not self.estimate_id:
            messagebox.showinfo("No Estimate", "No estimate selected to delete.")
            return
        
        estimate_name = self.current_project.quick_estimates.get(
            self.estimate_id, {}
        ).get('name', 'this estimate')
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{estimate_name}'?"):
            del self.current_project.quick_estimates[self.estimate_id]
            self.estimate_id = None
            self._clear_estimate_ui()
            self._refresh_estimate_dropdown(auto_select=True)
            
            if self.on_save:
                self.on_save()

    def _auto_unlock_allocation(self):
        """Unlock allocation if locked, with a user notification.
        
        Called when structural changes (inverter, topology, segments, groups)
        invalidate a locked allocation.
        """
        if not self.allocation_locked:
            return
        
        self.allocation_locked = False
        self.locked_allocation_result = None
        
        messagebox.showinfo(
            "Allocation Unlocked",
            "The allocation lock has been released because the estimate structure changed.\n\n"
            "Re-run Calculate Estimate and lock again from the Site Preview if needed.",
            parent=self
        )

    def _reconcile_locked_allocation(self):
        """Drop connections for tracker indices that no longer exist.

        Called after subtractive structural changes (delete segment, decrease qty,
        delete group). The device simply loses those strings — no prompt, no unlock.
        """
        if not self.allocation_locked or self.locked_allocation_result is None:
            return

        total_trackers = sum(
            seg.get('quantity', 0)
            for group in self.groups
            for seg in group.get('segments', [])
        )

        for inv in self.locked_allocation_result.get('inverters', []):
            inv['harness_map'] = [
                e for e in inv.get('harness_map', [])
                if e.get('tracker_idx', 0) < total_trackers
            ]
            inv['total_strings'] = sum(
                e.get('strings_taken', 0) for e in inv['harness_map']
            )

        # Update summary counts
        total_strings = sum(
            inv.get('total_strings', 0)
            for inv in self.locked_allocation_result.get('inverters', [])
        )
        split_tidxs = set(
            e['tracker_idx']
            for inv in self.locked_allocation_result.get('inverters', [])
            for e in inv.get('harness_map', [])
            if e.get('is_split')
        )
        summary = self.locked_allocation_result.setdefault('summary', {})
        summary['total_strings'] = total_strings
        summary['total_split_trackers'] = len(split_tidxs)

        # Also drop invalid entries from combiner assignments
        for cb in getattr(self, 'last_combiner_assignments', []):
            cb['connections'] = [
                c for c in cb.get('connections', [])
                if c.get('tracker_idx', 0) < total_trackers
            ]

    def _shift_locked_allocation(self, insertion_point, delta):
        """Shift tracker_idx values in locked allocation after a qty insert or delete.

        insertion_point: first global index affected by the change
        delta: +N trackers inserted (shift existing up), -N trackers removed (drop then shift down)
        """
        if not self.allocation_locked or self.locked_allocation_result is None:
            return

        if delta > 0:
            for inv in self.locked_allocation_result.get('inverters', []):
                for e in inv.get('harness_map', []):
                    if e.get('tracker_idx', 0) >= insertion_point:
                        e['tracker_idx'] += delta
            for cb in getattr(self, 'last_combiner_assignments', []):
                for c in cb.get('connections', []):
                    if c.get('tracker_idx', 0) >= insertion_point:
                        c['tracker_idx'] += delta
        else:
            removed_end = insertion_point - delta  # delta negative → removed_end > insertion_point
            for inv in self.locked_allocation_result.get('inverters', []):
                inv['harness_map'] = [
                    e for e in inv.get('harness_map', [])
                    if not (insertion_point <= e.get('tracker_idx', 0) < removed_end)
                ]
                for e in inv['harness_map']:
                    if e.get('tracker_idx', 0) >= removed_end:
                        e['tracker_idx'] += delta
                inv['total_strings'] = sum(e.get('strings_taken', 0) for e in inv['harness_map'])
            for cb in getattr(self, 'last_combiner_assignments', []):
                cb['connections'] = [
                    c for c in cb.get('connections', [])
                    if not (insertion_point <= c.get('tracker_idx', 0) < removed_end)
                ]
                for c in cb['connections']:
                    if c.get('tracker_idx', 0) >= removed_end:
                        c['tracker_idx'] += delta

        total_strings = sum(
            e.get('strings_taken', 0)
            for inv in self.locked_allocation_result.get('inverters', [])
            for e in inv.get('harness_map', [])
        )
        split_tidxs = set(
            e['tracker_idx']
            for inv in self.locked_allocation_result.get('inverters', [])
            for e in inv.get('harness_map', [])
            if e.get('is_split')
        )
        summary = self.locked_allocation_result.setdefault('summary', {})
        summary['total_strings'] = total_strings
        summary['total_split_trackers'] = len(split_tidxs)

    def _count_unallocated_strings(self):
        """Return the number of strings not yet assigned to any device in the locked allocation."""
        if not self.allocation_locked or self.locked_allocation_result is None:
            return 0
        allocated_trackers = {
            e.get('tracker_idx')
            for inv in self.locked_allocation_result.get('inverters', [])
            for e in inv.get('harness_map', [])
        }
        count = 0
        global_idx = 0
        for group in self.groups:
            for seg in group.get('segments', []):
                ref = seg.get('template_ref')
                raw_spt = 1
                if ref and ref in getattr(self, 'enabled_templates', {}):
                    raw_spt = self.enabled_templates[ref].get('strings_per_tracker', 1)
                spt = int(raw_spt) + (1 if raw_spt != int(raw_spt) else 0)
                for _ in range(seg.get('quantity', 0)):
                    if global_idx not in allocated_trackers:
                        count += spt
                    global_idx += 1
        return count

    def _clear_estimate_ui(self):
        """Clear the groups and details when switching/deleting estimates"""
        # Clear groups and pads
        self.groups.clear()
        self.pads.clear()
        self.device_names.clear()
        self.device_feeder_sizes.clear()
        self.last_combiner_assignments = []
        self.last_si_assignments = []
        self.allocation_locked = False
        self.locked_allocation_result = None
        self.manually_edited = False
        if hasattr(self, 'group_listbox'):
            self._refresh_group_listbox()
        
        # Clear details panel
        if hasattr(self, 'details_container'):
            self.clear_details_panel()
        
        # Clear results
        if hasattr(self, 'results_tree'):
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)
        
        # Reset inverter and topology fields
        if hasattr(self, 'inverter_select_var'):
            self.inverter_select_var.set('')
            self.selected_inverter = None
            self.inverter_info_label.config(text="No inverter selected", foreground='gray')
        if hasattr(self, 'topology_var'):
            self.topology_var.set('Distributed String')
        if hasattr(self, 'lv_collection_var'):
            self.lv_collection_var.set('Wire Harness')
        if hasattr(self, 'breaker_size_var'):
            self.breaker_size_var.set('400')
        if hasattr(self, 'dc_ac_ratio_var'):
            self.dc_ac_ratio_var.set('1.25')
        if hasattr(self, 'strings_per_inverter_var'):
            self.strings_per_inverter_var.set('--')
        if hasattr(self, 'isc_warning_label'):
            self.isc_warning_label.config(text="")

    def _schedule_autosave(self):
        """Debounced auto-save — saves estimate after a brief pause"""
        if getattr(self, '_loading', False):
            return
        if hasattr(self, '_autosave_after_id') and self._autosave_after_id:
            self.after_cancel(self._autosave_after_id)
        self._autosave_after_id = self.after(1000, self._do_autosave)
    
    def _do_autosave(self):
        """Execute the debounced save"""
        self._autosave_after_id = None
        self.save_estimate()

    def _mark_stale(self):
        """Mark results as stale and re-enable the calculate button"""
        self._results_stale = True
        if self._calc_btn:
            self._calc_btn.config(state='normal')

    def _mark_dirty(self):
        """Mark results stale and schedule autosave — convenience for the common pair."""
        self._mark_stale()
        self._schedule_autosave()

    def _get_int_var(self, var, default=0):
        """Safely get an integer from a tk StringVar."""
        try:
            return int(var.get())
        except (ValueError, AttributeError):
            return default

    def _get_float_var(self, var, default=0.0):
        """Safely get a float from a tk StringVar."""
        try:
            return float(var.get())
        except (ValueError, AttributeError):
            return default

    def _get_harness_sizes(self, seg):
        """Parse the effective harness config for a segment into a list of ints."""
        return self.parse_harness_config(self._get_effective_harness_config(seg))

    def _rebuild_group_details(self, group_idx):
        """Destroy and rebuild the details panel for a group, then refresh dependent state."""
        self._harness_combos = []
        for widget in self.details_container.winfo_children():
            widget.destroy()
        self.show_group_details(group_idx)
        self._refresh_wire_sizing_for_segments()
        self._mark_dirty()

    def _on_lv_collection_changed(self, *args):
        """Handle LV Collection Method change — update harness combo states in details panel"""
        lv_method = self.lv_collection_var.get()
        
        # Update any currently displayed harness combos
        if hasattr(self, '_harness_combos'):
            for combo, harness_var, segment in self._harness_combos:
                try:
                    if lv_method == 'String HR':
                        spt = segment.get('strings_per_tracker', 1)
                        override_display = '+'.join(['1'] * int(spt))
                        original_config = segment['harness_config']
                        harness_var.set(override_display)
                        segment['harness_config'] = original_config  # restore saved config
                        combo.config(state='disabled')
                    elif lv_method == 'Trunk Bus':
                        combo.config(state='disabled')
                    else:
                        # Wire Harness — restore the real config and re-enable
                        harness_var.set(segment['harness_config'])
                        combo.config(state='readonly')
                except tk.TclError:
                    pass  # Widget may have been destroyed

        # Refresh wire sizing table since string counts may have changed
        self._refresh_wire_sizing_for_segments()

    def _on_module_selected(self, event=None):
        """Legacy — module is now derived from templates. Kept for backward compat."""
        pass

    def _on_inverter_selected(self, event=None):
        """Handle inverter selection from dropdown"""
        selected_name = self.inverter_select_var.get()
        if selected_name in self.available_inverters:
            self.selected_inverter = self.available_inverters[selected_name]
            inv = self.selected_inverter
            type_str = inv.inverter_type.value if hasattr(inv, 'inverter_type') else 'String'
            self.inverter_info_label.config(
                text=f"{inv.rated_power_kw}kW AC  |  {inv.max_dc_power_kw}kW DC  |  {inv.get_total_string_capacity()} inputs  |  {type_str}",
                foreground='black'
            )
            self._update_strings_per_inverter()
            self._on_inverter_changed_wire_sizing()
            # Add inverter to project's selected_inverters so export embeds it
            if self.current_project and inv:
                inv_id = f"{inv.manufacturer} {inv.model}"
                if inv_id not in self.current_project.selected_inverters:
                    self.current_project.selected_inverters.append(inv_id)
            # Auto-save when inverter changes (but not during load)
            if not getattr(self, '_loading', False):
                self._auto_unlock_allocation()
                self._mark_stale()
                self.save_estimate()
        else:
            self.selected_inverter = None
            self.inverter_info_label.config(text="No inverter selected", foreground='gray')
            self.strings_per_inverter_var.set('--')
            self.isc_warning_label.config(text="")

    INVERTER_COLORS = [
        '#4A90D9',  # Blue
        '#2ECC71',  # Green
        '#9B59B6',  # Purple
        '#E74C3C',  # Red
        '#1ABC9C',  # Teal
        '#3498DB',  # Light Blue
        '#E91E63',  # Pink
        '#00BCD4',  # Cyan
        '#8BC34A',  # Light Green
        '#795548',  # Brown
        '#607D8B',  # Blue Gray
        '#CDDC39',  # Lime
        '#5C6BC0',  # Indigo
        '#26A69A',  # Dark Teal
        '#78909C',  # Steel
    ]

    def show_site_preview(self):
        """Open the site preview in a pop-out window"""
        # Raise existing window instead of stacking another one
        if self._site_preview_window is not None:
            try:
                if self._site_preview_window.winfo_exists():
                    self._site_preview_window.deiconify()
                    self._site_preview_window.lift()
                    self._site_preview_window.focus_force()
                    return
            except Exception:
                pass
            self._site_preview_window = None

        inv_summary = getattr(self, 'last_totals', {}).get('inverter_summary', {})

        if not inv_summary or not inv_summary.get('allocation_result'):
            messagebox.showinfo("No Data", "Run Calculate Estimate first to generate preview data.")
            return

        topology = self.topology_var.get()
        row_spacing_ft = self.groups[0].get('row_spacing_ft', 20.0) if self.groups else 20.0

        # Compute device info for preview
        totals = getattr(self, 'last_totals', {})
        total_combiners = sum(totals.get('combiners_by_breaker', {}).values())

        lv_method = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
        if topology == 'Distributed String':
            num_devices = totals.get('string_inverters', 0)
            device_label = 'SI'
        elif topology == 'Central Inverter':
            # For Central Inverter, allocation groups = CBs
            alloc = totals.get('inverter_summary', {}).get('allocation_result', {})
            num_devices = alloc.get('summary', {}).get('total_inverters', total_combiners)
            device_label = 'LBD' if lv_method == 'Trunk Bus' else 'CB'
        else:
            num_devices = total_combiners
            device_label = 'LBD' if lv_method == 'Trunk Bus' else 'CB'

        # Restore inspect mode from previous session
        initial_inspect = getattr(self, '_last_inspect_mode', False)

        preview = SitePreviewWindow(
            self, inv_summary, topology, self.INVERTER_COLORS,
            self.groups, self.enabled_templates, row_spacing_ft,
            num_devices=num_devices, device_label=device_label,
            initial_inspect=initial_inspect, pads=self.pads,
            device_names=self.device_names,
            device_feeder_sizes=self.device_feeder_sizes,
            device_feeder_parallel_counts=self.device_feeder_parallel_counts,
            measurements=self.measurements
        )
        self._site_preview_window = preview

        # When window closes, save state back
        def _on_preview_close():
            self._last_inspect_mode = preview.inspect_mode
            self.pads = preview.pads  # Save pad positions back
            self.device_names = dict(preview.device_names)  # Save renamed devices back
            self.device_feeder_sizes = dict(preview.device_feeder_sizes)  # Save feeder sizes back
            self.device_feeder_parallel_counts = dict(preview.device_feeder_parallel_counts)  # Save parallel counts back
            self.measurements = list(preview.measurements)  # Save measurements back

            # If CB assignments were edited, refresh the estimate results
            if hasattr(self, 'last_combiner_assignments') and self.last_combiner_assignments:
                self._refresh_combiner_results_from_assignments()

            self._schedule_autosave()
            self._site_preview_window = None
            preview.destroy()
        preview.protocol("WM_DELETE_WINDOW", _on_preview_close)

    def _on_strings_per_inverter_changed(self, *args):
        """When user manually edits strings/device, reverse-calculate DC:AC ratio.
        
        For Central Inverter, this field is strings/CB — don't reverse-calc DC:AC.
        """
        if self._updating_spi:
            return
        if not self.selected_inverter or not self.selected_module:
            return
        
        topology = self.topology_var.get()
        spi = self._get_int_var(self.strings_per_inverter_var, 0)
        if spi <= 0:
            return
        
        if topology != 'Central Inverter':
            # Reverse-calculate DC:AC ratio from strings/device
            modules_per_string = self._get_int_var(self.modules_per_string_var, 28)
            module_wattage = self.selected_module.wattage
            actual_ratio = self.selected_inverter.dc_ac_ratio(spi, module_wattage, modules_per_string)
            
            if actual_ratio > 0:
                self._updating_spi = True
                self.dc_ac_ratio_var.set(f"{actual_ratio:.2f}")
                self._updating_spi = False
            
            # Update Isc warning for new string count
            module_isc = self.selected_module.isc
            total_isc = spi * module_isc
            max_isc = getattr(self.selected_inverter, 'max_short_circuit_current', None)
            
            if max_isc and total_isc > max_isc:
                self.isc_warning_label.config(
                    text=f"⚠️ Isc limit exceeded: {total_isc:.1f}A > {max_isc:.0f}A max"
                )
            else:
                self.isc_warning_label.config(text="")
        else:
            # Central Inverter: field is strings/CB, no DC:AC coupling
            self.isc_warning_label.config(text="")
        
        self._mark_dirty()

    def _update_distance_hints(self):
        """Update the hint text next to distance inputs based on topology."""
        if not hasattr(self, 'distance_hint_label'):
            return
        
        use_routed = self.use_routed_var.get() if hasattr(self, 'use_routed_var') else False
        
        
        if use_routed:
            if not self.pads:
                self.distance_hint_label.config(
                    text="⚠ No pads placed — add pads in Site Preview first",
                    foreground='orange'
                )
            else:
                num_pads = len(self.pads)
                self.distance_hint_label.config(
                    text=f"✓ Using routed distances from {num_pads} pad{'s' if num_pads > 1 else ''} (avg inputs ignored)",
                    foreground='green'
                )
            return
        
        topology = self.topology_var.get()
        if topology == 'Distributed String':
            self.distance_hint_label.config(text="(DC feeders N/A for distributed — AC homeruns are primary cable)", foreground='gray')
        elif topology == 'Central Inverter':
            self.distance_hint_label.config(text="(Long DC feeders to central pad — short AC from inverter)", foreground='gray')
        elif topology == 'Centralized String':
            self.distance_hint_label.config(text="(DC feeders to inverter bank — short AC from bank)", foreground='gray')
        else:
            self.distance_hint_label.config(text="", foreground='gray')

    def _get_estimate_tracker_dims_ft(self, template_ref):
        """Get (width_ft, length_ft) for a tracker from its template reference.
        
        Width = E-W dimension (across tracker row).
        Length = N-S dimension (along tracker).
        
        Returns (width_ft, length_ft) or None if template not found.
        """
        if not template_ref or template_ref not in self.enabled_templates:
            return None
        
        tdata = self.enabled_templates[template_ref]
        ms = tdata.get('module_spec', {})
        mps = tdata.get('modules_per_string', 28)
        strings_per_tracker = tdata.get('strings_per_tracker', 1)
        orientation = tdata.get('module_orientation', 'Portrait')
        module_spacing_m = tdata.get('module_spacing_m', 0.02)
        
        mod_w_m = ms.get('width_mm', 1134) / 1000.0
        mod_l_m = ms.get('length_mm', 2278) / 1000.0
        
        if orientation == 'Portrait':
            mod_across = mod_l_m   # E-W (width of tracker)
            mod_along = mod_w_m    # N-S (length of tracker)
        else:
            mod_across = mod_w_m
            mod_along = mod_l_m
        
        # Width (E-W): module across dimension × modules_high (stacked columns)
        modules_high = tdata.get('modules_high', 1)
        width_m = mod_across * modules_high
        
        # Length (N-S): all modules laid end-to-end (full strings + partial) plus gaps and motor
        full_spt = int(strings_per_tracker)
        partial_mods = round((strings_per_tracker - full_spt) * mps) if strings_per_tracker != full_spt else 0
        modules_in_row = full_spt * mps + partial_mods
        
        motor_gap_m = tdata.get('motor_gap_m', 0)
        has_motor = tdata.get('has_motor', True)
        if not has_motor:
            motor_gap_m = 0
        
        length_m = modules_in_row * mod_along + max(modules_in_row - 1, 0) * module_spacing_m + motor_gap_m

        width_ft = width_m * 3.28084
        length_ft = length_m * 3.28084
        
        return (width_ft, length_ft)

    def _get_harness_config_for_tracker_type(self, strings_per_tracker):
        """Find the harness config used for trackers with the given string count."""
        for group in self.groups:
            for seg in group['segments']:
                if seg['strings_per_tracker'] == strings_per_tracker and seg['quantity'] > 0:
                    return self.parse_harness_config(self._get_effective_harness_config(seg))
                
        # Fallback: single harness equal to full string count
        return [int(strings_per_tracker)]

    def _build_combiner_assignments(self, totals, topology):
        """Build structured combiner box assignments for Device Configurator.
        
        Produces self.last_combiner_assignments — a list of dicts, one per CB,
        each containing the tracker/harness connections that feed it.
        
        For Centralized String: 1 CB per inverter, connections from allocation harness_map.
        For Central Inverter: distribute trackers proportionally across N combiners.
        For Distributed String: no CBs — returns empty list.
        """
        self.last_combiner_assignments = []
        self.last_si_assignments = []

        if topology == 'Distributed String':
            self._build_si_assignments(totals)
            return
        
        inv_summary = totals.get('inverter_summary', {})
        allocation_result = inv_summary.get('allocation_result')
        module_isc = self.selected_module.isc if self.selected_module else 0
        
        try:
            breaker_size = int(self.breaker_size_var.get())
        except (ValueError, AttributeError):
            breaker_size = 400
        
        # NEC factor — use project setting if available
        nec_factor = 1.56
        if self.current_project:
            nec_factor = getattr(self.current_project, 'nec_safety_factor', 1.56)
        
        # Standard breaker sizes for per-CB calculation
        BREAKER_SIZES = [100, 125, 150, 175, 200, 225, 250, 300, 350, 400, 450, 500, 600, 700, 800]
        
        # Build flat tracker list with segment metadata for harness config lookup
        tracker_segment_map = []  # flat list: one entry per tracker with its segment ref
        for group in self.groups:
            for seg in group['segments']:
                harness_sizes = self._get_harness_sizes(seg)
                for _ in range(seg['quantity']):
                    tracker_segment_map.append({
                        'spt': seg['strings_per_tracker'],
                        'harness_sizes': list(harness_sizes),
                        'wire_gauge': self._get_wire_gauge_for_segment(seg, 'whip'),
                    })
        
        if topology == 'Centralized String' and allocation_result:
            # 1 CB per inverter — connections directly from allocation harness_map
            for inv_idx, inv in enumerate(allocation_result['inverters']):
                cb_name = self.device_names.get(inv_idx, f"CB-{inv_idx + 1:02d}")
                connections = self._build_connections_from_harness_map(
                    inv['harness_map'], tracker_segment_map, module_isc, nec_factor
                )
                # Calculate per-CB breaker from actual total current
                total_current = sum(
                    c['num_strings'] * c['module_isc'] * c['nec_factor']
                    for c in connections
                )
                calc_breaker = breaker_size  # fallback to global
                for bs in BREAKER_SIZES:
                    if bs >= total_current:
                        calc_breaker = bs
                        break
                
                self.last_combiner_assignments.append({
                    'combiner_name': cb_name,
                    'device_idx': inv_idx,
                    'breaker_size': calc_breaker,
                    'module_isc': module_isc,
                    'nec_factor': nec_factor,
                    'connections': connections,
                })
        
        elif topology == 'Central Inverter' and allocation_result:
            # Central Inverter: allocation groups = CBs (or LBDs for Trunk Bus).
            # Use allocation harness_map (same approach as Centralized String).
            _lv = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
            _dev_prefix = 'LBD' if _lv == 'Trunk Bus' else 'CB'
            for inv_idx, inv in enumerate(allocation_result['inverters']):
                cb_name = self.device_names.get(inv_idx, f"{_dev_prefix}-{inv_idx + 1:02d}")
                connections = self._build_connections_from_harness_map(
                    inv['harness_map'], tracker_segment_map, module_isc, nec_factor
                )
                # Calculate per-CB breaker from actual total current
                total_current = sum(
                    c['num_strings'] * c['module_isc'] * c['nec_factor']
                    for c in connections
                )
                calc_breaker = breaker_size  # fallback to global
                for bs in BREAKER_SIZES:
                    if bs >= total_current:
                        calc_breaker = bs
                        break
                
                self.last_combiner_assignments.append({
                    'combiner_name': cb_name,
                    'device_idx': inv_idx,
                    'breaker_size': calc_breaker,
                    'module_isc': module_isc,
                    'nec_factor': nec_factor,
                    'connections': connections,
                })

    def _build_si_assignments(self, totals):
        """Build self.last_si_assignments for Distributed String topology.

        Mirrors _build_combiner_assignments but produces one entry per string
        inverter rather than per combiner box. Each entry carries the same
        tracker/harness connection list format used by the Device Configurator.
        """
        self.last_si_assignments = []

        inv_summary = totals.get('inverter_summary', {})
        allocation_result = inv_summary.get('allocation_result')
        if not allocation_result:
            return

        module_isc = self.selected_module.isc if self.selected_module else 0

        nec_factor = 1.56
        if self.current_project:
            nec_factor = getattr(self.current_project, 'nec_safety_factor', 1.56)

        # Build flat tracker list for harness config / wire gauge lookup
        tracker_segment_map = []
        for group in self.groups:
            for seg in group['segments']:
                harness_sizes = self._get_harness_sizes(seg)
                for _ in range(seg['quantity']):
                    tracker_segment_map.append({
                        'spt': seg['strings_per_tracker'],
                        'harness_sizes': list(harness_sizes),
                        'wire_gauge': self._get_wire_gauge_for_segment(seg, 'whip'),
                    })

        for inv_idx, inv in enumerate(allocation_result.get('inverters', [])):
            inv_name = self.device_names.get(inv_idx, f"SI-{inv_idx + 1:02d}")
            connections = self._build_connections_from_harness_map(
                inv['harness_map'], tracker_segment_map, module_isc, nec_factor
            )
            self.last_si_assignments.append({
                'inverter_name': inv_name,
                'device_idx': inv_idx,
                'module_isc': module_isc,
                'nec_factor': nec_factor,
                'inverter_spec': self.selected_inverter,
                'connections': connections,
            })

    def _read_combiner_bom_from_device_config(self):
        """Read combiner BOM data from Device Configurator (single source of truth).

        Populates self.last_totals['combiners_by_breaker'] and ['combiner_details']
        from the Device Configurator's combiner_configs, which reflect the user's
        NEC multiplier, breaker overrides, and fuse sizes.

        Also syncs breaker sizes back into self.last_combiner_assignments.
        
        Returns True if Device Configurator had data, False if fallback is needed.
        """
        if not hasattr(self, 'last_totals') or not self.last_totals:
            return False
        
        # Get Device Configurator
        main_app = getattr(self, 'main_app', None)
        if not main_app or not hasattr(main_app, 'device_configurator'):
            return False
        
        dc = main_app.device_configurator
        if not hasattr(dc, 'combiner_configs') or not dc.combiner_configs:
            return False
        
        # Only use DC data if it's in QE mode
        if getattr(dc, 'data_source', 'blocks') != 'quick_estimate':
            return False
        
        totals = self.last_totals
        totals['combiners_by_breaker'] = {}
        totals['combiner_details'] = []
        
        # Sync breaker sizes back to last_combiner_assignments
        dc_breakers = {}
        for combiner_id, config in dc.combiner_configs.items():
            dc_breakers[combiner_id] = config.get_display_breaker_size()
        
        for cb in self.last_combiner_assignments:
            cb_name = cb.get('combiner_name', '')
            if cb_name in dc_breakers:
                cb['breaker_size'] = dc_breakers[cb_name]
        
        # Build totals from Device Configurator configs
        # Group by (breaker_size, fuse_holder_rating, max_inputs) for CB matching
        cb_groups = {}  # (breaker_size, fuse_holder_rating) -> {max_inputs, count}
        
        for combiner_id, config in dc.combiner_configs.items():
            if not config.connections:
                continue
            
            breaker_size = config.get_display_breaker_size()
            num_inputs = len(config.connections)
            
            # Get fuse holder rating from the DC's actual fuse sizes
            max_fuse = max(conn.get_display_fuse_size() for conn in config.connections)
            fuse_holder_rating = self.get_fuse_holder_category(max_fuse)
            
            # Count by breaker size
            if breaker_size not in totals['combiners_by_breaker']:
                totals['combiners_by_breaker'][breaker_size] = 0
            totals['combiners_by_breaker'][breaker_size] += 1
            
            # Group for CB part matching
            key = (breaker_size, fuse_holder_rating)
            if key not in cb_groups:
                cb_groups[key] = {'max_inputs': 0, 'count': 0}
            cb_groups[key]['max_inputs'] = max(cb_groups[key]['max_inputs'], num_inputs)
            cb_groups[key]['count'] += 1
        
        # Match CB parts for each group
        for (breaker_size, fuse_holder_rating), group_info in cb_groups.items():
            matched_cb = self.find_combiner_box(group_info['max_inputs'], breaker_size, fuse_holder_rating)
            if matched_cb:
                totals['combiner_details'].append({
                    'part_number': matched_cb.get('part_number', ''),
                    'description': matched_cb.get('description', ''),
                    'max_inputs': matched_cb.get('max_inputs', 0),
                    'breaker_size': breaker_size,
                    'fuse_holder_rating': fuse_holder_rating,
                    'strings_per_cb': group_info['max_inputs'],
                    'quantity': group_info['count'],
                    'block_name': 'Site Total'
                })
            else:
                totals['combiner_details'].append({
                    'part_number': 'NO MATCH',
                    'description': f'No CB found: {group_info["max_inputs"]} inputs, {breaker_size}A, {fuse_holder_rating}',
                    'max_inputs': group_info['max_inputs'],
                    'breaker_size': breaker_size,
                    'fuse_holder_rating': fuse_holder_rating,
                    'strings_per_cb': group_info['max_inputs'],
                    'quantity': group_info['count'],
                    'block_name': 'Site Total'
                })
        
        self.last_totals = totals
        return True

    def _rebuild_combiner_totals_from_assignments(self):
        """Fallback: rebuild combiner BOM totals from last_combiner_assignments.
        
        Used when the Device Configurator isn't available or isn't in QE mode.
        """
        if not self.last_combiner_assignments or not hasattr(self, 'last_totals') or not self.last_totals:
            return
        
        totals = self.last_totals
        totals['combiners_by_breaker'] = {}
        totals['combiner_details'] = []
        
        module_isc = self.selected_module.isc if self.selected_module else 0
        nec_factor = 1.56
        if self.current_project:
            nec_factor = getattr(self.current_project, 'nec_safety_factor', 1.56)
        
        for cb in self.last_combiner_assignments:
            bs = cb['breaker_size']
            if bs not in totals['combiners_by_breaker']:
                totals['combiners_by_breaker'][bs] = 0
            totals['combiners_by_breaker'][bs] += 1
        
        for bs, qty in totals['combiners_by_breaker'].items():
            max_h_strings = 1
            max_inputs = 0
            for cb in self.last_combiner_assignments:
                if cb['breaker_size'] == bs:
                    num_inputs = len(cb['connections'])
                    max_inputs = max(max_inputs, num_inputs)
                    for conn in cb['connections']:
                        max_h_strings = max(max_h_strings, conn['num_strings'])
            
            fuse_current = module_isc * nec_factor * max_h_strings
            fuse_holder_rating = self.get_fuse_holder_category(fuse_current)
            
            matched_cb = self.find_combiner_box(max_inputs, bs, fuse_holder_rating)
            if matched_cb:
                totals['combiner_details'].append({
                    'part_number': matched_cb.get('part_number', ''),
                    'description': matched_cb.get('description', ''),
                    'max_inputs': matched_cb.get('max_inputs', 0),
                    'breaker_size': bs,
                    'fuse_holder_rating': fuse_holder_rating,
                    'strings_per_cb': max_inputs,
                    'quantity': qty,
                    'block_name': 'Site Total'
                })
            else:
                totals['combiner_details'].append({
                    'part_number': 'NO MATCH',
                    'description': f'No CB found: {max_inputs} inputs, {bs}A, {fuse_holder_rating}',
                    'max_inputs': max_inputs,
                    'breaker_size': bs,
                    'fuse_holder_rating': fuse_holder_rating,
                    'strings_per_cb': max_inputs,
                    'quantity': qty,
                    'block_name': 'Site Total'
                })
        
        self.last_totals = totals

    def _refresh_combiner_results_from_assignments(self):
        """Rebuild combiner BOM totals and refresh display.
        
        Called after manual CB edits in the site preview to keep the estimate
        results in sync without re-running the full calculation.
        """
        if not self.last_combiner_assignments or not hasattr(self, 'last_totals') or not self.last_totals:
            return
        
        # Read from Device Configurator if available, else fallback
        if not self._read_combiner_bom_from_device_config():
            self._rebuild_combiner_totals_from_assignments()
        
        self._redraw_results_tree()

    def _redraw_results_tree(self):
        """Redraw the results treeview from self.last_totals without recalculating.
        
        This is a display-only refresh — it does NOT re-run calculate_estimate().
        """
        if not hasattr(self, 'last_totals') or not self.last_totals:
            return
        
        totals = self.last_totals
        
        # Clear existing results
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        self.checked_items.clear()
        
        def insert_section(label):
            self.results_tree.insert('', 'end', values=('', f'--- {label} ---', '', '', '', '', '', ''), tags=('section',))

        def insert_row(item, part_number, qty, unit, unit_cost='', ext_cost='', description=''):
            iid = self.results_tree.insert('', 'end', values=('☑', item, part_number, description, qty, unit, unit_cost, ext_cost), tags=('checked',))
            self.checked_items.add(iid)
        
        # Combiner Boxes
        if totals.get('combiners_by_breaker'):
            insert_section('COMBINER BOXES')
            total_cbs = 0
            for breaker_size in sorted(totals['combiners_by_breaker'].keys()):
                qty = totals['combiners_by_breaker'][breaker_size]
                total_cbs += qty
                insert_row(f"Combiner Box ({breaker_size}A breaker)", '', qty, 'ea')
            if len(totals['combiners_by_breaker']) > 1:
                insert_row('Total Combiner Boxes', '', total_cbs, 'ea')
            
            for detail in totals.get('combiner_details', []):
                if detail['part_number'] != 'NO MATCH':
                    insert_row(
                        f"  └ Site Total: ({detail['max_inputs']}-input, {detail['fuse_holder_rating']})",
                        detail['part_number'], detail['quantity'], 'ea',
                        description=self._lookup_description(detail['part_number'])
                    )
                else:
                    insert_row(
                        f"  └ Site Total: ⚠ {detail['description']}",
                        '', detail['quantity'], 'ea'
                    )
        
        # Inverters
        if totals.get('string_inverters', 0) > 0:
            insert_section('INVERTERS')
            inv_summary = totals.get('inverter_summary', {})
            actual_ratio = inv_summary.get('actual_dc_ac', 0)
            inv_name = "Inverter"
            if self.selected_inverter:
                inv_name = f"{self.selected_inverter.manufacturer} {self.selected_inverter.model}"
            insert_row(f"{inv_name} (DC:AC {actual_ratio:.2f})", '', totals['string_inverters'], 'ea')
        
        # Harnesses
        if totals.get('harnesses_by_size'):
            insert_section('HARNESSES')
            for size in sorted(totals['harnesses_by_size'].keys(), reverse=True):
                qty = totals['harnesses_by_size'][size]
                pos_pn, pos_unit, pos_ext = self.lookup_part_and_price('harness', num_strings=size, polarity='positive', qty=qty)
                neg_pn, neg_unit, neg_ext = self.lookup_part_and_price('harness', num_strings=size, polarity='negative', qty=qty)
                pos_desc = self._lookup_description(pos_pn)
                neg_desc = self._lookup_description(neg_pn)
                if size == 1:
                    iid = self.results_tree.insert('', 'end', values=('☐', f"{size}-String Harness (Pos, CU)", pos_pn, pos_desc, qty, 'ea', pos_unit, pos_ext), tags=('unchecked',))
                    iid2 = self.results_tree.insert('', 'end', values=('☐', f"{size}-String Harness (Neg, CU)", neg_pn, neg_desc, qty, 'ea', neg_unit, neg_ext), tags=('unchecked',))
                else:
                    insert_row(f"{size}-String Harness (Pos, CU)", pos_pn, qty, 'ea', pos_unit, pos_ext, description=pos_desc)
                    insert_row(f"{size}-String Harness (Neg, CU)", neg_pn, qty, 'ea', neg_unit, neg_ext, description=neg_desc)
        
        # Inline DC string fuses (Wire Harness only)
        if totals.get('inline_fuses_by_rating'):
            import os as _os
            is_first_solar = totals.get('has_first_solar', False)

            fuse_library = {}
            try:
                _cur = _os.path.dirname(_os.path.abspath(__file__))
                _root = _os.path.dirname(_os.path.dirname(_cur))
                with open(_os.path.join(_root, 'data', 'fuse_library.json'), 'r') as _f:
                    fuse_library = json.load(_f)
            except Exception:
                pass

            def _fuse_pn(rating):
                for pn, spec in fuse_library.items():
                    if spec.get('fuse_rating_amps') == rating:
                        return pn
                candidates = [(spec.get('fuse_rating_amps', 0), pn)
                              for pn, spec in fuse_library.items()
                              if spec.get('fuse_rating_amps', 0) >= rating]
                return min(candidates)[1] if candidates else 'N/A'

            insert_section('FUSES')
            for rating in sorted(totals['inline_fuses_by_rating'].keys()):
                qty = totals['inline_fuses_by_rating'][rating]
                pn = _fuse_pn(rating)
                try:
                    from src.utils.pricing_lookup import PricingLookup as _PL
                    _up = _PL().get_price(pn) if pn != 'N/A' else None
                    unit_cost_str = f"${_up:,.2f}" if _up else ''
                    ext_cost_str = f"${_up * qty:,.2f}" if _up else ''
                except Exception:
                    unit_cost_str = ''
                    ext_cost_str = ''
                fuse_desc = self._lookup_description(pn)
                item_name = f"{rating}A Inline DC String Fuse (Pos)"
                if is_first_solar:
                    self.results_tree.insert('', 'end',
                        values=('☐', item_name, pn, fuse_desc, qty, 'ea', unit_cost_str, ext_cost_str),
                        tags=('unchecked',))
                else:
                    insert_row(item_name, pn, qty, 'ea', unit_cost_str, ext_cost_str, description=fuse_desc)

        # Trunk Bus items
        if totals.get('trunk_cable_by_size'):
            insert_section('TRUNK BUS CABLE')
            for size in sorted(totals['trunk_cable_by_size'].keys()):
                total_ft = totals['trunk_cable_by_size'][size]
                insert_row(f"Trunk Bus {size} (Pos)", '', f"{total_ft:.0f}", 'ft')
                insert_row(f"Trunk Bus {size} (Neg)", '', f"{total_ft:.0f}", 'ft')
        
        if totals.get('lbd_by_size'):
            insert_section('LOAD BREAK DISCONNECTS')
            for rating in sorted(totals['lbd_by_size'].keys()):
                qty = totals['lbd_by_size'][rating]
                insert_row(f"LBD ({rating}A)", '', qty, 'ea')
        
        if totals.get('ipc_by_tap'):
            insert_section('INSULATION PIERCING CONNECTORS')
            for tap_count in sorted(totals['ipc_by_tap'].keys()):
                qty = totals['ipc_by_tap'][tap_count]
                insert_row(f"{tap_count}-Tap IPC", '', qty, 'ea')
        
        # Extenders — split by wire gauge, then by polarity
        ext_gauges = sorted(set(g for (_, g) in totals['extenders_pos_by_length'].keys()) |
                           set(g for (_, g) in totals['extenders_neg_by_length'].keys()))
        for gauge in ext_gauges:
            # Positive
            pos_items = {length: qty for (length, g), qty in totals['extenders_pos_by_length'].items() if g == gauge}
            if pos_items:
                insert_section(f'EXTENDERS — POSITIVE ({gauge}, CU)')
                for length in sorted(pos_items.keys()):
                    qty = pos_items[length]
                    e_pn, e_unit, e_ext = self.lookup_part_and_price('extender', polarity='positive', length_ft=length, qty=qty, num_strings=self._gauge_to_string_count('extender', gauge))
                    insert_row(f"Extender {length}ft (Pos)", e_pn, qty, 'ea', e_unit, e_ext, description=self._lookup_description(e_pn))

            # Negative
            neg_items = {length: qty for (length, g), qty in totals['extenders_neg_by_length'].items() if g == gauge}
            if neg_items:
                insert_section(f'EXTENDERS — NEGATIVE ({gauge}, CU)')
                for length in sorted(neg_items.keys()):
                    qty = neg_items[length]
                    e_pn, e_unit, e_ext = self.lookup_part_and_price('extender', polarity='negative', length_ft=length, qty=qty, num_strings=self._gauge_to_string_count('extender', gauge))
                    insert_row(f"Extender {length}ft (Neg)", e_pn, qty, 'ea', e_unit, e_ext, description=self._lookup_description(e_pn))

        # Whips — split by wire gauge, then by polarity
        whip_gauges = sorted(set(g for (_, g) in totals.get('whips_by_length', {}).keys()))
        if whip_gauges:
            display = totals.get('_display', {})
            topology = display.get('topology', self.topology_var.get() if hasattr(self, 'topology_var') else '')
            device_label = 'to inverter' if topology == 'Distributed String' else 'to combiner'

            for gauge in whip_gauges:
                gauge_items = {length: qty for (length, g), qty in totals['whips_by_length'].items() if g == gauge}
                if not gauge_items:
                    continue

                # Positive whips (half of each length bucket)
                insert_section(f'WHIPS — POSITIVE ({device_label}) ({gauge}, CU)')
                for length in sorted(gauge_items.keys()):
                    qty = gauge_items[length] // 2
                    w_pn, w_unit, w_ext = self.lookup_part_and_price('whip', polarity='positive', length_ft=length, qty=qty, num_strings=self._gauge_to_string_count('whip', gauge))
                    insert_row(f"Whip {length}ft (Pos)", w_pn, qty, 'ea', w_unit, w_ext, description=self._lookup_description(w_pn))

                # Negative whips
                insert_section(f'WHIPS — NEGATIVE ({device_label}) ({gauge}, CU)')
                for length in sorted(gauge_items.keys()):
                    qty = gauge_items[length] // 2
                    w_pn, w_unit, w_ext = self.lookup_part_and_price('whip', polarity='negative', length_ft=length, qty=qty, num_strings=self._gauge_to_string_count('whip', gauge))
                    insert_row(f"Whip {length}ft (Neg)", w_pn, qty, 'ea', w_unit, w_ext, description=self._lookup_description(w_pn))
        
        # DC Feeders / AC Homeruns — per-device sizes grouped
        display = totals.get('_display', {})
        use_routed = display.get('use_routed', False)
        topo = display.get('topology', 'Distributed String')
        label_suffix = " (routed)" if use_routed else ""
        feeders_by_size = totals.get('feeders_by_size', {})
        feeder_mat = 'AL' if self.wire_sizing.get('feeder_material', 'aluminum') == 'aluminum' else 'CU'

        def _unpack_feeder_key(k):
            """Accept either tuple (size, parallel) or legacy string size; return (size, parallel)."""
            if isinstance(k, tuple):
                return k[0], int(k[1]) if len(k) > 1 else 1
            return k, 1

        if topo == 'Distributed String':
            # Primary cable is AC homerun (per-device sizes)
            if feeders_by_size:
                insert_section('AC HOMERUNS')
                for raw_key, data in sorted(feeders_by_size.items(), key=lambda kv: _unpack_feeder_key(kv[0])):
                    wire_size, parallel = _unpack_feeder_key(raw_key)
                    count = data['count']
                    total_ft = data['total_ft']
                    # avg distance uses raw (unmultiplied) distance so the label stays intuitive
                    dist_ft = data.get('distance_ft', total_ft / max(parallel, 1))
                    avg_ft = dist_ft / count if count > 0 else 0
                    parallel_suffix = f" ×{parallel} parallel" if parallel > 1 else ""
                    insert_row(
                        f"AC Homerun {wire_size}, {feeder_mat} — avg {avg_ft:.0f}ft{label_suffix} × {count} runs{parallel_suffix}",
                        '', f"{total_ft:.0f}", 'ft'
                    )
        else:
            # Primary cable is DC feeder (per-device sizes)
            if feeders_by_size:
                insert_section('DC FEEDERS')
                for raw_key, data in sorted(feeders_by_size.items(), key=lambda kv: _unpack_feeder_key(kv[0])):
                    wire_size, parallel = _unpack_feeder_key(raw_key)
                    count = data['count']
                    total_ft = data['total_ft']
                    dist_ft = data.get('distance_ft', total_ft / max(parallel, 1))
                    avg_ft = dist_ft / count if count > 0 else 0
                    parallel_suffix = f" ×{parallel} parallel" if parallel > 1 else ""
                    insert_row(
                        f"DC Feeder {wire_size}, {feeder_mat} — avg {avg_ft:.0f}ft{label_suffix} × {count} runs{parallel_suffix} (pos)",
                        '', f"{total_ft:.0f}", 'ft'
                    )
                    insert_row(
                        f"DC Feeder {wire_size}, {feeder_mat} — avg {avg_ft:.0f}ft{label_suffix} × {count} runs{parallel_suffix} (neg)",
                        '', f"{total_ft:.0f}", 'ft'
                    )

            # Secondary AC homeruns (blanket size, not per-device)
            if totals.get('ac_homerun_count', 0) > 0:
                ac_wire_size = self.get_wire_size_for('ac_homerun')
                ac_count = totals['ac_homerun_count']
                ac_total = totals['ac_homerun_total_ft']
                ac_avg = ac_total / ac_count if ac_count > 0 else 0
                insert_section(f'AC HOMERUNS ({ac_wire_size}, {feeder_mat})')
                insert_row(f"AC Homerun {ac_wire_size}, {feeder_mat} — avg {ac_avg:.0f}ft × {ac_count} runs", '', f"{ac_total:.0f}", 'ft')
    
    def _build_connections_from_harness_map(self, harness_map, tracker_segment_map, module_isc, nec_factor):
        """Convert an allocation harness_map into Device Configurator connection dicts."""
        connections = []
        
        # Group harness_map entries by tracker to assign harness labels
        tracker_harness_counter = {}  # tracker_idx -> next harness number
        
        for entry in harness_map:
            tidx = entry['tracker_idx']
            strings_taken = entry['strings_taken']
            spt = entry['strings_per_tracker']
            is_split = entry.get('is_split', False)
            
            # Get segment info for this tracker
            t_info = tracker_segment_map[tidx] if tidx < len(tracker_segment_map) else None
            wire_gauge = t_info['wire_gauge'] if t_info else '10 AWG'
            original_harnesses = t_info['harness_sizes'] if t_info else [spt]

            if not is_split:
                # Full tracker — one connection per harness in the config
                pos_cursor = 0
                for h_idx, h_size in enumerate(original_harnesses):
                    connections.append({
                        'tracker_idx': tidx,
                        'tracker_label': f"T{tidx + 1:02d}",
                        'harness_label': f"H{h_idx + 1:02d}",
                        'num_strings': h_size,
                        'start_string_pos': pos_cursor,
                        'module_isc': module_isc,
                        'nec_factor': nec_factor,
                        'wire_gauge': self.get_wire_size_for('whip', h_size),
                    })
                    pos_cursor += h_size
            else:
                # Split tracker — distribute strings_taken across harnesses
                remaining = strings_taken
                harness_cursor = tracker_harness_counter.get(tidx, 0)
                # Compute physical start position from split_position.
                # Head starts at 0, tail starts at (spt - strings_taken).
                # This is correct for fresh allocations where splits are always head/tail.
                split_pos = entry.get('split_position', 'head')
                if split_pos == 'tail':
                    pos_cursor = spt - strings_taken
                else:
                    pos_cursor = 0
                
                while remaining > 0 and harness_cursor < len(original_harnesses):
                    h_size = original_harnesses[harness_cursor]
                    take = min(remaining, h_size)
                    connections.append({
                        'tracker_idx': tidx,
                        'tracker_label': f"T{tidx + 1:02d}",
                        'harness_label': f"H{harness_cursor + 1:02d}",
                        'num_strings': take,
                        'start_string_pos': pos_cursor,
                        'module_isc': module_isc,
                        'nec_factor': nec_factor,
                        'wire_gauge': self.get_wire_size_for('whip', take),
                    })
                    pos_cursor += take
                    remaining -= take
                    if take >= h_size:
                        harness_cursor += 1
                
                tracker_harness_counter[tidx] = harness_cursor
        
        return connections
    
    def _get_wire_gauge_for_segment(self, seg, cable_type):
        """Get the wire gauge for a segment from the wire sizing table."""
        spt = seg.get('strings_per_tracker', 1)
        harness_sizes = self._get_harness_sizes(seg)
        max_harness = max(harness_sizes) if harness_sizes else spt
        
        # Look up from self.wire_sizing
        if hasattr(self, 'wire_sizing') and self.wire_sizing:
            # Wire sizing is keyed by string count
            sizing = self.wire_sizing.get(str(max_harness)) or self.wire_sizing.get(str(spt))
            if sizing and cable_type in sizing:
                return sizing[cable_type]
        
        return '10 AWG'
    
    def _derive_harnesses_for_split(self, harness_config_sizes, spt, device_positions):
        """Derive per-device harness configs from a segment's harness config and split boundaries.
        
        Walks the harness config north-to-south (physical position order) and
        cuts any harness that straddles a device boundary.
        
        Args:
            harness_config_sizes: list of ints, e.g. [6, 6] for "6+6"
            spt: total strings per tracker
            device_positions: list of dicts, each with:
                - 'inv_idx': device index
                - 'strings_taken': how many strings this device owns
                - 'start_pos': first physical position owned (0-based, north)
                
        Returns:
            list of dicts (same order as device_positions), each with:
                - 'inv_idx': device index
                - 'strings_taken': total strings
                - 'harnesses': list of ints (harness sizes for this portion)
        """
        # Build the harness ranges: which physical positions each harness covers
        harness_ranges = []
        pos_cursor = 0
        for h_size in harness_config_sizes:
            harness_ranges.append((pos_cursor, pos_cursor + h_size - 1))  # inclusive
            pos_cursor += h_size
        
        # Sort device portions by start_pos (north first)
        sorted_portions = sorted(device_positions, key=lambda p: p['start_pos'])
        
        result = []
        for portion in sorted_portions:
            p_start = portion['start_pos']
            p_end = p_start + portion['strings_taken'] - 1  # inclusive
            
            portion_harnesses = []
            for h_start, h_end in harness_ranges:
                # Find overlap between this harness range and this device's range
                overlap_start = max(h_start, p_start)
                overlap_end = min(h_end, p_end)
                overlap = overlap_end - overlap_start + 1
                
                if overlap > 0:
                    portion_harnesses.append(overlap)
            
            result.append({
                'inv_idx': portion['inv_idx'],
                'strings_taken': portion['strings_taken'],
                'start_pos': portion['start_pos'],
                'harnesses': portion_harnesses,
            })
        
        return result

    def _calc_inline_fuse_rating(self, module_spec) -> int:
        """Return the inline DC string fuse rating for a single string.
        Uses fixed 1.25 NEC factor regardless of project setting."""
        _FUSE_RATINGS = [5, 10, 15, 20, 25, 30, 35, 40, 45]
        try:
            if isinstance(module_spec, dict):
                isc = float(module_spec.get('isc', 0))
            elif module_spec is not None:
                isc = float(module_spec.isc)
            elif self.selected_module:
                isc = float(self.selected_module.isc)
            else:
                return 15
            nec_min = isc * 1.25
            for rating in _FUSE_RATINGS:
                if rating >= nec_min:
                    return rating
        except (ValueError, TypeError):
            pass
        return 15

    def _adjust_harnesses_for_splits(self, totals):
        """Derive harness configs for split trackers based on allocation boundaries.
        
        Walks the segment's harness config north-to-south and cuts at the split
        boundary. Harnesses cannot span device boundaries.
        
        Also builds self._split_tracker_details for use by whip and extender
        calculations — maps each split tracker to per-device harness assignments.
        """
        self._split_tracker_details = {}
        
        inv_summary = totals.get('inverter_summary', {})
        allocation_result = inv_summary.get('allocation_result')
        
        if not allocation_result:
            return
        
        # Collect split tracker info from harness_map
        split_trackers = {}
        
        for inv_idx, inv in enumerate(allocation_result['inverters']):
            for entry in inv['harness_map']:
                if entry['is_split']:
                    tidx = entry['tracker_idx']
                    if tidx not in split_trackers:
                        split_trackers[tidx] = []
                    entry_with_inv = dict(entry)
                    entry_with_inv['inv_idx'] = inv_idx
                    split_trackers[tidx].append(entry_with_inv)
        
        if not split_trackers:
            return
        
        # Build tracker_to_segment map for looking up harness configs
        tracker_seg_map = getattr(self, '_tracker_to_segment', [])
        
        for tidx, entries in split_trackers.items():
            spt = entries[0]['strings_per_tracker']
            # Use per-tracker segment lookup to get correct harness config
            # (multiple segments can share the same SPT with different configs)
            if tidx < len(tracker_seg_map) and tracker_seg_map[tidx]:
                seg = tracker_seg_map[tidx]['seg']
                harness_config = self._get_effective_harness_config(seg)
                original_harness_sizes = self.parse_harness_config(harness_config)
            else:
                original_harness_sizes = self._get_harness_config_for_tracker_type(spt)
            
            # Build device_positions with physical start positions
            # Entries come from harness_map which is ordered by allocation sequence.
            # For head/tail splits: head starts at 0, tail starts after head.
            device_positions = []
            pos_cursor = 0
            
            # Sort by split_position: head first, then middle, then tail
            position_order = {'head': 0, 'middle': 1, 'tail': 2, 'full': 0}
            sorted_entries = sorted(entries, key=lambda e: position_order.get(e.get('split_position', 'full'), 0))
            
            # Check if we have start_string_pos from Edit Devices (most accurate)
            # This comes through the combiner assignments path
            has_physical_pos = False
            if self.last_combiner_assignments:
                for cb in self.last_combiner_assignments:
                    for conn in cb.get('connections', []):
                        if conn['tracker_idx'] == tidx and 'start_string_pos' in conn:
                            has_physical_pos = True
                            break
                    if has_physical_pos:
                        break
            
            if has_physical_pos:
                # Build from combiner assignments with physical positions
                for cb_idx, cb in enumerate(self.last_combiner_assignments):
                    for conn in cb.get('connections', []):
                        if conn['tracker_idx'] == tidx:
                            device_positions.append({
                                'inv_idx': cb_idx,
                                'strings_taken': conn['num_strings'],
                                'start_pos': conn.get('start_string_pos', 0),
                            })
            elif any('start_physical_pos' in e for e in entries):
                # Use physical positions from harness_map (from manual Edit Devices)
                for entry in entries:
                    device_positions.append({
                        'inv_idx': entry['inv_idx'],
                        'strings_taken': entry['strings_taken'],
                        'start_pos': entry['start_physical_pos'],
                    })
            else:
                # Fall back to allocation order (head=north, tail=south)
                for entry in sorted_entries:
                    device_positions.append({
                        'inv_idx': entry['inv_idx'],
                        'strings_taken': entry['strings_taken'],
                        'start_pos': pos_cursor,
                    })
                    pos_cursor += entry['strings_taken']
            # Derive harnesses for each portion
            portion_details = self._derive_harnesses_for_split(
                original_harness_sizes, spt, device_positions
            )
            
            # Store for whip and extender calculations
            self._split_tracker_details[tidx] = {
                'spt': spt,
                'original_config': original_harness_sizes,
                'portions': portion_details,
            }
            
            # Adjust harness totals: remove original, add derived
            for size in original_harness_sizes:
                if size in totals['harnesses_by_size']:
                    totals['harnesses_by_size'][size] -= 1
                    if totals['harnesses_by_size'][size] <= 0:
                        del totals['harnesses_by_size'][size]
            
            all_new = []
            for p in portion_details:
                all_new.extend(p['harnesses'])
            for size in all_new:
                if size not in totals['harnesses_by_size']:
                    totals['harnesses_by_size'][size] = 0
                totals['harnesses_by_size'][size] += 1

    def _update_spi_label(self):
        """Update the Strings/Device label and show/hide central inverter count."""
        if not hasattr(self, 'spi_label'):
            return
        topology = self.topology_var.get()
        
        if topology == 'Distributed String':
            self.spi_label.config(text="Strings/SI:")
        elif topology == 'Centralized String':
            self.spi_label.config(text="Strings/CB:")
        else:  # Central Inverter
            self.spi_label.config(text="Strings/CB:")
        
        # Show/hide central inverter count field
        if hasattr(self, 'central_inv_count_label'):
            if topology == 'Central Inverter':
                self.central_inv_count_label.pack(side='left', padx=(10, 5))
                self.central_inv_count_spinbox.pack(side='left', padx=(0, 15))
            else:
                self.central_inv_count_label.pack_forget()
                self.central_inv_count_spinbox.pack_forget()

    def _update_strings_per_inverter(self):
        """Auto-calculate strings per inverter from DC:AC ratio and show Isc warning if needed"""
        if self._updating_spi:
            return
        # For Central Inverter, this field is strings/CB — managed by the user and
        # by calculate_estimate. Don't touch it here, including the early-return
        # resets below that would clobber the user's input.
        if self.topology_var.get() == 'Central Inverter':
            self.isc_warning_label.config(text="")
            return
        if not self.selected_inverter or not self.selected_module:
            self.strings_per_inverter_var.set('--')
            self.isc_warning_label.config(text="")
            return
        
        target_ratio = self._get_float_var(self.dc_ac_ratio_var, 0.0)
        if target_ratio <= 0:
            self.strings_per_inverter_var.set('--')
            return
        
        modules_per_string = self._get_int_var(self.modules_per_string_var, 28)
        
        module_wattage = self.selected_module.wattage
        string_power_kw = (module_wattage * modules_per_string) / 1000
        
        if string_power_kw <= 0:
            self.strings_per_inverter_var.set('--')
            return
        
        topology = self.topology_var.get()
        
        # Calculate power-based string count (always applies)
        target_dc_kw = target_ratio * self.selected_inverter.rated_power_kw
        power_based_strings = round(target_dc_kw / string_power_kw)
        
        # For Central Inverter, the field means "strings per CB" — don't overwrite it
        # from DC:AC ratio. DC:AC drives central inverter count separately.
        if topology == 'Central Inverter':
            strings_per_inv = power_based_strings
            strings_per_inv = max(strings_per_inv, 1)
            # Don't write to strings_per_inverter_var — it holds strings/CB
            # Still calculate actual_ratio for display purposes below
        else:
            strings_per_inv = power_based_strings
            strings_per_inv = max(strings_per_inv, 1)
            self._updating_spi = True
            self.strings_per_inverter_var.set(str(strings_per_inv))
            self._updating_spi = False
        
        # Calculate actual DC:AC ratio achieved
        actual_ratio = self.selected_inverter.dc_ac_ratio(
            strings_per_inv, module_wattage, modules_per_string
        )
        

        # Show warning if input limit is capping the ratio (Distributed String only)
        if topology == 'Distributed String' and strings_per_inv < power_based_strings:
            self.inverter_info_label.config(
                text=f"{self.selected_inverter.rated_power_kw}kW AC  |  Capped by string inputs ({self.selected_inverter.get_total_string_capacity()})  |  Max DC:AC ≈ {actual_ratio:.2f}",
                foreground='orange'
            )
        else:
            # No capping — reset to normal inverter info
            inv = self.selected_inverter
            type_str = inv.inverter_type.value if hasattr(inv, 'inverter_type') else 'String'
            self.inverter_info_label.config(
                text=f"{inv.rated_power_kw}kW AC  |  {inv.max_dc_power_kw}kW DC  |  {inv.get_total_string_capacity()} inputs  |  {type_str}",
                foreground='black'
            )
        
        # Check Isc hard limit
        module_isc = self.selected_module.isc
        total_isc = strings_per_inv * module_isc
        max_isc = getattr(self.selected_inverter, 'max_short_circuit_current', None)
        
        if max_isc and total_isc > max_isc:
            self.isc_warning_label.config(
                text=f"⚠️ Isc limit exceeded: {total_isc:.1f}A > {max_isc:.0f}A max"
            )
        else:
            self.isc_warning_label.config(text="")


    # ==================== UI Setup ====================
    
    def setup_ui(self):
        """Create and arrange UI components"""
        # Main container with padding
        main_container = ttk.Frame(self, padding="10")
        main_container.pack(fill='both', expand=True)
        
        # Top bar: Title + Estimate selector
        top_bar = ttk.Frame(main_container)
        top_bar.pack(fill='x', pady=(0, 10))
        
        # Title
        title_label = ttk.Label(top_bar, text="Quick Estimate", font=('Helvetica', 14, 'bold'))
        title_label.pack(side='left', padx=(0, 20))
        
        # Estimate selector
        ttk.Label(top_bar, text="Estimate:").pack(side='left', padx=(0, 5))
        self.estimate_var = tk.StringVar()
        self.estimate_combo = ttk.Combobox(
            top_bar,
            textvariable=self.estimate_var,
            width=30,
            state='normal'
        )
        self.estimate_combo.pack(side='left', padx=(0, 5))
        self.estimate_combo.bind('<<ComboboxSelected>>', self._on_estimate_selected)
        self.estimate_combo.bind('<Return>', self._on_estimate_name_changed)
        self.estimate_combo.bind('<FocusOut>', self._on_estimate_name_changed)
        
        ttk.Button(top_bar, text="New", command=self.new_estimate, width=6).pack(side='left', padx=2)
        ttk.Button(top_bar, text="Copy", command=self.copy_estimate, width=6).pack(side='left', padx=2)
        ttk.Button(top_bar, text="Delete", command=self.delete_estimate, width=6).pack(side='left', padx=2)
        
        # Description
        desc_label = ttk.Label(main_container, text="Early-stage BOM estimation for bid and preliminary designs", foreground='gray')
        desc_label.pack(anchor='w', pady=(0, 10))
        
        # Main content area - PanedWindow for resizable split
        paned = ttk.PanedWindow(main_container, orient='horizontal')
        paned.pack(fill='both', expand=True, pady=(0, 10))
        
        # Left panel - Tree view
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)
        
        # Right panel - Details
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=3)
        
        self.setup_tree_panel(left_frame)
        self.setup_details_panel(right_frame)
        
        # Bottom section - Calculate button and results
        bottom_frame = ttk.Frame(main_container)
        bottom_frame.pack(fill='both', expand=True)
        
        # Global settings + Wire sizing side-by-side container
        settings_container = ttk.Frame(bottom_frame)
        settings_container.pack(fill='x', pady=(0, 10))
        
        settings_frame = ttk.LabelFrame(settings_container, text="Global Settings", padding="5")
        settings_frame.pack(side='left', fill='y')
        
        # Row 1: Module display (derived from templates)
        module_row = ttk.Frame(settings_frame)
        module_row.pack(fill='x', pady=(0, 5))
        
        ttk.Label(module_row, text="Module:", font=('Helvetica', 9, 'bold')).pack(side='left', padx=(0, 5))
        self.module_info_label = ttk.Label(module_row, text="No templates linked — link a tracker template to a segment", foreground='gray')
        self.module_info_label.pack(side='left', padx=(5, 0))
        
        # Row 2: Inverter selection
        inverter_row = ttk.Frame(settings_frame)
        inverter_row.pack(fill='x', pady=(0, 5))
        
        ttk.Label(inverter_row, text="Inverter:").pack(side='left', padx=(0, 5))
        self.inverter_select_var = tk.StringVar()
        self.inverter_combo = ttk.Combobox(
            inverter_row,
            textvariable=self.inverter_select_var,
            values=sorted(self.available_inverters.keys()),
            state='readonly',
            width=50
        )
        self.inverter_combo.pack(side='left', padx=(0, 15))
        self.inverter_combo.bind('<<ComboboxSelected>>', self._on_inverter_selected)
        self.disable_combobox_scroll(self.inverter_combo)
        
        # Inverter info display
        self.inverter_info_label = ttk.Label(inverter_row, text="No inverter selected", foreground='gray')
        self.inverter_info_label.pack(side='left', padx=(5, 0))
        
        # Row 3: Topology and DC:AC ratio
        topology_row = ttk.Frame(settings_frame)
        topology_row.pack(fill='x', pady=(0, 5))
        
        ttk.Label(topology_row, text="Topology:").pack(side='left', padx=(0, 5))
        self.topology_var = tk.StringVar(value='Distributed String')
        topology_combo = ttk.Combobox(
            topology_row,
            textvariable=self.topology_var,
            values=['Distributed String', 'Centralized String', 'Central Inverter'],
            state='readonly',
            width=20
        )
        topology_combo.pack(side='left', padx=(0, 15))
        self.disable_combobox_scroll(topology_combo)
        self.topology_var.trace_add('write', lambda *args: (self._auto_unlock_allocation(), self._update_spi_label(), self._update_distance_hints(), self._on_topology_changed_wire_sizing(), self._mark_stale(), self._schedule_autosave()))
        
        ttk.Label(topology_row, text="DC:AC Ratio:").pack(side='left', padx=(0, 5))
        self.dc_ac_ratio_var = tk.StringVar(value='1.25')
        ttk.Spinbox(
            topology_row, from_=1.0, to=2.0, increment=0.05,
            textvariable=self.dc_ac_ratio_var, width=6, format='%.2f'
        ).pack(side='left', padx=(0, 15))
        self.dc_ac_ratio_var.trace_add('write', lambda *args: (self._auto_unlock_allocation(), self._update_strings_per_inverter(), self._mark_stale(), self._schedule_autosave()))
        
        ttk.Label(topology_row, text="LV Collection:").pack(side='left', padx=(0, 5))
        self.lv_collection_var = tk.StringVar(value='Wire Harness')
        lv_collection_combo = ttk.Combobox(
            topology_row,
            textvariable=self.lv_collection_var,
            values=['String HR', 'Wire Harness', 'Trunk Bus'],
            state='readonly',
            width=14
        )
        lv_collection_combo.pack(side='left', padx=(0, 15))
        self.disable_combobox_scroll(lv_collection_combo)
        self.lv_collection_var.trace_add('write', lambda *args: (self._auto_unlock_allocation(), self._on_lv_collection_changed(), self._mark_stale(), self._schedule_autosave()))

        ttk.Label(topology_row, text="Breaker Size:").pack(side='left', padx=(0, 5))
        self.breaker_size_var = tk.StringVar(value='400')
        breaker_combo = ttk.Combobox(
            topology_row,
            textvariable=self.breaker_size_var,
            values=['200', '300', '400', '600', '800'],
            state='readonly',
            width=6
        )
        breaker_combo.pack(side='left', padx=(0, 15))
        self.disable_combobox_scroll(breaker_combo)
        self.breaker_size_var.trace_add('write', lambda *args: (self._on_breaker_changed_wire_sizing(), self._mark_stale(), self._schedule_autosave()))

        self.spi_label = ttk.Label(topology_row, text="Strings/Device:")
        self.spi_label.pack(side='left', padx=(0, 5))
        self.strings_per_inverter_var = tk.StringVar(value='--')
        self._updating_spi = False  # Guard to prevent infinite loop
        spi_spinbox = ttk.Spinbox(
            topology_row, from_=1, to=100,
            textvariable=self.strings_per_inverter_var, width=6,
            font=('Helvetica', 10, 'bold')
        )
        spi_spinbox.pack(side='left', padx=(0, 15))
        self.strings_per_inverter_var.trace_add('write', self._on_strings_per_inverter_changed)
        
        # Central inverter count (only visible for Central Inverter topology)
        self.central_inv_count_label = ttk.Label(topology_row, text="Central Inverters:")
        self.central_inv_count_var = tk.StringVar(value='1')
        self.central_inv_count_spinbox = ttk.Spinbox(
            topology_row, from_=1, to=50,
            textvariable=self.central_inv_count_var, width=4
        )
        self.central_inv_count_var.trace_add('write', lambda *args: (self._mark_stale(), self._schedule_autosave()))
        # Hidden by default — shown when topology is Central Inverter

        # Isc warning label (hidden by default)
        self.isc_warning_label = ttk.Label(topology_row, text="", foreground='red')
        self.isc_warning_label.pack(side='left', padx=(5, 0))
        
        # Row 4: Other settings
        settings_inner = ttk.Frame(settings_frame)
        settings_inner.pack(fill='x')
        
        # Modules per string — hidden var, derived from template
        self.modules_per_string_var = tk.StringVar(value=str(getattr(self, 'modules_per_string_default', 28)))
        self.modules_per_string_var.trace_add('write', lambda *args: (self._update_strings_per_inverter(), self._mark_stale(), self._schedule_autosave()))

        # Wire sizing dict — populated dynamically based on harness configs
        # Keys: 'temp_rating', 'feeder_material', 'by_string_count', 'dc_feeder', 'ac_homerun', 'user_overrides'
        self.wire_sizing = {
            'temp_rating': '90C',
            'feeder_material': 'aluminum',
            'by_string_count': {},
            'dc_feeder': '',
            'ac_homerun': '',
            'user_overrides': {}
        }
        
        ttk.Label(settings_inner, text="Polarity:").pack(side='left', padx=(10, 5))
        self.polarity_convention_var = tk.StringVar(value='Negative Always South')
        polarity_combo = ttk.Combobox(
            settings_inner,
            textvariable=self.polarity_convention_var,
            values=[
                'Negative Always South',
                'Negative Always North',
                'Negative Toward Device',
                'Positive Toward Device'
            ],
            state='readonly',
            width=22
        )
        polarity_combo.pack(side='left')
        self.disable_combobox_scroll(polarity_combo)
        self.polarity_convention_var.trace_add('write', lambda *args: (self._mark_stale(), self._schedule_autosave()))
        
        # Row 5: Topology-dependent distance inputs
        distance_row = ttk.Frame(settings_frame)
        distance_row.pack(fill='x', pady=(5, 0))
        
        ttk.Label(distance_row, text="Avg DC Feeder Run (ft):").pack(side='left', padx=(0, 5))
        self.dc_feeder_distance_var = tk.StringVar(value='500')
        ttk.Spinbox(distance_row, from_=0, to=5000, increment=50,
                     textvariable=self.dc_feeder_distance_var, width=8).pack(side='left', padx=(0, 15))
        self.dc_feeder_distance_var.trace_add('write', lambda *args: (self._mark_stale(), self._schedule_autosave()))
        
        ttk.Label(distance_row, text="Avg AC Homerun (ft):").pack(side='left', padx=(0, 5))
        self.ac_homerun_distance_var = tk.StringVar(value='500')
        ttk.Spinbox(distance_row, from_=0, to=5000, increment=50,
                     textvariable=self.ac_homerun_distance_var, width=8).pack(side='left', padx=(0, 15))
        self.ac_homerun_distance_var.trace_add('write', lambda *args: (self._mark_stale(), self._schedule_autosave()))
        
        self.use_routed_var = tk.BooleanVar(value=False)
        self.use_routed_cb = ttk.Checkbutton(
            distance_row, text="Use Routed Distances",
            variable=self.use_routed_var,
            command=lambda: (self._mark_stale(), self._update_distance_hints(), self._schedule_autosave())
        )
        self.use_routed_cb.pack(side='left', padx=(10, 0))
        
        # Topology hint label
        self.distance_hint_label = ttk.Label(distance_row, text="", foreground='gray')
        self.distance_hint_label.pack(side='left', padx=(5, 0))
        self._update_distance_hints()
        
        # Wire Sizing frame (right side of Global Settings)
        self._build_wire_sizing_frame(settings_container)
        
        # Button row
        button_row = ttk.Frame(bottom_frame)
        button_row.pack(fill='x', pady=(0, 10))
        
        self._calc_btn = ttk.Button(button_row, text="Calculate Estimate", command=self.calculate_estimate, state='disabled')
        self._calc_btn.pack(side='left', padx=(0, 10))
        
        export_btn = ttk.Button(button_row, text="Export to Excel", command=self.export_to_excel)
        export_btn.pack(side='left')
        
        # packet_btn = ttk.Button(button_row, text="Export Packet", command=self.export_packet)
        # packet_btn.pack(side='left', padx=(10, 0))
        
        pdf_btn = ttk.Button(button_row, text="Export PDF", command=self.export_pdf_only)
        pdf_btn.pack(side='left', padx=(5, 0))
        
        preview_btn = ttk.Button(button_row, text="Site Preview", command=self.show_site_preview)
        preview_btn.pack(side='left', padx=(10, 0))

        diag_btn = ttk.Button(button_row, text="Run Diagnostics", command=self._run_diagnostics)
        diag_btn.pack(side='left', padx=(10, 0))
        
        # Results frame (full width)
        results_frame = ttk.LabelFrame(bottom_frame, text="Estimated BOM (Rolled-Up Totals)", padding="10")
        results_frame.pack(fill='both', expand=True)
        
        # Results treeview
        columns = ('include', 'item', 'part_number', 'description', 'quantity', 'unit', 'unit_cost', 'ext_cost')
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=8)
        self.results_tree.heading('include', text='')
        self.results_tree.heading('item', text='Item')
        self.results_tree.heading('part_number', text='Part Number')
        self.results_tree.heading('description', text='Description')
        self.results_tree.heading('quantity', text='Quantity')
        self.results_tree.heading('unit', text='Unit')
        self.results_tree.heading('unit_cost', text='Unit Cost')
        self.results_tree.heading('ext_cost', text='Ext. Cost')
        self.results_tree.column('include', width=30, anchor='center')
        self.results_tree.column('item', width=230, anchor='w')
        self.results_tree.column('part_number', width=160, anchor='w')
        self.results_tree.column('description', width=200, anchor='w')
        self.results_tree.column('quantity', width=80, anchor='center')
        self.results_tree.column('unit', width=50, anchor='center')
        self.results_tree.column('unit_cost', width=80, anchor='e')
        self.results_tree.column('ext_cost', width=90, anchor='e')

        # Tags for checked/unchecked styling
        self.results_tree.tag_configure('checked', foreground='black')
        self.results_tree.tag_configure('unchecked', foreground='gray60')
        self.results_tree.tag_configure('section', foreground='black')

        # Click handler for checkbox column
        self.results_tree.bind('<Button-1>', self._on_results_tree_click)
        
        # Scrollbar for results
        results_scroll = ttk.Scrollbar(results_frame, orient='vertical', command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=results_scroll.set)
        
        self.results_tree.pack(side='left', fill='both', expand=True)
        results_scroll.pack(side='right', fill='y')

        # Select all / none buttons
        check_btn_frame = ttk.Frame(results_frame)
        check_btn_frame.pack(fill='x', pady=(5, 0))
        ttk.Button(check_btn_frame, text="Select All", command=self._results_select_all).pack(side='left', padx=(0, 5))
        ttk.Button(check_btn_frame, text="Select None", command=self._results_select_none).pack(side='left')

    def setup_tree_panel(self, parent):
        """Setup the left row list panel"""
        # Label
        tree_label = ttk.Label(parent, text="Tracker Groups", font=('Helvetica', 10, 'bold'))
        tree_label.pack(anchor='w', pady=(0, 5))
        
        # Listbox frame
        list_frame = ttk.Frame(parent)
        list_frame.pack(fill='both', expand=True)
        
        self.group_listbox = tk.Listbox(list_frame, selectmode='browse', font=('Helvetica', 10))
        self.group_listbox.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        list_scroll = ttk.Scrollbar(list_frame, orient='vertical', command=self.group_listbox.yview)
        list_scroll.pack(side='right', fill='y')
        self.group_listbox.configure(yscrollcommand=list_scroll.set)
        
        # Bind selection
        self.group_listbox.bind('<<ListboxSelect>>', self.on_group_select)
        
        # Buttons frame
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Button(btn_frame, text="+ Group", command=self.add_group).pack(side='left', padx=(0, 5))
        ttk.Button(btn_frame, text="Copy", command=self.copy_selected_group).pack(side='left', padx=(0, 5))
        ttk.Button(btn_frame, text="Delete", command=self.delete_selected_group).pack(side='left', padx=(0, 5))
        
        # Move buttons
        move_frame = ttk.Frame(parent)
        move_frame.pack(fill='x', pady=(5, 0))
        ttk.Button(move_frame, text="▲ Up", command=self.move_group_up).pack(side='left', padx=(0, 5))
        ttk.Button(move_frame, text="▼ Down", command=self.move_group_down).pack(side='left')

    def setup_details_panel(self, parent):
        """Setup the right details panel"""
        # Details label
        self.details_label = ttk.Label(parent, text="Details", font=('Helvetica', 10, 'bold'))
        self.details_label.pack(anchor='w', pady=(0, 5))
        
        # Container for dynamic content
        self.details_container = ttk.Frame(parent)
        self.details_container.pack(fill='both', expand=True)
        
        # Placeholder
        self.placeholder_label = ttk.Label(self.details_container, text="Select a group to view details", foreground='gray')
        self.placeholder_label.pack(pady=20)

    def clear_details_panel(self):
        """Clear the details panel"""
        self._harness_combos = []
        for widget in self.details_container.winfo_children():
            widget.destroy()
        
        self.placeholder_label = ttk.Label(self.details_container, text="Select a group to view details", foreground='gray')
        self.placeholder_label.pack(pady=20)
        
        self.details_label.config(text="Details")

    def on_group_select(self, event):
        """Handle row selection changes"""
        if getattr(self, '_updating_listbox', False):
            return
        
        sel = self.group_listbox.curselection()
        if not sel:
            return
        
        new_idx = sel[0]
        
        # Sync the previous row's listbox text before switching (inline, no selection_set)
        if self.selected_group_idx is not None and self.selected_group_idx != new_idx and self.selected_group_idx < len(self.groups):
            prev = self.selected_group_idx
            group = self.groups[prev]
            total_trackers = sum(seg['quantity'] for seg in group['segments'])
            total_strings = sum(seg['quantity'] * seg['strings_per_tracker'] for seg in group['segments'])
            display = f"{group['name']}  ({total_trackers}T / {total_strings}S)"
            self._updating_listbox = True
            self.group_listbox.delete(prev)
            self.group_listbox.insert(prev, display)
            self.group_listbox.selection_clear(0, tk.END)
            self.group_listbox.selection_set(new_idx)
            self._updating_listbox = False
        
        if new_idx < 0 or new_idx >= len(self.groups):
            self.clear_details_panel()
            return
        
        # Clear existing details
        self._harness_combos = []
        for widget in self.details_container.winfo_children():
            widget.destroy()
        
        self.selected_group_idx = new_idx
        self.show_group_details(new_idx)

    def show_group_details(self, group_idx: int):
        """Show segment editor for the selected group"""
        if group_idx < 0 or group_idx >= len(self.groups):
            return
        
        group = self.groups[group_idx]
        self.details_label.config(text=f"Group: {group['name']}")
        
        # Create scrollable frame
        canvas = tk.Canvas(self.details_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.details_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        
        # Two-column layout: group inputs (left) | segments (right)
        two_col_frame = ttk.Frame(scrollable_frame)
        two_col_frame.pack(fill='both', expand=True)

        # Row name
        form_frame = ttk.Frame(two_col_frame, padding="10")
        form_frame.pack(side='left', fill='y')
        
        ttk.Label(form_frame, text="Group Name:").grid(row=0, column=0, sticky='w', pady=5)
        name_var = tk.StringVar(value=group['name'])
        name_entry = ttk.Entry(form_frame, textvariable=name_var, width=25)
        name_entry.grid(row=0, column=1, sticky='w', pady=5, padx=(10, 0))
        
        def update_name(*args):
            group['name'] = name_var.get()
            self.details_label.config(text=f"Group: {name_var.get()}")
            self._schedule_autosave()
        name_var.trace_add('write', update_name)
        
        # Device Position dropdown
        ttk.Label(form_frame, text="Device Position:").grid(row=1, column=0, sticky='w', pady=5)
        device_pos_var = tk.StringVar(value=group.get('device_position', 'middle'))
        device_pos_combo = ttk.Combobox(form_frame, textvariable=device_pos_var,
                                         values=['north', 'middle', 'south'],
                                         state='readonly', width=10)
        device_pos_combo.grid(row=1, column=1, sticky='w', pady=5, padx=(10, 0))
        
        def on_device_position_change(*args):
            group['device_position'] = device_pos_var.get()
            self._mark_stale()
            self._schedule_autosave()
        device_pos_combo.bind('<<ComboboxSelected>>', on_device_position_change)
        
        # Driveline Angle
        ttk.Label(form_frame, text="Driveline Angle (°):").grid(row=2, column=0, sticky='w', pady=5)
        angle_var = tk.StringVar(value=str(group.get('driveline_angle', 0.0)))
        angle_spinbox = ttk.Spinbox(form_frame, from_=-45, to=45, increment=0.5,
                                     textvariable=angle_var, width=8)
        angle_spinbox.grid(row=2, column=1, sticky='w', pady=5, padx=(10, 0))
        
        def on_angle_change(*args):
            try:
                val = float(angle_var.get())
                val = max(-45.0, min(45.0, val))
                group['driveline_angle'] = val
            except ValueError:
                pass
            self._mark_stale()
            self._schedule_autosave()
        angle_var.trace_add('write', on_angle_change)

        # Azimuth
        ttk.Label(form_frame, text="Azimuth (°):").grid(row=3, column=0, sticky='w', pady=5)
        azimuth_var = tk.StringVar(value=str(group.get('azimuth', 180)))
        azimuth_spinbox = ttk.Spinbox(form_frame, from_=0, to=359, increment=1,
                                      textvariable=azimuth_var, width=8)
        azimuth_spinbox.grid(row=3, column=1, sticky='w', pady=5, padx=(10, 0))
        ttk.Label(form_frame, text="(180 = south-facing)", foreground='gray').grid(
            row=3, column=2, columnspan=2, sticky='w', pady=5, padx=(10, 0))

        def on_azimuth_change(*args):
            try:
                val = float(azimuth_var.get())
                val = val % 360
                group['azimuth'] = val
            except ValueError:
                pass
            self._mark_stale()
            self._schedule_autosave()
        azimuth_var.trace_add('write', on_azimuth_change)

        # Tracker Alignment
        ttk.Label(form_frame, text="Tracker Alignment:").grid(row=4, column=0, sticky='w', pady=5)
        alignment_var = tk.StringVar(value=group.get('tracker_alignment', 'motor'))
        alignment_combo = ttk.Combobox(form_frame, textvariable=alignment_var,
                                       values=['top', 'motor', 'bottom'],
                                       state='readonly', width=10)
        alignment_combo.grid(row=4, column=1, sticky='w', pady=5, padx=(10, 0))
        self.disable_combobox_scroll(alignment_combo)

        def on_alignment_change(*args):
            group['tracker_alignment'] = alignment_var.get()
            self._mark_stale()
            self._schedule_autosave()
        alignment_combo.bind('<<ComboboxSelected>>', on_alignment_change)

        # Row Spacing + GCR (bidirectional)
        ttk.Label(form_frame, text="Row Spacing (ft):").grid(row=5, column=0, sticky='w', pady=5)
        row_spacing_var = tk.StringVar(value=f"{group.get('row_spacing_ft', 20.0):.3f}")
        row_spacing_entry = ttk.Entry(form_frame, textvariable=row_spacing_var, width=10)
        row_spacing_entry.grid(row=5, column=1, sticky='w', pady=5, padx=(10, 0))

        ttk.Label(form_frame, text="GCR:").grid(row=5, column=2, sticky='w', pady=5, padx=(15, 0))
        gcr_var = tk.StringVar(value="--")
        gcr_entry = ttk.Entry(form_frame, textvariable=gcr_var, width=8)
        gcr_entry.grid(row=5, column=3, sticky='w', pady=5, padx=(5, 0))

        def _update_gcr_from_row_spacing(*args):
            try:
                rs_ft = float(row_spacing_var.get())
                if rs_ft <= 0:
                    return
                rs_m = rs_ft / 3.28084
                mod_len_m = self._get_group_module_length_m(group)
                if mod_len_m:
                    gcr_var.set(f"{mod_len_m / rs_m:.3f}")
                else:
                    gcr_var.set("--")
                group['row_spacing_ft'] = rs_ft
                self._mark_stale()
                self._schedule_autosave()
            except ValueError:
                pass

        def _update_row_spacing_from_gcr(*args):
            try:
                gcr_str = gcr_var.get().strip()
                if gcr_str == "--":
                    return
                gcr = float(gcr_str)
                if gcr <= 0 or gcr > 1.0:
                    return
                mod_len_m = self._get_group_module_length_m(group)
                if not mod_len_m:
                    return
                rs_m = mod_len_m / gcr
                rs_ft = rs_m * 3.28084
                row_spacing_var.set(f"{rs_ft:.3f}")
                group['row_spacing_ft'] = rs_ft
                self._mark_stale()
                self._schedule_autosave()
            except ValueError:
                pass

        row_spacing_entry.bind('<FocusOut>', _update_gcr_from_row_spacing)
        row_spacing_entry.bind('<Return>', _update_gcr_from_row_spacing)
        gcr_entry.bind('<FocusOut>', _update_row_spacing_from_gcr)
        gcr_entry.bind('<Return>', _update_row_spacing_from_gcr)

        # Populate GCR on load
        _update_gcr_from_row_spacing()

        # Strings per Device override
        ttk.Label(form_frame, text="Strings/Device:").grid(row=6, column=0, sticky='w', pady=5)
        spi_frame = ttk.Frame(form_frame)
        spi_frame.grid(row=6, column=1, columnspan=3, sticky='w', pady=5, padx=(10, 0))

        spi_override_var = tk.StringVar(
            value=str(group['strings_per_inv']) if group.get('strings_per_inv') is not None else ''
        )
        spi_entry = ttk.Entry(spi_frame, textvariable=spi_override_var, width=6)
        spi_entry.pack(side='left')

        # Compute the global-derived value to show as placeholder hint
        def _get_global_spi_hint():
            try:
                spi = self._get_int_var(self.strings_per_inverter_var, 0)
                return f"(global: {spi})" if spi > 0 else "(global default)"
            except Exception:
                return ""

        spi_hint_label = ttk.Label(spi_frame, text=_get_global_spi_hint(), foreground='gray')
        spi_hint_label.pack(side='left', padx=(6, 0))

        def _on_spi_change(*args):
            val = spi_override_var.get().strip()
            if val == '':
                group['strings_per_inv'] = None
            else:
                try:
                    group['strings_per_inv'] = int(val)
                except ValueError:
                    pass
            self._mark_stale()
            self._schedule_autosave()
        spi_override_var.trace_add('write', _on_spi_change)

        # Segments section
        seg_frame = ttk.LabelFrame(two_col_frame, text="Segments (left to right)", padding="10")
        seg_frame.pack(side='left', fill='both', expand=True, pady=10, padx=(0, 10))
        
        # String count display
        self.group_string_count_label = ttk.Label(seg_frame, text="", font=('Helvetica', 10, 'bold'))
        self.group_string_count_label.pack(anchor='w', pady=(0, 10))
        
        # Headers
        header_frame = ttk.Frame(seg_frame)
        header_frame.pack(fill='x')
        ttk.Label(header_frame, text="Qty", width=6).pack(side='left', padx=2)
        ttk.Label(header_frame, text="Tracker Template", width=30).pack(side='left', padx=2)
        ttk.Label(header_frame, text="Harness Config", width=14).pack(side='left', padx=2)
        ttk.Label(header_frame, text="", width=9).pack(side='left', padx=2)
        
        # Container for segment rows
        self.segment_rows_container = ttk.Frame(seg_frame)
        self.segment_rows_container.pack(fill='x', pady=(5, 0))
        
        # Add existing segment rows
        for i, seg in enumerate(group['segments']):
            self._add_segment_ui(group, group_idx, i, seg)
        
        # Update count
        self._update_group_string_count(group)
        
        # Add segment button
        add_btn = ttk.Button(seg_frame, text="+ Add Segment",
                            command=lambda: self._add_segment_to_group(group, group_idx))
        add_btn.pack(anchor='w', pady=(10, 0))
    
    def _add_segment_ui(self, group: dict, group_idx: int, seg_idx: int, segment: dict):
        """Add a segment configuration row to the UI"""
        row_frame = ttk.Frame(self.segment_rows_container)
        row_frame.pack(fill='x', pady=2)
        
        # Quantity
        qty_var = tk.StringVar(value=str(segment['quantity']))
        qty_spinbox = ttk.Spinbox(row_frame, from_=1, to=10000, textvariable=qty_var, width=6)
        qty_spinbox.pack(side='left', padx=2)
        
        # Template dropdown
        compatible_keys = self.get_compatible_templates(group)
        template_display_map = {}  # display_name -> template_key
        display_names = []
        
        # Add "Unlinked" option for backward compat
        unlinked_label = f"Unlinked ({segment.get('strings_per_tracker', 3)}S)"
        display_names.append(unlinked_label)
        template_display_map[unlinked_label] = None
        
        for key in compatible_keys:
            display = self.get_template_display_name(key)
            display_names.append(display)
            template_display_map[display] = key
        
        # Determine current selection
        current_ref = segment.get('template_ref')
        if current_ref and current_ref in self.enabled_templates:
            current_display = self.get_template_display_name(current_ref)
        else:
            current_display = unlinked_label
        
        template_var = tk.StringVar(value=current_display)
        template_combo = ttk.Combobox(row_frame, textvariable=template_var,
                                      values=display_names, width=28, state='readonly')
        template_combo.pack(side='left', padx=2)
        self.disable_combobox_scroll(template_combo)
        
        # Harness config (still user-specified)
        harness_var = tk.StringVar(value=segment['harness_config'])
        harness_options = self.get_harness_options(segment.get('strings_per_tracker', 3))
        harness_combo = ttk.Combobox(row_frame, textvariable=harness_var,
                                     values=harness_options, width=12)
        harness_combo.pack(side='left', padx=2)
        self.disable_combobox_scroll(harness_combo)
        
        # Track combo for LV collection method disabling
        self._harness_combos.append((harness_combo, harness_var, segment))
        if self.lv_collection_var.get() == 'String HR':
            spt = segment.get('strings_per_tracker', 1)
            override_display = '+'.join(['1'] * int(spt))
            original_config = segment['harness_config']
            harness_var.set(override_display)
            segment['harness_config'] = original_config  # restore — don't overwrite saved config
            harness_combo.config(state='disabled')
        elif self.lv_collection_var.get() == 'Trunk Bus':
            harness_combo.config(state='disabled')
        
        # Move up button
        up_btn = ttk.Button(row_frame, text="▲", width=2,
                           command=lambda si=seg_idx: self._move_segment(group, group_idx, si, -1))
        up_btn.pack(side='left', padx=(2, 0))
        if seg_idx == 0:
            up_btn.config(state='disabled')
        
        # Move down button
        down_btn = ttk.Button(row_frame, text="▼", width=2,
                             command=lambda si=seg_idx: self._move_segment(group, group_idx, si, 1))
        down_btn.pack(side='left', padx=(0, 2))
        if seg_idx == len(group['segments']) - 1:
            down_btn.config(state='disabled')
        
        # Delete button
        del_btn = ttk.Button(row_frame, text="×", width=3,
                            command=lambda: self._delete_segment(group, group_idx, seg_idx))
        del_btn.pack(side='left', padx=2)
        
        # Callbacks
        def on_template_change(*args):
            selected_display = template_var.get()
            selected_key = template_display_map.get(selected_display)
            segment['template_ref'] = selected_key

            if selected_key and selected_key in self.enabled_templates:
                # Derive strings_per_tracker from template
                tdata = self.enabled_templates[selected_key]
                new_spt = tdata.get('strings_per_tracker', 3)
                segment['strings_per_tracker'] = new_spt

                # Update harness options for new string count; suppress the
                # on_harness_change trace so wire sizing only runs once below.
                new_options = self.get_harness_options(new_spt)
                harness_combo['values'] = new_options

                self._suppress_harness_trace = True
                try:
                    derived = self._get_default_harness_config_from_template(selected_key)
                    if derived and derived in new_options:
                        harness_var.set(derived)
                        segment['harness_config'] = derived
                    elif harness_var.get() not in new_options:
                        harness_var.set(new_options[0])
                        segment['harness_config'] = new_options[0]
                finally:
                    self._suppress_harness_trace = False

            self._update_group_string_count(group)
            self._auto_unlock_allocation()
            self._mark_stale()
            self._schedule_autosave()

            # Update derived module from templates
            self._derive_module_from_templates()

            # Refresh wire sizing once (trace suppressed above)
            self._refresh_wire_sizing_for_segments()
        template_combo.bind('<<ComboboxSelected>>', on_template_change)
        
        def on_qty_change(*args):
            try:
                new_qty = max(1, int(qty_var.get()))
            except ValueError:
                return
            old_qty = segment.get('quantity', 1)
            segment['quantity'] = new_qty
            delta = new_qty - old_qty
            self._update_group_string_count(group)
            if delta != 0 and self.allocation_locked and self.locked_allocation_result is not None:
                # Compute global start index of this segment
                seg_global_start = 0
                found = False
                for g in self.groups:
                    for s in g.get('segments', []):
                        if s is segment:
                            found = True
                            break
                        seg_global_start += s.get('quantity', 0)
                    if found:
                        break
                if found:
                    insertion_point = seg_global_start + (old_qty if delta > 0 else new_qty)
                    self._shift_locked_allocation(insertion_point, delta)
                else:
                    self._reconcile_locked_allocation()
            self._mark_stale()
            self._schedule_autosave()
        qty_var.trace_add('write', on_qty_change)
        
        def on_harness_change(*args):
            if self._suppress_harness_trace:
                return
            segment['harness_config'] = harness_var.get()
            self._refresh_wire_sizing_for_segments()
            self._mark_stale()
            self._schedule_autosave()
        harness_var.trace_add('write', on_harness_change)
    
    def _add_segment_to_group(self, group: dict, group_idx: int):
        """Add a new segment to the group and refresh UI"""
        # Default to first compatible template if available
        compatible = self.get_compatible_templates(group)
        if compatible:
            default_ref = compatible[0]
            default_spt = self.enabled_templates[default_ref].get('strings_per_tracker', 3)
        else:
            default_ref = None
            default_spt = 3
        
        # Auto-derive harness config from template motor position
        default_harness = self._get_default_harness_config_from_template(default_ref)
        if not default_harness:
            default_harness = str(default_spt)
        
        group['segments'].append({
            'quantity': 1,
            'strings_per_tracker': default_spt,
            'harness_config': default_harness,
            'template_ref': default_ref
        })
        new_seg_idx = len(group['segments']) - 1
        self._add_segment_ui(group, group_idx, new_seg_idx, group['segments'][-1])
        # The previous last row's down button must now be enabled
        if new_seg_idx > 0:
            rows = self.segment_rows_container.winfo_children()
            if len(rows) >= 2:
                prev_children = rows[-2].winfo_children()
                if len(prev_children) > 4:
                    prev_children[4].config(state='normal')  # down_btn is index 4
        self._update_group_string_count(group)
        self._refresh_wire_sizing_for_segments()
        self._mark_dirty()
    
    def _move_segment(self, group: dict, group_idx: int, seg_idx: int, direction: int):
        """Move a segment up (-1) or down (+1) within the group"""
        new_idx = seg_idx + direction
        if new_idx < 0 or new_idx >= len(group['segments']):
            return
        group['segments'][seg_idx], group['segments'][new_idx] = group['segments'][new_idx], group['segments'][seg_idx]
        self._auto_unlock_allocation()
        self._rebuild_group_details(group_idx)

    def _delete_segment(self, group: dict, group_idx: int, seg_idx: int):
        """Delete a segment from the group"""
        if len(group['segments']) <= 1:
            return  # Keep at least one segment
        del group['segments'][seg_idx]
        self._reconcile_locked_allocation()
        self._rebuild_group_details(group_idx)
    
    def _update_group_string_count(self, group: dict):
        """Update the string/tracker count label for a group"""
        if not hasattr(self, 'group_string_count_label'):
            return
        total_trackers = sum(seg['quantity'] for seg in group['segments'])
        total_strings = sum(seg['quantity'] * seg['strings_per_tracker'] for seg in group['segments'])
        self.group_string_count_label.config(
            text=f"Total Strings: {total_strings:,}  |  Total Trackers: {total_trackers:,}")

    # ==================== Calculation Methods ====================

    def calculate_estimate(self, silent=False):
        """Calculate and display the rolled-up BOM estimate"""
        # Clear previous results
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        self.checked_items.clear()
        
        # Sync all group listbox text before calculating
        self._refresh_group_listbox(preserve_selection=True)
        
        # Aggregated totals
        totals = {
            'combiners_by_breaker': defaultdict(int),
            'combiner_details': [],
            'string_inverters': 0,
            'trackers_by_string': defaultdict(int),
            'harnesses_by_size': defaultdict(int),
            'inline_fuses_by_rating': defaultdict(int),
            'whips_by_length': defaultdict(int),
            'extenders_pos_by_length': defaultdict(int),
            'extenders_neg_by_length': defaultdict(int),
            'total_whip_length': 0,
            'dc_feeder_total_ft': 0,
            'dc_feeder_count': 0,
            'ac_homerun_total_ft': 0,
            'ac_homerun_count': 0,
            # Trunk Bus items
            'trunk_cable_by_size': defaultdict(float),
            'lbd_by_size': defaultdict(int),
            'ipc_by_tap': defaultdict(int),
        }
        
        # Topology and strings-per-inverter (used throughout calculation)
        topology = self.topology_var.get()
        
        if topology == 'Central Inverter':
            # For Central Inverter, allocation groups = CBs.
            # Compute strings_per_cb from library BEFORE allocation runs.
            breaker_size_early = self._get_int_var(self.breaker_size_var, 400)
            strings_per_inv = 0  # Will be set after we know module_isc
            _central_inv_breaker = breaker_size_early
        else:
            strings_per_inv = self._get_int_var(self.strings_per_inverter_var, 0)
        
        # Validate — need at least one linked template or a legacy fallback module
        self._derive_module_from_templates()
        lv_method = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
        if not self.selected_module:
            if not silent:
                messagebox.showwarning(
                    "No Module Available",
                    "No tracker templates are linked and no legacy module data was found.\n\n"
                    "Please link a tracker template to at least one segment before calculating."
                )
            return
        
        # ==================== Build tracker sequence from groups ====================
        try:
            modules_per_string = int(self.modules_per_string_var.get())
        except ValueError:
            modules_per_string = 28

        tracker_sequence = []
        total_all_trackers = 0
        total_all_strings = 0
        total_all_harnesses = 0
        max_harness_strings = 0
        self._tracker_to_segment = []
        
        # Track unique modules across all groups
        unique_modules = {}  # "Manufacturer Model (WattageW)" -> module_spec_dict
        
        # Per-segment module data for geometry calculations
        segment_module_data = []  # list of {module_spec_dict, modules_per_string, qty, spt}

        for group in self.groups:
            for seg in group['segments']:
                qty = seg['quantity']
                spt = seg['strings_per_tracker']
                harness_config = self._get_effective_harness_config(seg)

                if qty <= 0:
                    continue
                
                # Resolve module from template
                seg_module = None
                seg_mps = modules_per_string
                ref = seg.get('template_ref')
                if ref and ref in self.enabled_templates:
                    tdata = self.enabled_templates[ref]
                    seg_module = tdata.get('module_spec', {})
                    seg_mps = tdata.get('modules_per_string', modules_per_string)
                    
                    # Track unique modules
                    mod_key = f"{seg_module.get('manufacturer', '?')} {seg_module.get('model', '?')} ({seg_module.get('wattage', '?')}W)"
                    if mod_key not in unique_modules:
                        unique_modules[mod_key] = seg_module
                
                segment_module_data.append({
                    'module_spec': seg_module,
                    'modules_per_string': seg_mps,
                    'qty': qty,
                    'spt': spt
                })

                # Determine if this segment has partial strings
                full_spt = int(spt)
                has_partial = (spt != full_spt)
                
                # Add to flat tracker sequence (one entry per physical tracker)
                # For partial-string trackers, pair adjacent trackers:
                #   Left of pair: full_spt + 1 (owns the shared string)
                #   Right of pair: full_spt
                #   Unpaired odd tracker: full_spt (half-string not counted)
                num_pairs = qty // 2 if has_partial else 0
                unpaired = qty % 2 if has_partial else 0
                
                for i in range(qty):
                    if has_partial:
                        if i % 2 == 0 and i + 1 < qty:
                            effective_spt = full_spt + 1  # Left of pair — owns shared string
                        elif i % 2 == 1:
                            effective_spt = full_spt  # Right of pair
                        else:
                            effective_spt = full_spt  # Unpaired odd tracker
                    else:
                        effective_spt = spt
                    
                    tracker_sequence.append(effective_spt)
                    self._tracker_to_segment.append({
                        'group_idx': self.groups.index(group),
                        'seg': seg,
                        'device_position': group.get('device_position', 'middle'),
                    })

                total_all_trackers += qty
                if has_partial:
                    total_all_strings += full_spt * qty + num_pairs  # Only paired halves count
                else:
                    total_all_strings += qty * spt

                # Count trackers by effective string count
                if has_partial:
                    num_pairs_count = qty // 2
                    # Left of pair
                    left_spt = full_spt + 1
                    totals['trackers_by_string'][left_spt] += num_pairs_count
                    # Right of pair + any unpaired
                    right_count = num_pairs_count + (qty % 2)
                    totals['trackers_by_string'][full_spt] += right_count
                else:
                    totals['trackers_by_string'][spt] += qty

                # Count harnesses by size (skip for Trunk Bus — no harnesses)
                if lv_method != 'Trunk Bus':
                    harness_sizes = self.parse_harness_config(harness_config)
                    for size in harness_sizes:
                        if size > max_harness_strings:
                            max_harness_strings = size
                        totals['harnesses_by_size'][size] += qty
                        total_all_harnesses += qty

                # Count inline fuses (Wire Harness only, positive side, size >= 2)
                if lv_method == 'Wire Harness':
                    for size in harness_sizes:
                        if size >= 2:
                            fuse_rating = self._calc_inline_fuse_rating(seg_module)
                            totals['inline_fuses_by_rating'][fuse_rating] += size * qty
        
        # Build harness-count-per-spt lookup for whip calculation
        harness_count_by_spt = {}
        harness_sizes_by_spt = {}
        for group in self.groups:
            for seg in group['segments']:
                spt = seg['strings_per_tracker']
                if spt not in harness_count_by_spt:
                    harness_sizes = self._get_harness_sizes(seg)
                    harness_count_by_spt[spt] = len(harness_sizes)
                    harness_sizes_by_spt[spt] = harness_sizes

        # Build per-tracker harness sizes (indexed by global tracker index)
        # This handles segments with same SPT but different harness configs
        tracker_harness_sizes_list = []
        for group in self.groups:
            for seg in group['segments']:
                harness_sizes = self._get_harness_sizes(seg)
                for _ in range(seg['quantity']):
                    tracker_harness_sizes_list.append(list(harness_sizes))

        # Warn about unpaired partial strings
        unpaired_warnings = []
        for group in self.groups:
            for seg in group['segments']:
                seg_spt = seg['strings_per_tracker']
                if seg_spt != int(seg_spt) and seg['quantity'] % 2 != 0:
                    ref = seg.get('template_ref', 'Unlinked')
                    unpaired_warnings.append(f"{group['name']}: {seg['quantity']}x {seg_spt}S has 1 unpaired half-string")
        
        if unpaired_warnings:
            messagebox.showwarning(
                "Unpaired Partial Strings",
                "The following segments have an odd number of partial-string trackers, "
                "leaving half-strings unpaired:\n\n" + "\n".join(unpaired_warnings)
            )

        # ==================== Module geometry (primary module for global calcs) ====================
        module_isc = self.selected_module.isc
        nec_factor = 1.56
        if self.current_project:
            nec_factor = getattr(self.current_project, 'nec_safety_factor', 1.56)
        module_width_mm = self.selected_module.width_mm
        module_width_ft = module_width_mm / 304.8
        string_length_ft = module_width_ft * modules_per_string

        # ==================== Build spatial tracker entries ====================
        fallback_width_ft = 6.0
        fallback_length_ft = 180.0

        # Pre-compute tracker count per group for auto-layout fallback
        group_tracker_counts = []
        for group in self.groups:
            count = sum(seg['quantity'] for seg in group['segments'] if seg['quantity'] > 0)
            group_tracker_counts.append(count)

        tracker_entries = []
        flat_idx = 0
        auto_x_cursor = 0.0  # Running X for auto-layout

        for grp_idx, group in enumerate(self.groups):
            saved_x = group.get('position_x')
            saved_y = group.get('position_y')

            if saved_x is not None and saved_y is not None:
                group_x = saved_x
                group_y = saved_y
            else:
                # Auto-layout: stack groups left-to-right at y=0
                group_x = auto_x_cursor
                group_y = 0.0

            grp_row_spacing = group.get('row_spacing_ft', 20.0)

            # Compute group reference motor Y (same as SitePreviewWindow uses)
            # This is the motor_y of the FIRST segment's template in the group
            group_ref_motor_y_ft = 0.0
            for seg in group['segments']:
                ref_check = seg.get('template_ref')
                if ref_check and ref_check in self.enabled_templates:
                    tdata_check = self.enabled_templates[ref_check]
                    if tdata_check.get('has_motor', True):
                        # Compute this template's motor_y using same logic
                        ms_c = tdata_check.get('module_spec', {})
                        orient_c = tdata_check.get('module_orientation', 'Portrait')
                        mps_c = tdata_check.get('modules_per_string', 28)
                        spacing_c = tdata_check.get('module_spacing_m', 0.02)
                        placement_c = tdata_check.get('motor_placement_type', 'between_strings')
                        pos_after_c = tdata_check.get('motor_position_after_string', None)
                        str_idx_c = tdata_check.get('motor_string_index', None)
                        split_n_c = tdata_check.get('motor_split_north', mps_c // 2)
                        mod_along_c = (ms_c.get('width_mm', 1000) if orient_c == 'Portrait' else ms_c.get('length_mm', 2000)) / 1000
                        
                        # Partial string on north adds offset
                        spt_c = tdata_check.get('strings_per_tracker', 1)
                        partial_north_m_c = 0
                        if spt_c != int(spt_c) and tdata_check.get('partial_string_side', 'north') == 'north':
                            partial_north_mods_c = round((spt_c - int(spt_c)) * mps_c)
                            partial_north_m_c = partial_north_mods_c * (mod_along_c + spacing_c)
                        
                        if placement_c == 'between_strings':
                            p = pos_after_c if pos_after_c is not None else (str_idx_c if str_idx_c is not None else 1)
                            mn = p * mps_c
                            motor_y_m = partial_north_m_c + (mn * mod_along_c + (mn - 1) * spacing_c + spacing_c) if mn > 0 else 0.0
                            group_ref_motor_y_ft = motor_y_m * 3.28084
                        elif placement_c == 'middle_of_string':
                            s = str_idx_c if str_idx_c is not None else 1
                            mb = (s - 1) * mps_c + split_n_c
                            motor_y_m = partial_north_m_c + (mb * mod_along_c + (mb - 1) * spacing_c + spacing_c)
                            group_ref_motor_y_ft = motor_y_m * 3.28084
                        break  # Use first template's motor as reference

            # Driveline angle for this group
            driveline_angle_deg = group.get('driveline_angle', 0.0)
            driveline_tan = math.tan(math.radians(driveline_angle_deg)) if driveline_angle_deg != 0 else 0.0

            tracker_within_group = 0
            for seg in group['segments']:
                qty = seg['quantity']
                spt = seg['strings_per_tracker']
                if qty <= 0:
                    continue

                ref = seg.get('template_ref')
                dims = self._get_estimate_tracker_dims_ft(ref)
                t_length = dims[1] if dims else fallback_length_ft
                
                # Partial string pairing (same logic as tracker_sequence)
                full_spt = int(spt)
                has_partial = (spt != full_spt)

                for i in range(qty):
                    if has_partial:
                        if i % 2 == 0 and i + 1 < qty:
                            effective_spt = full_spt + 1
                        else:
                            effective_spt = full_spt
                    else:
                        effective_spt = int(spt)
                    
                    local_x_offset = tracker_within_group * grp_row_spacing
                    tracker_entries.append({
                        'original_idx': flat_idx,
                        'spt': effective_spt,
                        'x': group_x + local_x_offset,
                        'y': group_y + local_x_offset * driveline_tan,
                        'length_ft': t_length,
                        'motor_y_ft': group_ref_motor_y_ft,
                        'row_spacing_ft': grp_row_spacing,
                    })
                    flat_idx += 1
                    tracker_within_group += 1

            # Advance auto-layout cursor for next group
            group_width = group_tracker_counts[grp_idx] * grp_row_spacing
            auto_x_cursor += group_width + grp_row_spacing * 2  # Extra gap between groups

        # For Central Inverter, compute strings_per_cb from library now that we have module_isc
        if topology == 'Central Inverter':
            _ci_fuse_current = module_isc * nec_factor * max(max_harness_strings, 1)
            _ci_fuse_rating = self.get_fuse_holder_category(_ci_fuse_current)
            _ci_library = self.load_combiner_library()
            _ci_matching = [
                cb for cb in _ci_library.values()
                if (cb.get('breaker_size', 0) == _central_inv_breaker and
                    cb.get('fuse_holder_rating', '') == _ci_fuse_rating)
            ]
            if _ci_matching:
                _ci_matching.sort(key=lambda c: c.get('max_inputs', 0), reverse=True)
                _ci_max_inputs = _ci_matching[0].get('max_inputs', 24)
            else:
                _ci_max_inputs = 24
            
            # User's strings/CB input takes priority; library max is the ceiling
            user_strings_per_cb = self._get_int_var(self.strings_per_inverter_var, 0)
            if user_strings_per_cb > 0:
                # Cap at library max inputs
                strings_per_inv = min(user_strings_per_cb, _ci_max_inputs)
            else:
                # No user input — default to library max
                strings_per_inv = _ci_max_inputs
                self._updating_spi = True
                self.strings_per_inverter_var.set(str(strings_per_inv))
                self._updating_spi = False

        # ==================== Allocation ====================
        allocation_result = None

        if self.selected_inverter and strings_per_inv > 0 and total_all_strings > 0:
            if self.allocation_locked and self.locked_allocation_result is not None:
                # Use the locked (frozen) allocation — skip spatial recalculation
                allocation_result = self.locked_allocation_result
                spatial_runs = allocation_result.get('spatial_runs', 1)
            elif tracker_entries:
                # Run allocation per group (respects per-group strings_per_inv override)
                # then merge results with remapped tracker indices
                merged_inverters = []
                flat_offset = 0  # running global tracker index offset per group

                for grp_idx, group in enumerate(self.groups):
                    grp_count = sum(seg['quantity'] for seg in group['segments'])
                    grp_entries = tracker_entries[flat_offset:flat_offset + grp_count]

                    grp_spi = group.get('strings_per_inv') or strings_per_inv
                    grp_pitch = group.get('row_spacing_ft', 20.0)

                    if grp_entries and grp_spi > 0:
                        grp_result = allocate_strings_spatial(grp_entries, grp_spi, grp_pitch)
                        merged_inverters.extend(grp_result.get('inverters', []))

                    flat_offset += grp_count

                if merged_inverters:
                    # Build merged summary
                    total_inv_strings = sum(inv['total_strings'] for inv in merged_inverters)
                    split_tidxs = set()
                    for inv in merged_inverters:
                        for entry in inv.get('harness_map', []):
                            if entry.get('is_split'):
                                split_tidxs.add(entry['tracker_idx'])
                    total_split = len(split_tidxs)
                    inv_sizes = [inv['total_strings'] for inv in merged_inverters]
                    allocation_result = {
                        'inverters': merged_inverters,
                        'spatial_runs': len(self.groups),
                        'summary': {
                            'total_inverters': len(merged_inverters),
                            'total_strings': total_inv_strings,
                            'total_trackers': len(tracker_entries),
                            'total_split_trackers': total_split,
                            'max_strings_per_inverter': max(inv_sizes) if inv_sizes else 0,
                            'min_strings_per_inverter': min(inv_sizes) if inv_sizes else 0,
                            'num_larger_inverters': 0,
                            'num_smaller_inverters': 0,
                            'tracker_type_counts': {},
                        }
                    }
                    spatial_runs = len(self.groups)
                else:
                    allocation_result = allocate_strings_sequential(tracker_sequence, strings_per_inv)
                    spatial_runs = 1

            module_wattage = self.selected_module.wattage
            # Use site-level DC:AC (total DC power / total AC capacity)
            total_alloc_strings = allocation_result['summary']['total_strings']
            total_alloc_invs = allocation_result['summary']['total_inverters']
            total_dc_kw = (total_alloc_strings * modules_per_string * module_wattage) / 1000
            total_ac_kw = 0.0

            if topology == 'Central Inverter':
                # For Central Inverter, allocation groups are CBs, not inverters.
                # DC:AC uses the user-specified central inverter count.
                central_inv_count = self._get_int_var(self.central_inv_count_var, 1)
                total_ac_kw = central_inv_count * self.selected_inverter.rated_power_kw
                actual_dc_ac = round(total_dc_kw / total_ac_kw, 3) if total_ac_kw > 0 else 0.0
            else:
                if total_alloc_invs > 0:
                    total_ac_kw = total_alloc_invs * self.selected_inverter.rated_power_kw
                    actual_dc_ac = round(total_dc_kw / total_ac_kw, 3)
                else:
                    actual_dc_ac = 0.0         

            if topology == 'Central Inverter':
                totals['string_inverters'] = self._get_int_var(self.central_inv_count_var, 1)
            else:
                totals['string_inverters'] = allocation_result['summary']['total_inverters']
            # Recount split trackers by unique tracker_idx (avoids double-counting)
            _split_tidxs = set()
            for _inv in allocation_result.get('inverters', []):
                for _entry in _inv.get('harness_map', []):
                    if _entry.get('is_split'):
                        _split_tidxs.add(_entry['tracker_idx'])
            _true_split_count = len(_split_tidxs)
            
            totals['inverter_summary'] = {
                'strings_per_inverter': strings_per_inv,
                'total_inverters': allocation_result['summary']['total_inverters'],
                'total_split_trackers': _true_split_count,
                'total_strings': allocation_result['summary']['total_strings'],
                'actual_dc_ac': actual_dc_ac,
                'allocation_result': allocation_result,
                'allocations': [],  # Backward compat
                'central_inverter_count': self._get_int_var(self.central_inv_count_var, 1) if topology == 'Central Inverter' else 0,
            }
            totals['total_dc_kw'] = round(total_dc_kw, 2)
            totals['total_ac_kw'] = round(total_ac_kw, 2)
            totals['total_modules'] = total_alloc_strings * modules_per_string

            # For Central Inverter, update device/combiner count from allocation
            # (allocation respects group boundaries, may differ from library estimate)
            if topology == 'Central Inverter':
                num_combiners = allocation_result['summary']['total_inverters']
                num_devices = num_combiners

        # ==================== Topology-driven device & combiner counting ===================

        total_inverters_count = totals.get('inverter_summary', {}).get('total_inverters', 0)
        num_devices = 0
        num_combiners = 0
        strings_per_cb = 0

        # Global breaker size (will add UI field later, default 400A for now)
        breaker_size = self._get_int_var(self.breaker_size_var, 400)

        if topology == 'Distributed String':
            num_devices = total_inverters_count

        elif topology == 'Centralized String':
            num_devices = total_inverters_count
            num_combiners = total_inverters_count
            strings_per_cb = strings_per_inv

        elif topology == 'Central Inverter':
            if lv_method == 'Trunk Bus':
                # Trunk Bus: allocation groups = LBDs, not combiner boxes.
                # num_devices is still needed for DC feeder distance calc.
                if totals.get('inverter_summary', {}).get('allocation_result'):
                    num_devices = totals['inverter_summary']['allocation_result']['summary']['total_inverters']
                # Do NOT populate combiners_by_breaker — no combiner boxes for Trunk Bus.
            else:
                # CB count from library: target breaker + fuse holder rating → max inputs
                fuse_current = module_isc * nec_factor * max(max_harness_strings, 1)
                fuse_holder_rating = self.get_fuse_holder_category(fuse_current)
                combiner_library = self.load_combiner_library()

                matching_cbs = [
                    cb_data for cb_data in combiner_library.values()
                    if (cb_data.get('breaker_size', 0) == breaker_size and
                        cb_data.get('fuse_holder_rating', '') == fuse_holder_rating)
                ]

                if matching_cbs:
                    matching_cbs.sort(key=lambda c: c.get('max_inputs', 0), reverse=True)
                    max_inputs_per_cb = matching_cbs[0].get('max_inputs', 24)
                else:
                    max_inputs_per_cb = 24  # fallback

                if total_all_strings > 0:
                    num_combiners = math.ceil(total_all_strings / max_inputs_per_cb)
                    strings_per_cb = math.ceil(total_all_strings / num_combiners)
                # Allocation may produce more CBs due to group boundaries.
                # Update num_combiners from allocation if available.
                if totals.get('inverter_summary', {}).get('allocation_result'):
                    num_combiners = totals['inverter_summary']['allocation_result']['summary']['total_inverters']
                num_devices = num_combiners

        # Preliminary combiner count (used for DC feeder/AC homerun distance calc)
        # Actual CB part matching is done later via Device Configurator or fallback
        # For Central Inverter, post-allocation block sets the authoritative count.
        if num_combiners > 0 and topology != 'Central Inverter':
            if breaker_size not in totals['combiners_by_breaker']:
                totals['combiners_by_breaker'][breaker_size] = 0
            totals['combiners_by_breaker'][breaker_size] += num_combiners

        # ==================== Harness split adjustment ====================
        # For now it's a no-op since allocations=[] above.
        self._adjust_harnesses_for_splits(totals)

        # Adjust inline fuse counts for split trackers to match harness adjustments
        if lv_method == 'Wire Harness' and self._split_tracker_details:
            _tsm = getattr(self, '_tracker_to_segment', [])
            for tidx, split_info in self._split_tracker_details.items():
                _seg_module = None
                if tidx < len(_tsm):
                    _seg = _tsm[tidx].get('seg')
                    if _seg:
                        _ref = _seg.get('template_ref')
                        if _ref and _ref in self.enabled_templates:
                            _seg_module = self.enabled_templates[_ref].get('module_spec')
                _fuse_rating = self._calc_inline_fuse_rating(_seg_module)
                for size in split_info['original_config']:
                    if size >= 2:
                        totals['inline_fuses_by_rating'][_fuse_rating] -= size
                for portion in split_info['portions']:
                    for size in portion['harnesses']:
                        if size >= 2:
                            totals['inline_fuses_by_rating'][_fuse_rating] += size
            # Clean up zero/negative entries
            zero_keys = [r for r, v in totals['inline_fuses_by_rating'].items() if v <= 0]
            for r in zero_keys:
                del totals['inline_fuses_by_rating'][r]

        # ==================== Trunk Bus calculation ====================
        if lv_method == 'Trunk Bus' and allocation_result:
            from src.utils.cable_sizing import recommend_trunk_cable_size, select_lbd_size
            
            for inv_data in allocation_result.get('inverters', []):
                block_strings = inv_data['total_strings']
                if block_strings <= 0:
                    continue
                
                # Count trackers in this LBD block
                block_tracker_indices = [ti for ti, _ in inv_data['tracker_indices']]
                num_trackers_in_block = len(block_tracker_indices)
                
                # Trunk cable length = (trackers - 1) × row_spacing
                if num_trackers_in_block > 1:
                    # Get row_spacing from tracker_entries for this block's trackers
                    block_row_spacings = []
                    for tidx in block_tracker_indices:
                        if tidx < len(tracker_entries):
                            block_row_spacings.append(tracker_entries[tidx].get('row_spacing_ft', 20.0))
                    avg_row_spacing = sum(block_row_spacings) / len(block_row_spacings) if block_row_spacings else 20.0
                    trunk_length_ft = (num_trackers_in_block - 1) * avg_row_spacing
                else:
                    trunk_length_ft = 0.0
                
                # Auto-size trunk cable
                trunk_cable_size = recommend_trunk_cable_size(
                    block_strings, module_isc, nec_factor
                )
                totals['trunk_cable_by_size'][trunk_cable_size] += trunk_length_ft
                
                # Auto-size LBD
                lbd_size = select_lbd_size(block_strings, module_isc, nec_factor)
                totals['lbd_by_size'][lbd_size] += 1
                
                # Count IPCs: 1 per tracker per polarity (×2), grouped by tap count
                for tidx, strings_taken in inv_data['tracker_indices']:
                    # Determine tap count: use strings_taken from this tracker
                    # Auto-select: ≤2 strings → 2-tap, ≤4 strings → 4-tap
                    if strings_taken <= 2:
                        tap_count = 2
                    else:
                        tap_count = 4
                    totals['ipc_by_tap'][tap_count] += 2  # pos + neg

        # ==================== Whip calculation (skipped for Trunk Bus) ====================
        if lv_method != 'Trunk Bus' and total_all_trackers > 0 and num_devices > 0:
            whip_distances = self.calculate_whip_distances_from_positions(
                allocation_result, topology, num_devices
            )

            split_details = getattr(self, '_split_tracker_details', {})
            seen_whip_trackers = set()
            
            for entry in whip_distances:
                if len(entry) == 4:
                    distance_ft, spt, tidx, inv_idx = entry
                else:
                    distance_ft, spt = entry[0], entry[1]
                    tidx, inv_idx = -1, -1
                whip_length = self.round_whip_length(distance_ft)
                
                if tidx in split_details:
                    # Split tracker — collect ALL portions for this inv_idx
                    # (a single inverter can own multiple portions when a harness
                    #  straddles the split boundary)
                    portion_harnesses = 0
                    for portion in split_details[tidx]['portions']:
                        if portion['inv_idx'] == inv_idx:
                            portion_harnesses += len(portion['harnesses'])
                    
                    if portion_harnesses == 0:
                        continue
                    
                    num_harnesses = portion_harnesses
                else:
                    # Non-split tracker — skip duplicates, use original harness count
                    if tidx in seen_whip_trackers:
                        debug_whip_skipped.append(f"T{tidx+1:02d} inv={inv_idx} DUPLICATE skipped")
                        continue
                    seen_whip_trackers.add(tidx)
                    num_harnesses = harness_count_by_spt.get(spt, 1)
                
                # Determine individual harness sizes for wire gauge lookup
                if tidx in split_details:
                    # Collect harness sizes from ALL portions for this inv_idx
                    ind_harness_sizes = []
                    for portion in split_details[tidx]['portions']:
                        if portion['inv_idx'] == inv_idx:
                            ind_harness_sizes.extend(portion['harnesses'])
                    if not ind_harness_sizes:
                        ind_harness_sizes = split_details[tidx]['portions'][0]['harnesses']
                else:
                    if tidx < len(tracker_harness_sizes_list):
                        ind_harness_sizes = tracker_harness_sizes_list[tidx]
                    else:
                        ind_harness_sizes = harness_sizes_by_spt.get(spt, [spt])
                
                for h_str_count in ind_harness_sizes:
                    gauge = self.get_wire_size_for('whip', h_str_count)
                    key = (whip_length, gauge)
                    totals['whips_by_length'][key] += 2  # pos + neg
                    totals['total_whip_length'] += whip_length * 2

        split_details = getattr(self, '_split_tracker_details', {})
        tracker_seg_map = getattr(self, '_tracker_to_segment', [])
        
        # Count how many split trackers exist per (group_idx, seg identity) so we can reduce bulk qty
        split_tracker_seg_counts = {}  # (group_idx, id(seg)) -> count of split trackers
        for tidx in split_details:
            if tidx < len(tracker_seg_map):
                info = tracker_seg_map[tidx]
                key = (info['group_idx'], id(info['seg']))
                split_tracker_seg_counts[key] = split_tracker_seg_counts.get(key, 0) + 1
                
        # Process non-split trackers in bulk (original logic minus split count)
        for group_idx, group in enumerate(self.groups):
            device_position = group.get('device_position', 'middle')
            for seg in group['segments']:
                if seg['quantity'] <= 0:
                    continue
                
                key = (group_idx, id(seg))
                num_splits_in_seg = split_tracker_seg_counts.get(key, 0)
                non_split_qty = seg['quantity'] - num_splits_in_seg
                
                if non_split_qty > 0:
                    extender_pairs = self.calculate_extender_lengths_per_segment(seg, device_position)
                    harness_sizes = self._get_harness_sizes(seg)
                    for pair_idx, (pos_len, neg_len) in enumerate(extender_pairs):
                        h_str_count = harness_sizes[pair_idx] if pair_idx < len(harness_sizes) else 1
                        gauge = self.get_wire_size_for('extender', h_str_count)
                        pos_rounded = self.round_whip_length(pos_len)
                        neg_rounded = self.round_whip_length(neg_len)
                        pos_key = (pos_rounded, gauge)
                        neg_key = (neg_rounded, gauge)
                        totals['extenders_pos_by_length'][pos_key] += non_split_qty
                        totals['extenders_neg_by_length'][neg_key] += non_split_qty
                        
        # Process split trackers individually — each portion gets its own extenders
        for tidx, details in split_details.items():
            if tidx >= len(tracker_seg_map):
                continue
            
            seg_info = tracker_seg_map[tidx]
            seg = seg_info['seg']
            device_position = seg_info['device_position']
            
            for portion in details['portions']:
                string_offset = portion.get('start_pos', 0)
                extender_pairs = self.calculate_extender_lengths_per_segment(
                    seg, device_position, string_offset,
                    harness_sizes_override=portion['harnesses'])
                portion_harness_sizes = portion['harnesses']
                
                for pair_idx, (pos_len, neg_len) in enumerate(extender_pairs):
                    h_str_count = portion_harness_sizes[pair_idx] if pair_idx < len(portion_harness_sizes) else 1
                    gauge = self.get_wire_size_for('extender', h_str_count)
                    pos_rounded = self.round_whip_length(pos_len)
                    neg_rounded = self.round_whip_length(neg_len)
                    pos_key = (pos_rounded, gauge)
                    neg_key = (neg_rounded, gauge)
                    totals['extenders_pos_by_length'][pos_key] += 1
                    totals['extenders_neg_by_length'][neg_key] += 1

        # Adjust extenders for trackers with inter-row N-S offset to their device.
        # Instead of adding offset uniformly, recompute with shifted target_y so
        # each extender naturally lengthens toward the CB direction.
        ns_offsets = getattr(self, '_tracker_ns_to_device', {})
        if ns_offsets:
            for (tidx, inv_idx), signed_ns in ns_offsets.items():
                if abs(signed_ns) < 1.0:
                    continue  # Skip negligible offsets

                if tidx >= len(tracker_seg_map):
                    continue

                seg_info = tracker_seg_map[tidx]
                seg = seg_info['seg']
                device_position = seg_info['device_position']

                if tidx in split_details:
                    for portion in split_details[tidx]['portions']:
                        if portion['inv_idx'] != inv_idx:
                            continue
                        string_offset = portion.get('start_pos', 0)
                        h_override = portion['harnesses']

                        # Base extenders (no offset)
                        base_pairs = self.calculate_extender_lengths_per_segment(
                            seg, device_position, string_offset,
                            harness_sizes_override=h_override)
                        # Recomputed extenders with signed N-S shift
                        adjusted_pairs = self.calculate_extender_lengths_per_segment(
                            seg, device_position, string_offset,
                            target_y_offset=signed_ns,
                            harness_sizes_override=h_override)
                        portion_harness_sizes = portion['harnesses']

                        for pair_idx in range(len(base_pairs)):
                            h_str_count = portion_harness_sizes[pair_idx] if pair_idx < len(portion_harness_sizes) else 1
                            gauge = self.get_wire_size_for('extender', h_str_count)
                            # Remove base
                            base_pos, base_neg = base_pairs[pair_idx]
                            old_pos_key = (self.round_whip_length(base_pos), gauge)
                            old_neg_key = (self.round_whip_length(base_neg), gauge)
                            totals['extenders_pos_by_length'][old_pos_key] -= 1
                            totals['extenders_neg_by_length'][old_neg_key] -= 1
                            # Add adjusted
                            adj_pos, adj_neg = adjusted_pairs[pair_idx]
                            new_pos_key = (self.round_whip_length(adj_pos), gauge)
                            new_neg_key = (self.round_whip_length(adj_neg), gauge)
                            totals['extenders_pos_by_length'][new_pos_key] += 1
                            totals['extenders_neg_by_length'][new_neg_key] += 1
                else:
                    # Non-split tracker
                    base_pairs = self.calculate_extender_lengths_per_segment(
                        seg, device_position)
                    adjusted_pairs = self.calculate_extender_lengths_per_segment(
                        seg, device_position, target_y_offset=signed_ns)
                    harness_sizes = self._get_harness_sizes(seg)

                    for pair_idx in range(len(base_pairs)):
                        h_str_count = harness_sizes[pair_idx] if pair_idx < len(harness_sizes) else 1
                        gauge = self.get_wire_size_for('extender', h_str_count)
                        # Remove base
                        base_pos, base_neg = base_pairs[pair_idx]
                        old_pos_key = (self.round_whip_length(base_pos), gauge)
                        old_neg_key = (self.round_whip_length(base_neg), gauge)
                        totals['extenders_pos_by_length'][old_pos_key] -= 1
                        totals['extenders_neg_by_length'][old_neg_key] -= 1
                        # Add adjusted
                        adj_pos, adj_neg = adjusted_pairs[pair_idx]
                        new_pos_key = (self.round_whip_length(adj_pos), gauge)
                        new_neg_key = (self.round_whip_length(adj_neg), gauge)
                        totals['extenders_pos_by_length'][new_pos_key] += 1
                        totals['extenders_neg_by_length'][new_neg_key] += 1

            # Clean up any zero-count entries
            for key_dict in [totals['extenders_pos_by_length'], totals['extenders_neg_by_length']]:
                zero_keys = [k for k, v in key_dict.items() if v <= 0]
                for k in zero_keys:
                    del key_dict[k]

        # ==================== DC Feeder and AC Homerun ====================
        dc_feeder_avg_ft = self._get_float_var(self.dc_feeder_distance_var, 500.0)
        ac_homerun_avg_ft = self._get_float_var(self.ac_homerun_distance_var, 500.0)

        total_inverters = totals.get('inverter_summary', {}).get('total_inverters', 0)
        total_combiners = sum(totals['combiners_by_breaker'].values())
        
        # Default feeder size for devices without a per-device override
        if topology == 'Distributed String':
            default_feeder_size = self.wire_sizing.get('ac_homerun', '') or '4/0 AWG'
        else:
            default_feeder_size = self.wire_sizing.get('dc_feeder', '') or '4/0 AWG'
        
        use_routed = self.use_routed_var.get() if hasattr(self, 'use_routed_var') else False

        # Build per-device distance list: [(dev_idx, label, distance_ft), ...]
        per_device_distances = []
        
        if use_routed and self.pads and allocation_result:
            try:
                routed = self.calculate_routed_feeder_distances(
                    allocation_result, topology, min((g.get('row_spacing_ft', 20.0) for g in self.groups), default=20.0)
                )
            except Exception as e:
                print(f"[Routed distance error] {e}")
                import traceback
                traceback.print_exc()
                routed = {'feeder_distances': [], 'feeder_total_ft': 0, 'feeder_count': 0}
            
            per_device_distances = routed['feeder_distances']  # already (dev_idx, label, ft)
            totals['routed_feeder_details'] = routed['feeder_distances']
        else:
            # Build per-device using average distance
            if topology == 'Distributed String':
                num_feeders = total_inverters
                avg_ft = ac_homerun_avg_ft
            else:
                num_feeders = total_combiners
                avg_ft = dc_feeder_avg_ft
            per_device_distances = [(i, f"Dev-{i+1:02d}", avg_ft) for i in range(num_feeders)]
        
        # Group primary device-to-pad cable by (per-device feeder size, parallel count)
        # - 'count'       = number of physical runs (devices)
        # - 'distance_ft' = sum of raw run distances (not multiplied) — used for avg display
        # - 'total_ft'    = distance_ft × parallel_count (actual cable footage needed)
        feeders_by_size = defaultdict(lambda: {'count': 0, 'distance_ft': 0.0, 'total_ft': 0.0})
        for dev_idx, label, dist_ft in per_device_distances:
            size = self.device_feeder_sizes.get(dev_idx, default_feeder_size)
            try:
                parallel = int(self.device_feeder_parallel_counts.get(dev_idx, 1))
                if parallel < 1:
                    parallel = 1
            except (ValueError, TypeError):
                parallel = 1
            key = (size, parallel)
            feeders_by_size[key]['count'] += 1
            feeders_by_size[key]['distance_ft'] += dist_ft
            feeders_by_size[key]['total_ft'] += dist_ft * parallel
        
        totals['feeders_by_size'] = dict(feeders_by_size)
        
        # Aggregate totals (backward compat) — total_feeder_ft reflects parallel-multiplied cable
        total_feeder_ft = sum(v['total_ft'] for v in feeders_by_size.values())
        total_feeder_count = sum(v['count'] for v in feeders_by_size.values())
        
        if topology == 'Distributed String':
            totals['dc_feeder_count'] = 0
            totals['dc_feeder_total_ft'] = 0
            totals['ac_homerun_count'] = total_feeder_count
            totals['ac_homerun_total_ft'] = total_feeder_ft
        elif topology == 'Centralized String':
            totals['dc_feeder_count'] = total_feeder_count
            totals['dc_feeder_total_ft'] = total_feeder_ft
            totals['ac_homerun_count'] = total_inverters
            totals['ac_homerun_total_ft'] = total_inverters * ac_homerun_avg_ft
        elif topology == 'Central Inverter':
            totals['dc_feeder_count'] = total_feeder_count
            totals['dc_feeder_total_ft'] = total_feeder_ft
            central_inv_count = self._get_int_var(self.central_inv_count_var, 1)
            totals['central_inverter_count'] = central_inv_count
            totals['ac_homerun_count'] = central_inv_count
            totals['ac_homerun_total_ft'] = central_inv_count * ac_homerun_avg_ft

        # ==================== Display Results ====================
        
        total_inverters = totals.get('inverter_summary', {}).get('total_inverters', 0)
        total_combiners = sum(totals['combiners_by_breaker'].values())

        # Stash display context into totals so _redraw_results_tree has everything
        totals['_display'] = {
            'topology': topology,
            'use_routed': use_routed if 'use_routed' in dir() else False,
            'dc_feeder_avg_ft': dc_feeder_avg_ft if 'dc_feeder_avg_ft' in dir() else 0,
            'ac_homerun_avg_ft': ac_homerun_avg_ft if 'ac_homerun_avg_ft' in dir() else 0,
        }
        totals['unique_modules'] = unique_modules
        totals['segment_module_data'] = segment_module_data
        totals['has_first_solar'] = any(
            'first solar' in (m.get('manufacturer', '') or '').lower()
            for m in unique_modules.values()
        )

        # Store totals for Excel export
        self.last_totals = totals
        self._results_stale = False
        
        # Combiner assignments: preserve if allocation is locked (user edited devices),
        # otherwise rebuild from scratch with correct positions.
        # For Trunk Bus, still build assignments (used for device positions/labels) but
        # they represent LBDs, not combiner boxes.
        if self.allocation_locked and self.last_combiner_assignments:
            # Edit Devices locked the allocation — its assignments are the source of truth
            pass
        else:
            # Fresh or unlocked — rebuild from allocation + harness configs
            self._build_combiner_assignments(totals, topology)

        # Read combiner BOM from Device Configurator (single source of truth)
        # Falls back to simple assignment-based totals if DC isn't available
        # For Trunk Bus, skip combiner BOM entirely — LBDs are handled separately.
        if lv_method == 'Trunk Bus':
            dc_had_data = True  # Skip combiner BOM rebuild
        elif topology == 'Central Inverter' and not self.allocation_locked:
            dc_had_data = False
        else:
            dc_had_data = self._read_combiner_bom_from_device_config()
        if not dc_had_data:
            self._rebuild_combiner_totals_from_assignments()
            
            self.last_totals = totals
            totals['combiners_by_breaker'] = {}
            totals['combiner_details'] = []
            
            # Group by breaker size
            for cb in self.last_combiner_assignments:
                bs = cb['breaker_size']
                if bs not in totals['combiners_by_breaker']:
                    totals['combiners_by_breaker'][bs] = 0
                totals['combiners_by_breaker'][bs] += 1
            
            # Rebuild combiner details with per-breaker matching
            module_isc = self.selected_module.isc if self.selected_module else 0
            for bs, qty in totals['combiners_by_breaker'].items():
                # Find max harness strings for fuse holder rating
                max_h_strings = 1
                max_inputs = 0
                for cb in self.last_combiner_assignments:
                    if cb['breaker_size'] == bs:
                        num_inputs = len(cb['connections'])
                        max_inputs = max(max_inputs, num_inputs)
                        for conn in cb['connections']:
                            max_h_strings = max(max_h_strings, conn['num_strings'])
                
                fuse_current = module_isc * 1.56 * max_h_strings
                fuse_holder_rating = self.get_fuse_holder_category(fuse_current)
                strings_per_cb = max_inputs
                
                matched_cb = self.find_combiner_box(strings_per_cb, bs, fuse_holder_rating)
                if matched_cb:
                    totals['combiner_details'].append({
                        'part_number': matched_cb.get('part_number', ''),
                        'description': matched_cb.get('description', ''),
                        'max_inputs': matched_cb.get('max_inputs', 0),
                        'breaker_size': bs,
                        'fuse_holder_rating': fuse_holder_rating,
                        'strings_per_cb': strings_per_cb,
                        'quantity': qty,
                        'block_name': 'Site Total'
                    })
                else:
                    totals['combiner_details'].append({
                        'part_number': 'NO MATCH',
                        'description': f'No CB found: {strings_per_cb} inputs, {bs}A, {fuse_holder_rating}',
                        'max_inputs': strings_per_cb,
                        'breaker_size': bs,
                        'fuse_holder_rating': fuse_holder_rating,
                        'strings_per_cb': strings_per_cb,
                        'quantity': qty,
                        'block_name': 'Site Total'
                    })

            # Update last_totals since we modified totals
            self.last_totals = totals

        # Push fresh combiner/SI assignments to Device Configurator if in QE mode
        _has_assignments = self.last_combiner_assignments or self.last_si_assignments
        if _has_assignments and not self.allocation_locked:
            main_app = getattr(self, 'main_app', None)
            if main_app and hasattr(main_app, 'device_configurator'):
                dc = main_app.device_configurator
                if getattr(dc, 'data_source', 'blocks') == 'quick_estimate':
                    try:
                        dc.load_from_quick_estimate()
                    except Exception as e:
                        print(f"[QE] Device Configurator refresh failed: {e}")

        self._redraw_results_tree()

    def _write_block_details_sheet(self, wb):
        """Write a Block Details sheet with per-device part breakdowns."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os as _os

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        ws = wb.create_sheet("Block Details")
        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1, value="Block Details — Per-Device Part Breakdown").font = title_font
        row += 2

        topology = self.topology_var.get() if hasattr(self, 'topology_var') else ''
        totals = getattr(self, 'last_totals', {}) or {}
        devices = list(getattr(self, 'last_combiner_assignments', []))

        if topology == 'Distributed String':
            inv_summary = totals.get('inverter_summary', {})
            allocation_result = inv_summary.get('allocation_result')
            if allocation_result:
                devices = []
                module_isc = self.selected_module.isc if self.selected_module else 0
                nec_factor = getattr(self.current_project, 'nec_safety_factor', 1.56) if self.current_project else 1.56
                tracker_segment_map = []
                for group in self.groups:
                    for seg in group['segments']:
                        harness_sizes = self._get_harness_sizes(seg)
                        for _ in range(seg['quantity']):
                            tracker_segment_map.append({
                                'spt': seg['strings_per_tracker'],
                                'harness_sizes': list(harness_sizes),
                                'wire_gauge': self._get_wire_gauge_for_segment(seg, 'whip'),
                            })
                for inv_idx, inv in enumerate(allocation_result['inverters']):
                    inv_name = self.device_names.get(inv_idx, f"INV-{inv_idx + 1:02d}")
                    connections = self._build_connections_from_harness_map(
                        inv.get('harness_map', []), tracker_segment_map, module_isc, nec_factor
                    )
                    devices.append({
                        'combiner_name': inv_name,
                        'device_idx': inv_idx,
                        'breaker_size': None,
                        'module_isc': module_isc,
                        'nec_factor': nec_factor,
                        'connections': connections,
                        '_is_distributed': True,
                    })

        if not devices:
            ws.cell(row=row, column=1, value="No device data available.").font = section_font
            return

        default_feeder_size = (
            self.wire_sizing.get('ac_homerun', '4/0 AWG')
            if topology == 'Distributed String'
            else self.wire_sizing.get('dc_feeder', '4/0 AWG')
        )
        dc_feeder_avg_ft = self._get_float_var(self.dc_feeder_distance_var, 500.0) if hasattr(self, 'dc_feeder_distance_var') else 500.0
        ac_homerun_avg_ft = self._get_float_var(self.ac_homerun_distance_var, 500.0) if hasattr(self, 'ac_homerun_distance_var') else 500.0

        # Site-level extender totals for proportional per-device allocation
        ext_pos_totals = totals.get('extenders_pos_by_length', {})
        ext_neg_totals = totals.get('extenders_neg_by_length', {})
        site_total_strings = sum(
            sum(conn['num_strings'] for conn in d['connections']) for d in devices
        ) or 1

        # Average whip length per gauge from aggregate totals (used for PN lookup)
        whip_avg_by_gauge = {}
        for (length, gauge), count in totals.get('whips_by_length', {}).items():
            if gauge not in whip_avg_by_gauge:
                whip_avg_by_gauge[gauge] = {'total_ft': 0.0, 'count': 0}
            whip_avg_by_gauge[gauge]['total_ft'] += length * count
            whip_avg_by_gauge[gauge]['count'] += count
        for gauge in list(whip_avg_by_gauge):
            d = whip_avg_by_gauge[gauge]
            whip_avg_by_gauge[gauge] = d['total_ft'] / d['count'] if d['count'] > 0 else 10.0

        # Load part libraries for description and PN lookups
        _cur = _os.path.dirname(_os.path.abspath(__file__))
        _root = _os.path.dirname(_os.path.dirname(_cur))

        def _load_lib(filename):
            try:
                with open(_os.path.join(_root, 'data', filename), 'r') as _f:
                    return json.load(_f)
            except Exception:
                return {}

        harness_library = _load_lib('harness_library.json')
        extender_library = _load_lib('extender_library.json')
        whip_library = _load_lib('whip_library.json')
        fuse_library = _load_lib('fuse_library.json')

        def _lib_desc(lib, pn):
            return lib.get(pn, {}).get('description', '')

        has_inline_fuses = bool(totals.get('inline_fuses_by_rating'))

        def _fuse_pn(rating):
            for pn, spec in fuse_library.items():
                if spec.get('fuse_rating_amps') == rating:
                    return pn
            candidates = [(spec.get('fuse_rating_amps', 0), pn)
                          for pn, spec in fuse_library.items()
                          if spec.get('fuse_rating_amps', 0) >= rating]
            return min(candidates)[1] if candidates else 'N/A'

        col_headers = ['Component Type', 'Part Number', 'Description', 'Quantity', 'Unit']

        _split_details_bds = getattr(self, '_split_tracker_details', {})
        _tracker_seg_map_bds = getattr(self, '_tracker_to_segment', [])

        # Accumulates {(comp_type, part_num, desc, unit): total_qty} across all devices
        _summary = {}

        for dev in devices:
            dev_name = dev['combiner_name']
            connections = dev['connections']
            dev_idx = dev['device_idx']
            is_distributed = dev.get('_is_distributed', False)

            feeder_size = self.device_feeder_sizes.get(dev_idx, default_feeder_size)
            try:
                parallel = int(self.device_feeder_parallel_counts.get(dev_idx, 1))
                if parallel < 1:
                    parallel = 1
            except (ValueError, TypeError):
                parallel = 1

            total_strings = sum(conn['num_strings'] for conn in connections)
            parallel_txt = f", {parallel}× parallel" if parallel > 1 else ""
            summary = f"{total_strings} strings allocated, {feeder_size} feeder{parallel_txt}"

            ws.merge_cells(f'A{row}:E{row}')
            hdr = ws.cell(row=row, column=1, value=f"{dev_name}  —  {summary}")
            hdr.font = section_font
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = section_fill
            row += 1

            for col, h in enumerate(col_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            def _wr(comp_type, part_num, desc, qty, unit, _row_ref=None):
                nonlocal row
                for c, v in enumerate([comp_type, part_num, desc, qty, unit], 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = thin_border
                    cell.alignment = left_align if c == 1 else center_align
                row += 1
                key = (comp_type, part_num or '', desc or '', unit or '')
                _summary[key] = _summary.get(key, 0) + (qty or 0)

            # Combiner box (non-distributed only)
            if not is_distributed and dev.get('breaker_size') and connections:
                n_inputs = len(connections)
                max_h = max(conn['num_strings'] for conn in connections)
                fuse_cat = self.get_fuse_holder_category(max_h * dev['module_isc'] * dev['nec_factor'])
                cb = self.find_combiner_box(n_inputs, dev['breaker_size'], fuse_cat)
                if cb:
                    _wr('Combiner Box', cb.get('part_number', ''), cb.get('description', ''), 1, 'ea')

            # Extenders — proportional share of site totals based on string count
            dev_fraction = total_strings / site_total_strings
            ext_gauges = sorted(
                set(g for (_, g) in ext_pos_totals) | set(g for (_, g) in ext_neg_totals)
            )
            for gauge in ext_gauges:
                sc = self._gauge_to_string_count('extender', gauge)
                pos_items = {length: qty for (length, g), qty in ext_pos_totals.items() if g == gauge}
                neg_items = {length: qty for (length, g), qty in ext_neg_totals.items() if g == gauge}
                for length in sorted(set(pos_items) | set(neg_items)):
                    pos_qty = round((pos_items.get(length, 0)) * dev_fraction)
                    neg_qty = round((neg_items.get(length, 0)) * dev_fraction)
                    if pos_qty > 0:
                        pn, _, _ = self.lookup_part_and_price('extender', polarity='positive', length_ft=length, qty=pos_qty, num_strings=sc)
                        _wr(f"Extender {length}ft (Pos)", pn, _lib_desc(extender_library, pn), pos_qty, 'ea')
                    if neg_qty > 0:
                        pn, _, _ = self.lookup_part_and_price('extender', polarity='negative', length_ft=length, qty=neg_qty, num_strings=sc)
                        _wr(f"Extender {length}ft (Neg)", pn, _lib_desc(extender_library, pn), neg_qty, 'ea')

            # Harnesses — counted from actual tracker assignments (exact, handles split trackers)
            _dev_harness_count = {}
            _seen_nonsplit = set()
            for conn in connections:
                tidx_h = conn.get('tracker_idx', -1)
                if tidx_h < 0 or tidx_h in _split_details_bds:
                    continue
                if tidx_h in _seen_nonsplit:
                    continue
                _seen_nonsplit.add(tidx_h)
                if tidx_h < len(_tracker_seg_map_bds):
                    _seg_h = _tracker_seg_map_bds[tidx_h]['seg']
                    for sz in self._get_harness_sizes(_seg_h):
                        _dev_harness_count[sz] = _dev_harness_count.get(sz, 0) + 1
                else:
                    sz = conn['num_strings']
                    _dev_harness_count[sz] = _dev_harness_count.get(sz, 0) + 1
            for tidx_h, split_info in _split_details_bds.items():
                if not any(conn.get('tracker_idx') == tidx_h for conn in connections):
                    continue
                for portion in split_info.get('portions', []):
                    if portion['inv_idx'] == dev_idx:
                        for sz in portion.get('harnesses', []):
                            _dev_harness_count[sz] = _dev_harness_count.get(sz, 0) + 1
                        break
            for sz in sorted(_dev_harness_count.keys(), reverse=True):
                dev_qty = _dev_harness_count[sz]
                if dev_qty > 0:
                    pos_pn, _, _ = self.lookup_part_and_price('harness', num_strings=sz, polarity='positive', qty=dev_qty)
                    neg_pn, _, _ = self.lookup_part_and_price('harness', num_strings=sz, polarity='negative', qty=dev_qty)
                    _wr(f"{sz}-String Harness (Pos)", pos_pn, _lib_desc(harness_library, pos_pn), dev_qty, 'ea')
                    _wr(f"{sz}-String Harness (Neg)", neg_pn, _lib_desc(harness_library, neg_pn), dev_qty, 'ea')

            # Inline fuses — one per string, derived from exact per-device harness counts
            if has_inline_fuses:
                dev_fuse_qty = sum(sz * cnt for sz, cnt in _dev_harness_count.items() if sz > 1)
                if dev_fuse_qty > 0:
                    fuse_rating = self._calc_inline_fuse_rating(self.selected_module)
                    pn = _fuse_pn(fuse_rating)
                    _wr(f"{fuse_rating}A Inline DC String Fuse (Pos)", pn, _lib_desc(fuse_library, pn), dev_fuse_qty, 'ea')

            # Whip cables by gauge
            whip_by_gauge = {}
            for conn in connections:
                g = conn.get('wire_gauge', '')
                whip_by_gauge[g] = whip_by_gauge.get(g, 0) + 1
            for gauge, count in sorted(whip_by_gauge.items()):
                avg_len = whip_avg_by_gauge.get(gauge, 0.0)
                sc = self._gauge_to_string_count('whip', gauge)
                pos_pn, _, _ = self.lookup_part_and_price('whip', polarity='positive', length_ft=avg_len, qty=count, num_strings=sc)
                neg_pn, _, _ = self.lookup_part_and_price('whip', polarity='negative', length_ft=avg_len, qty=count, num_strings=sc)
                _wr('Whip Cable (Pos)', pos_pn, _lib_desc(whip_library, pos_pn), count, 'ea')
                _wr('Whip Cable (Neg)', neg_pn, _lib_desc(whip_library, neg_pn), count, 'ea')

            # DC feeder / AC homerun
            if topology == 'Distributed String':
                total_ft = round(ac_homerun_avg_ft * parallel)
                _wr('AC Homerun', '', f"{feeder_size}, {ac_homerun_avg_ft:.0f}ft × {parallel}× parallel", total_ft, 'ft')
            else:
                total_ft = round(dc_feeder_avg_ft * parallel)
                _wr('DC Feeder (Pos)', '', f"{feeder_size}, {dc_feeder_avg_ft:.0f}ft × {parallel}× parallel", total_ft, 'ft')
                _wr('DC Feeder (Neg)', '', f"{feeder_size}, {dc_feeder_avg_ft:.0f}ft × {parallel}× parallel", total_ft, 'ft')
                if topology == 'Centralized String':
                    ac_size = self.get_wire_size_for('ac_homerun') if hasattr(self, 'get_wire_size_for') else ''
                    ac_ft = round(ac_homerun_avg_ft)
                    _wr('AC Homerun', '', f"{ac_size}, {ac_ft}ft", ac_ft, 'ft')

            row += 1  # blank row between devices

        # Summary table — all devices combined
        if _summary:
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            summary_title = ws.cell(row=row, column=1, value="Summary — All Devices Combined")
            summary_title.font = title_font
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            row += 1

            summary_headers = ['Component Type', 'Part Number', 'Description', 'Total Qty', 'Unit']
            for col, h in enumerate(summary_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            for (comp_type, part_num, desc, unit), total_qty in sorted(_summary.items()):
                for c, v in enumerate([comp_type, part_num, desc, total_qty, unit], 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = thin_border
                    cell.alignment = left_align if c == 1 else center_align
                row += 1

        # Auto-fit columns
        for col_idx in range(1, 6):
            max_len = max(len(col_headers[col_idx - 1]), 10)
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

    def _write_combiner_sheet(self, wb):
        """Write a Combiner Boxes sheet to the workbook from Device Configurator data."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get Device Configurator
        main_app = getattr(self, 'main_app', None)
        if not main_app or not hasattr(main_app, 'device_configurator'):
            return
        dc = main_app.device_configurator
        if not hasattr(dc, 'combiner_configs') or not dc.combiner_configs:
            return
        
        ws = wb.create_sheet("Combiner Boxes")
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cb_header_font = Font(bold=True, size=11)
        cb_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        mismatch_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        row = 1
        ws.merge_cells(f'A{row}:H{row}')
        ws.cell(row=row, column=1, value="Combiner Box Configuration Details").font = title_font
        row += 2
        
        # NEC factor info
        nec_factor = float(dc.nec_factor_var.get()) if hasattr(dc, 'nec_factor_var') else 1.56
        ws.cell(row=row, column=1, value="NEC Safety Factor:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=nec_factor)
        row += 2
        
        # Column headers
        headers = ['Combiner Box', 'Tracker', 'Harness', 'Strings', 'Isc (A)',
                    'Harness Current (A)', 'Fuse Size (A)', 'Cable Size',
                    'Total Current (A)', 'Breaker Size (A)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        header_row = row
        row += 1
        
        # Data rows — one section per combiner box
        for combiner_id in sorted(dc.combiner_configs.keys()):
            config = dc.combiner_configs[combiner_id]
            
            for i, conn in enumerate(config.connections):
                cb_cell = ws.cell(row=row, column=1, value=combiner_id)
                cb_cell.border = thin_border
                cb_cell.alignment = center_align
                ws.cell(row=row, column=2, value=conn.tracker_id).border = thin_border
                ws.cell(row=row, column=2).alignment = center_align
                ws.cell(row=row, column=3, value=conn.harness_id).border = thin_border
                ws.cell(row=row, column=3).alignment = center_align
                ws.cell(row=row, column=4, value=conn.num_strings).border = thin_border
                ws.cell(row=row, column=4).alignment = center_align
                ws.cell(row=row, column=5, value=round(conn.module_isc, 2)).border = thin_border
                ws.cell(row=row, column=5).alignment = center_align
                ws.cell(row=row, column=6, value=round(conn.harness_current, 2)).border = thin_border
                ws.cell(row=row, column=6).alignment = center_align
                
                fuse_cell = ws.cell(row=row, column=7, value=conn.get_display_fuse_size())
                fuse_cell.border = thin_border
                fuse_cell.alignment = center_align
                
                cable_cell = ws.cell(row=row, column=8, value=conn.get_display_cable_size())
                cable_cell.border = thin_border
                cable_cell.alignment = center_align
                if conn.is_cable_size_mismatch():
                    cable_cell.fill = mismatch_fill
                
                # Total current and breaker on first connection row only
                if i == 0:
                    ws.cell(row=row, column=9, value=round(config.total_input_current, 2)).border = thin_border
                    ws.cell(row=row, column=9).alignment = center_align
                    ws.cell(row=row, column=10, value=config.get_display_breaker_size()).border = thin_border
                    ws.cell(row=row, column=10).alignment = center_align
                else:
                    ws.cell(row=row, column=9, value='').border = thin_border
                    ws.cell(row=row, column=10, value='').border = thin_border
                
                row += 1
            
            # Blank row between combiner boxes
            row += 1
        
        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])  # Start with header width
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    def _write_string_inverter_sheet(self, wb):
        """Write a String Inverters sheet for Distributed String topology."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        main_app = getattr(self, 'main_app', None)
        if not main_app or not hasattr(main_app, 'device_configurator'):
            return
        dc = main_app.device_configurator
        if not hasattr(dc, 'string_inverter_configs') or not dc.string_inverter_configs:
            return

        ws = wb.create_sheet("String Inverters")

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        warn_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        row = 1
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value="String Inverter Configuration Details").font = title_font
        row += 2

        nec_factor = float(dc.nec_factor_var.get()) if hasattr(dc, 'nec_factor_var') else 1.56
        ws.cell(row=row, column=1, value="NEC Safety Factor:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=nec_factor)
        row += 2

        headers = ['Inverter', 'Tracker', 'Harness', 'Strings', 'Isc (A)',
                   'Harness Current (A)', 'Cable Size',
                   'Total DC Current (A)', 'MPPT Max (A)', 'Max AC Out (A)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        import re as _re
        _nat_key = lambda x: [int(c) if c.isdigit() else c.lower() for c in _re.split(r'(\d+)', x)]

        for si_id in sorted(dc.string_inverter_configs.keys(), key=_nat_key):
            si_cfg = dc.string_inverter_configs[si_id]
            mppt_max = si_cfg.get_mppt_max_current(0)
            ac_out = si_cfg.get_max_ac_output_current()
            total_dc = si_cfg.calculate_total_dc_current()

            for i, conn in enumerate(si_cfg.connections):
                over = mppt_max > 0 and conn.harness_current > mppt_max

                def _cell(r, c, v):
                    cl = ws.cell(row=r, column=c, value=v)
                    cl.border = thin_border
                    cl.alignment = center_align
                    if over:
                        cl.fill = warn_fill
                    return cl

                _cell(row, 1, si_id)
                _cell(row, 2, conn.tracker_id)
                _cell(row, 3, conn.harness_id)
                _cell(row, 4, conn.num_strings)
                _cell(row, 5, round(conn.module_isc, 2))
                _cell(row, 6, round(conn.harness_current, 2))
                _cell(row, 7, conn.actual_cable_size)

                if i == 0:
                    _cell(row, 8, round(total_dc, 2))
                    _cell(row, 9, round(mppt_max, 2) if mppt_max else '')
                    _cell(row, 10, round(ac_out, 2) if ac_out else '')
                else:
                    _cell(row, 8, '')
                    _cell(row, 9, '')
                    _cell(row, 10, '')

                row += 1
            row += 1  # blank row between inverters

        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    def export_to_excel(self, target_filepath=None, silent=False):
        """Export the quick estimate BOM to Excel
        
        Args:
            target_filepath: If provided, skip file dialog and write directly to this path.
            silent: If True, skip os.startfile and success messagebox.
        """
        
        # Run calculation first to ensure results are current
        self.calculate_estimate()

        unalloc = self._count_unallocated_strings()
        if unalloc > 0:
            messagebox.showerror(
                "Unallocated Strings",
                f"{unalloc} string(s) are not yet assigned to a device.\n\n"
                "Open Edit Devices in the Site Preview and drag all unallocated strings "
                "to a device before exporting.",
                parent=self,
            )
            return

        if not self.selected_module:
            messagebox.showwarning("No Module", "Please select a module before exporting.")
            return

        # Gather all the data we need
        modules_per_string = self._get_int_var(self.modules_per_string_var, 28)
        row_spacing = self.groups[0].get('row_spacing_ft', 20.0) if self.groups else 20.0
        
        module_isc = self.selected_module.isc
        module_width_mm = self.selected_module.width_mm
        module_width_ft = module_width_mm / 304.8
        string_length_ft = module_width_ft * modules_per_string
        
        # Build suggested filename (matching BOM Generator convention)
        def clean_filename(s):
            return "".join(c for c in s if c.isalnum() or c in (' ', '-', '_')).strip()
        
        client = "Unknown_Client"
        project_name = "Unknown_Project"
        estimate_name = "Estimate"
        
        if self.current_project and self.current_project.metadata:
            client = clean_filename(self.current_project.metadata.client or "Unknown_Client")
            project_name = clean_filename(self.current_project.metadata.name or "Unknown_Project")
        if self.estimate_id and self.current_project:
            est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
            estimate_name = clean_filename(est_data.get('name', 'Estimate'))
        
        lv_method_tag = clean_filename(self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness')
        suggested_filename = f"{client}_{project_name}_Ampacity Quick eBOM_{estimate_name}_{lv_method_tag}.xlsx"
        
        if target_filepath:
            filepath = target_filepath
        else:
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Export Quick Estimate BOM",
                initialfile=suggested_filename
            )
        
            if not filepath:
                return
        
        try:
            import re as _re
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def _group_range_label(names):
                """Return 'Group 3' for one name, 'Groups 1-4' when trailing numbers exist."""
                if len(names) == 1:
                    return names[0]
                nums = [_re.search(r'\d+$', n) for n in names]
                if all(nums):
                    return f"Groups {nums[0].group()}-{nums[-1].group()}"
                return ', '.join(names)
            
            wb = Workbook()
            info_ws = wb.active
            info_ws.title = "Project Info"
            ws = wb.create_sheet("Quick Estimate BOM")
            
            # Define styles
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            section_font = Font(bold=True, size=11)
            section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            label_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            warning_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            # ========== PROJECT INFO SHEET ==========
            info_row = 1
            
            info_ws.merge_cells(f'A{info_row}:E{info_row}')
            info_ws.cell(row=info_row, column=1, value="Quick Estimate — Project Info").font = title_font
            info_row += 2
            
            # Project info pairs
            info_items = []
            if self.current_project and self.current_project.metadata:
                meta = self.current_project.metadata
                if meta.name:
                    info_items.append(("Project Name:", meta.name))
                if meta.client:
                    info_items.append(("Customer:", meta.client))
                if meta.location:
                    info_items.append(("Location:", meta.location))
            
            # List all unique modules used
            if hasattr(self, 'last_totals') and self.last_totals.get('unique_modules'):
                seg_mod_data = self.last_totals.get('segment_module_data', [])
                for i, (mod_key, mod_data) in enumerate(self.last_totals['unique_modules'].items()):
                    prefix = "Module:" if i == 0 else f"Module {i+1}:"
                    info_items.append((prefix, mod_key))
                    info_items.append(("  Isc:", f"{mod_data.get('isc', '?')} A"))
                    info_items.append(("  Width:", f"{mod_data.get('width_mm', '?')} mm"))
                    for smd in seg_mod_data:
                        if smd.get('module_spec') == mod_data:
                            info_items.append(("  Modules/String:", str(smd['modules_per_string'])))
                            break
            else:
                info_items.append(("Module:", f"{self.selected_module.manufacturer} {self.selected_module.model} ({self.selected_module.wattage}W)"))
                info_items.append(("Module Isc:", f"{module_isc} A"))
                info_items.append(("Module Width:", f"{module_width_mm} mm"))
            # Consolidate groups that share the same row spacing into a single info row
            _spacing_map = {}
            for grp in self.groups:
                _s = round(grp.get('row_spacing_ft', 20.0), 3)
                _spacing_map.setdefault(_s, []).append(grp.get('name', 'Group'))
            for _s, _names in _spacing_map.items():
                info_items.append((f"Row Spacing ({_group_range_label(_names)}):", f"{_s:.3f} ft"))
            info_items.append(("LV Collection Method:", self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'))
            
            if self.selected_inverter:
                inv = self.selected_inverter
                info_items.append(("Inverter:", f"{inv.manufacturer} {inv.model} ({inv.rated_power_kw}kW AC)"))
                info_items.append(("Topology:", self.topology_var.get()))
                info_items.append(("DC:AC Ratio (target):", self.dc_ac_ratio_var.get()))
                if hasattr(self, 'last_totals') and self.last_totals.get('inverter_summary'):
                    inv_sum = self.last_totals['inverter_summary']
                    info_items.append(("DC:AC Ratio (actual):", f"{inv_sum.get('actual_dc_ac', 0):.2f}"))
                    info_items.append(("Total Inverters:", str(inv_sum.get('total_inverters', ''))))

                    lt = self.last_totals

                    # Build per-module-type totals from groups/segments
                    _mod_type = {}
                    for _grp in self.groups:
                        for _seg in _grp.get('segments', []):
                            _ref = _seg.get('template_ref')
                            _qty = _seg.get('quantity', 0)
                            _spt = _seg.get('strings_per_tracker', 1)
                            if _ref and _ref in self.enabled_templates and _qty > 0:
                                _td = self.enabled_templates[_ref]
                                _mps = _td.get('modules_per_string', 28)
                                _mod = _td.get('module_spec', {})
                                _lbl = f"{_mod.get('manufacturer', '?')}-{_mod.get('wattage', '?')}W"
                                if _lbl not in _mod_type:
                                    _mod_type[_lbl] = {'strings': 0, 'modules': 0}
                                _ss = _qty * int(_spt)
                                _mod_type[_lbl]['strings'] += _ss
                                _mod_type[_lbl]['modules'] += _ss * _mps

                    if len(_mod_type) > 1:
                        for _lbl, _d in _mod_type.items():
                            info_items.append((f"Strings — {_lbl}:", f"{_d['strings']:,}"))
                        info_items.append(("Total Strings:", f"{inv_sum.get('total_strings', 0):,}"))
                        if lt.get('total_modules') is not None:
                            for _lbl, _d in _mod_type.items():
                                info_items.append((f"Modules — {_lbl}:", f"{_d['modules']:,}"))
                            info_items.append(("Total Modules:", f"{lt['total_modules']:,}"))
                    else:
                        info_items.append(("Total Strings:", f"{inv_sum.get('total_strings', 0):,}"))
                        if lt.get('total_modules') is not None:
                            info_items.append(("Total Modules:", f"{lt['total_modules']:,}"))

                    if lt.get('total_dc_kw') is not None:
                        info_items.append(("DC Capacity:", f"{lt['total_dc_kw']:,.2f} kW"))
                    if lt.get('total_ac_kw') is not None:
                        info_items.append(("AC Capacity:", f"{lt['total_ac_kw']:,.2f} kW"))
            
            if self.estimate_id and self.current_project:
                est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
                info_items.append(("Estimate:", est_data.get('name', '')))
            
            for label, value in info_items:
                info_ws.cell(row=info_row, column=1, value=label).font = label_font
                info_ws.cell(row=info_row, column=2, value=value)
                info_row += 1
            
            # Add copper price from pricing manager
            try:
                from src.utils.pricing_lookup import PricingLookup
                _pricing = PricingLookup()
                copper_price = _pricing.get_current_copper_price()
                active_tier = _pricing.get_active_tier()
                info_ws.cell(row=info_row, column=1, value="Copper Price:").font = label_font
                info_ws.cell(row=info_row, column=2, value=f"${copper_price:.2f}/lb")
                info_row += 1
            except Exception:
                pass
            
            info_row += 1
            
            # ========== GROUP CONFIGURATION SUMMARY (info sheet) ==========
            info_ws.merge_cells(f'A{info_row}:E{info_row}')
            info_ws.cell(row=info_row, column=1, value="Group Configuration Summary").font = title_font
            info_row += 1
            
            row_headers = ['Group', 'Segment Configs', 'Total Strings', 'Total Trackers', 'Row Spacing (ft)']
            for col, header in enumerate(row_headers, 1):
                cell = info_ws.cell(row=info_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            info_row += 1

            for r in self.groups:
                group_strings = sum(int(s['quantity'] * s['strings_per_tracker']) for s in r['segments'])
                group_trackers = sum(s['quantity'] for s in r['segments'])
                seg_summary = ", ".join(
                    f"{s['quantity']}x{s['strings_per_tracker']}S({s['harness_config']})"
                    for s in r['segments'] if s['quantity'] > 0
                )
                row_spacing_val = r.get('row_spacing_ft', 20.0)
                group_data = [r['name'], seg_summary, group_strings, group_trackers, round(row_spacing_val, 3)]
                for col, value in enumerate(group_data, 1):
                    cell = info_ws.cell(row=info_row, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = center_align
                info_row += 1
            
            info_row += 1
            
            # ========== TRACKER SUMMARY (info sheet) ==========
            if hasattr(self, 'last_totals') and self.last_totals.get('trackers_by_string'):
                info_ws.merge_cells(f'A{info_row}:E{info_row}')
                info_ws.cell(row=info_row, column=1, value="Tracker Summary").font = title_font
                info_row += 1
                
                tracker_headers = ['Tracker Type', 'Quantity', 'Unit']
                for col, header in enumerate(tracker_headers, 1):
                    cell = info_ws.cell(row=info_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                info_row += 1
                
                total_trackers = 0
                for strings in sorted(self.last_totals['trackers_by_string'].keys()):
                    qty = self.last_totals['trackers_by_string'][strings]
                    total_trackers += qty
                    info_ws.cell(row=info_row, column=1, value=f"{strings}-String Trackers").border = thin_border
                    cell_qty = info_ws.cell(row=info_row, column=2, value=qty)
                    cell_qty.border = thin_border
                    cell_qty.alignment = center_align
                    cell_unit = info_ws.cell(row=info_row, column=3, value='ea')
                    cell_unit.border = thin_border
                    cell_unit.alignment = center_align
                    info_row += 1
                
                total_label = info_ws.cell(row=info_row, column=1, value="Total Trackers")
                total_label.font = label_font
                total_label.border = thin_border
                total_qty = info_ws.cell(row=info_row, column=2, value=total_trackers)
                total_qty.font = label_font
                total_qty.border = thin_border
                total_qty.alignment = center_align
                total_unit = info_ws.cell(row=info_row, column=3, value='ea')
                total_unit.border = thin_border
                total_unit.alignment = center_align
                info_row += 2
            
            # ========== INVERTER ALLOCATION (info sheet) ==========
            if hasattr(self, 'last_totals') and self.last_totals.get('inverter_summary'):
                inv_sum = self.last_totals['inverter_summary']
                
                info_ws.merge_cells(f'A{info_row}:E{info_row}')
                info_ws.cell(row=info_row, column=1, value="Device Allocation Summary").font = title_font
                info_row += 1

                alloc_headers = ['Device', 'Strings', 'Trackers', 'Pattern']
                for col, header in enumerate(alloc_headers, 1):
                    cell = info_ws.cell(row=info_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                info_row += 1
                
                allocation_result = inv_sum.get('allocation_result')
                if allocation_result:
                    for inv_idx, inv in enumerate(allocation_result['inverters']):
                        pattern_str = '-'.join(str(s) for s in inv['pattern'])
                        prefix = 'INV-' if self.topology_var.get() == 'Distributed String' else 'CB-'
                        inv_label = self.device_names.get(inv_idx, f"{prefix}{inv_idx + 1:02d}")
                        
                        inv_row = [
                            inv_label,
                            inv['total_strings'],
                            len(inv['tracker_indices']),
                            f"[{pattern_str}]"
                        ]
                        for col, value in enumerate(inv_row, 1):
                            cell = info_ws.cell(row=info_row, column=col, value=value)
                            cell.border = thin_border
                            cell.alignment = center_align
                        info_row += 1
            
            # Auto-fit columns on info sheet
            for col_idx in range(1, 6):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in info_ws[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                info_ws.column_dimensions[col_letter].width = min(max_length + 4, 55)
            
            # ========== PROJECT INFO HEADER (BOM sheet) ==========
            row = 1
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value="Quick Estimate — Project Info").font = title_font
            row += 2
            
            # Build the same info_items list used on the Project Info sheet
            bom_info_items = []
            if self.current_project and self.current_project.metadata:
                meta = self.current_project.metadata
                if meta.name:
                    bom_info_items.append(("Project Name:", meta.name))
                if meta.client:
                    bom_info_items.append(("Customer:", meta.client))
                if meta.location:
                    bom_info_items.append(("Location:", meta.location))
            
            if hasattr(self, 'last_totals') and self.last_totals.get('unique_modules'):
                seg_mod_data = self.last_totals.get('segment_module_data', [])
                for i, (mod_key, mod_data) in enumerate(self.last_totals['unique_modules'].items()):
                    prefix = "Module:" if i == 0 else f"Module {i+1}:"
                    bom_info_items.append((prefix, mod_key))
                    bom_info_items.append(("  Isc:", f"{mod_data.get('isc', '?')} A"))
                    bom_info_items.append(("  Width:", f"{mod_data.get('width_mm', '?')} mm"))
                    for smd in seg_mod_data:
                        if smd.get('module_spec') == mod_data:
                            bom_info_items.append(("  Modules/String:", str(smd['modules_per_string'])))
                            break
            elif self.selected_module:
                bom_info_items.append(("Module:", f"{self.selected_module.manufacturer} {self.selected_module.model} ({self.selected_module.wattage}W)"))
                bom_info_items.append(("Module Isc:", f"{module_isc} A"))
                bom_info_items.append(("Module Width:", f"{module_width_mm} mm"))
            for grp in self.groups:
                bom_info_items.append((f"Row Spacing ({grp.get('name', 'Group')}):", f"{grp.get('row_spacing_ft', 20.0):.3f} ft"))
            
            if self.selected_inverter:
                inv = self.selected_inverter
                bom_info_items.append(("Inverter:", f"{inv.manufacturer} {inv.model} ({inv.rated_power_kw}kW AC)"))
                bom_info_items.append(("Topology:", self.topology_var.get()))
                bom_info_items.append(("DC:AC Ratio (target):", self.dc_ac_ratio_var.get()))
                if hasattr(self, 'last_totals') and self.last_totals.get('inverter_summary'):
                    inv_sum = self.last_totals['inverter_summary']
                    bom_info_items.append(("DC:AC Ratio (actual):", f"{inv_sum.get('actual_dc_ac', 0):.2f}"))
                    bom_info_items.append(("Total Inverters:", str(inv_sum.get('total_inverters', ''))))
                    bom_info_items.append(("Total Strings:", f"{inv_sum.get('total_strings', 0):,}"))
                    lt = self.last_totals
                    if lt.get('total_modules') is not None:
                        bom_info_items.append(("Total Modules:", f"{lt['total_modules']:,}"))
                    if lt.get('total_dc_kw') is not None:
                        bom_info_items.append(("DC Capacity:", f"{lt['total_dc_kw']:,.2f} kW"))
                    if lt.get('total_ac_kw') is not None:
                        bom_info_items.append(("AC Capacity:", f"{lt['total_ac_kw']:,.2f} kW"))
            
            if self.estimate_id and self.current_project:
                est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
                bom_info_items.append(("Estimate:", est_data.get('name', '')))
            
            bom_info_items.append(("Date:", datetime.now().strftime('%Y-%m-%d')))
            
            # Add copper price from pricing manager
            try:
                from src.utils.pricing_lookup import PricingLookup
                _pricing = PricingLookup()
                copper_price = _pricing.get_current_copper_price()
                active_tier = _pricing.get_active_tier()
                bom_info_items.append(("Copper Price:", f"${copper_price:.2f}/lb"))
            except Exception:
                pass
            
            for label, value in bom_info_items:
                ws.cell(row=row, column=1, value=label).font = label_font
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1
            
            # ========== BOM RESULTS SECTION (BOM sheet) ==========
            ws.merge_cells(f'A{row}:G{row}')
            ws.cell(row=row, column=1, value="Estimated Bill of Materials").font = title_font
            row += 1
            
            # BOM headers
            bom_headers = ['Item', 'Part Number', 'Description', 'Quantity', 'Unit', 'Unit Cost', 'Ext. Cost']
            for col, header in enumerate(bom_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1
            
            # Pull results from the results treeview - only checked items
            bom_first_data_row = None
            for item_id in self.results_tree.get_children():
                values = self.results_tree.item(item_id, 'values')
                if len(values) < 8:
                    continue

                include = values[0]   # ☑ or ☐ or ''
                item_name = values[1]
                part_number = values[2]
                description = values[3]
                qty = values[4]
                unit = values[5]
                unit_cost = values[6]

                is_section = str(item_name).startswith('---')
                is_warning = '⚠' in str(item_name)

                # Skip unchecked non-section rows
                if not is_section and include != '☑':
                    continue

                cell_item = ws.cell(row=row, column=1, value=str(item_name).replace('---', '').strip() if is_section else item_name)
                cell_pn = ws.cell(row=row, column=2, value=part_number if not is_section else '')
                cell_desc = ws.cell(row=row, column=3, value='' if is_section else (description or ''))
                # Convert qty to number for Excel
                qty_val = ''
                if qty:
                    try:
                        qty_val = int(qty) if '.' not in str(qty) else float(qty)
                    except (ValueError, TypeError):
                        qty_val = qty
                cell_qty = ws.cell(row=row, column=4, value=qty_val)
                cell_unit = ws.cell(row=row, column=5, value=unit if unit else '')
                # Convert unit_cost to number for Excel
                unit_cost_val = ''
                if unit_cost:
                    try:
                        cleaned = str(unit_cost).replace('$', '').replace(',', '').strip()
                        unit_cost_val = float(cleaned) if cleaned else ''
                    except (ValueError, TypeError):
                        unit_cost_val = unit_cost
                cell_unit_cost = ws.cell(row=row, column=6, value=unit_cost_val)
                if isinstance(unit_cost_val, float):
                    cell_unit_cost.number_format = '"$"#,##0.00'

                # Ext. Cost: formula for all non-section rows
                if not is_section:
                    cell_ext_cost = ws.cell(row=row, column=7, value=f'=IF(F{row}="","",D{row}*F{row})')
                    cell_ext_cost.number_format = '"$"#,##0.00'
                    if bom_first_data_row is None:
                        bom_first_data_row = row
                else:
                    cell_ext_cost = ws.cell(row=row, column=7, value='')

                if is_section:
                    for c in [cell_item, cell_pn, cell_desc, cell_qty, cell_unit, cell_unit_cost, cell_ext_cost]:
                        c.font = section_font
                        c.fill = section_fill
                elif is_warning:
                    for c in [cell_item, cell_pn, cell_desc, cell_qty, cell_unit, cell_unit_cost, cell_ext_cost]:
                        c.fill = warning_fill

                for c in [cell_item, cell_pn, cell_desc, cell_qty, cell_unit, cell_unit_cost, cell_ext_cost]:
                    c.border = thin_border
                    c.alignment = center_align
                cell_item.alignment = wrap_align
                cell_desc.alignment = Alignment(horizontal='left', wrap_text=False)

                row += 1

            # Total Cost row
            if bom_first_data_row:
                row += 1
                total_label = ws.cell(row=row, column=6, value='Total Cost:')
                total_label.font = Font(bold=True)
                total_label.alignment = Alignment(horizontal='right')
                total_label.border = thin_border
                total_cell = ws.cell(row=row, column=7, value=f'=SUM(G{bom_first_data_row}:G{row - 2})')
                total_cell.font = Font(bold=True)
                total_cell.number_format = '"$"#,##0.00'
                total_cell.border = thin_border
                total_cell.alignment = center_align
                row += 1

            # ========== BLOCK DETAILS SHEET ==========
            self._write_block_details_sheet(wb)

            # ========== COMBINER BOXES / STRING INVERTERS SHEET ==========
            topology = self.topology_var.get() if hasattr(self, 'topology_var') else ''
            if topology == 'Distributed String':
                self._write_string_inverter_sheet(wb)
            else:
                self._write_combiner_sheet(wb)
            
            # ========== AUTO-FIT COLUMNS (BOM sheet) ==========
            for col_idx in range(1, 9):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                cap = 120 if col_idx == 3 else 55  # description column gets extra room
                ws.column_dimensions[col_letter].width = min(max_length + 4, cap)
            
            # Save
            wb.save(filepath)
            
            if not silent:
                # Try to open the file
                import os
                try:
                    os.startfile(filepath)
                except Exception:
                    pass
                
                messagebox.showinfo("Success", f"Quick Estimate BOM exported to:\n{filepath}")
            
            return True
            
        except PermissionError:
            messagebox.showerror(
                "Permission Error",
                f"Cannot write to {filepath}.\n\n"
                "The file may be open in Excel. Please close it and try again."
            )
            return False
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export BOM:\n{str(e)}")
            return False

    def export_packet(self):
        """Export a PDF containing the String Allocation and Harness Drawings."""
        import os

        self.calculate_estimate()

        if not self.selected_module:
            messagebox.showwarning("No Module", "Please select a module before exporting.")
            return

        inv_summary = getattr(self, 'last_totals', {}).get('inverter_summary', {})
        if not inv_summary or not inv_summary.get('allocation_result'):
            messagebox.showinfo("No Data", "Run Calculate Estimate first to generate preview data.")
            return

        def clean_fn(s):
            return "".join(c for c in s if c.isalnum() or c in (' ', '-', '_')).strip()

        client = "Unknown_Client"
        project_name = "Unknown_Project"
        estimate_name = "Estimate"

        if self.current_project and self.current_project.metadata:
            client = clean_fn(self.current_project.metadata.client or "Unknown_Client")
            project_name = clean_fn(self.current_project.metadata.name or "Unknown_Project")
        if self.estimate_id and self.current_project:
            est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
            estimate_name = clean_fn(est_data.get('name', 'Estimate'))

        suggested_name = f"{client}_{project_name}_Site PDF_{estimate_name}.pdf"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Export Site PDF with Harness Drawings",
            initialfile=suggested_name,
        )
        if not filepath:
            return

        success = self._generate_site_pdf(filepath, include_wiring=True)
        if success:
            try:
                os.startfile(filepath)
            except Exception:
                pass
            messagebox.showinfo("Success", f"PDF exported to:\n{filepath}")
        else:
            messagebox.showerror("Error", "Failed to generate PDF.")

    def _generate_site_pdf(self, filepath, include_wiring=True):
        """Generate the string allocation site PDF using current estimate data.
        
        Returns True on success, False on error.
        """
        from src.utils.site_pdf_generator import generate_site_pdf
        from .site_preview import SitePreviewWindow

        inv_summary = getattr(self, 'last_totals', {}).get('inverter_summary', {})
        topology = self.topology_var.get()
        row_spacing_ft = self.groups[0].get('row_spacing_ft', 20.0) if self.groups else 20.0

        # We need layout data. Create a temporary (hidden) SitePreviewWindow to build it,
        # then extract group_layout and device_positions.
        totals = getattr(self, 'last_totals', {})
        total_combiners = sum(totals.get('combiners_by_breaker', {}).values())

        lv_method = self.lv_collection_var.get() if hasattr(self, 'lv_collection_var') else 'Wire Harness'
        if topology == 'Distributed String':
            num_devices = totals.get('string_inverters', 0)
            device_label = 'SI'
        elif topology == 'Central Inverter':
            alloc = totals.get('inverter_summary', {}).get('allocation_result', {})
            num_devices = alloc.get('summary', {}).get('total_inverters', total_combiners)
            device_label = 'LBD' if lv_method == 'Trunk Bus' else 'CB'
        else:
            num_devices = total_combiners
            device_label = 'LBD' if lv_method == 'Trunk Bus' else 'CB'

        # Build layout via a temporary hidden preview window
        temp_preview = SitePreviewWindow(
            self, inv_summary, topology, self.INVERTER_COLORS,
            self.groups, self.enabled_templates, row_spacing_ft,
            num_devices=num_devices, device_label=device_label,
            initial_inspect=False, pads=self.pads,
            device_names=self.device_names,
            device_feeder_sizes=self.device_feeder_sizes,
            device_feeder_parallel_counts=self.device_feeder_parallel_counts
        )
        temp_preview.withdraw()  # Keep it hidden

        try:
            group_layout = temp_preview.group_layout
            device_positions = getattr(temp_preview, 'device_positions', [])

            # Build project info dict
            project_info = {
                'project_name': '',
                'customer': '',
                'location': '',
                'estimate_name': '',
                'topology': topology,
                'module_info': '',
                'total_strings': '',
                'total_devices': '',
                'dc_ac_ratio': '',
                'split_trackers': '',
                'revision': '0',
            }

            if self.current_project and self.current_project.metadata:
                meta = self.current_project.metadata
                project_info['project_name'] = meta.name or ''
                project_info['customer'] = meta.client or ''
                project_info['location'] = meta.location or ''

            if self.estimate_id and self.current_project:
                est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
                project_info['estimate_name'] = est_data.get('name', '')

            if self.selected_module:
                mod = self.selected_module
                project_info['module_info'] = f"{mod.manufacturer} {mod.model} ({mod.wattage}W)"
                project_info['module_name'] = f"{mod.manufacturer} {mod.model}"

            if self.selected_inverter:
                inv = self.selected_inverter
                project_info['inverter_info'] = f"{inv.manufacturer} {inv.model}"

            alloc_result = inv_summary.get('allocation_result', {})
            summary = alloc_result.get('summary', {})
            project_info['total_strings'] = str(summary.get('total_strings', ''))
            project_info['total_devices'] = f"{num_devices} {device_label}s"
            project_info['dc_ac_ratio'] = f"{inv_summary.get('actual_dc_ac', 0):.2f}"
            project_info['split_trackers'] = str(summary.get('total_split_trackers', ''))
            project_info['inverter_qty'] = str(totals.get('string_inverters', ''))
            project_info['dc_capacity_kw'] = f"{totals.get('total_dc_kw', 0):,.2f}"
            project_info['ac_capacity_kw'] = f"{totals.get('total_ac_kw', 0):,.2f}"
            project_info['total_modules'] = f"{totals.get('total_modules', 0):,}"

            # --- Build tracker summary for system summary table ---
            tracker_counts = {}
            for group in self.groups:
                for seg in group.get('segments', []):
                    ref = seg.get('template_ref')
                    qty = seg.get('quantity', 0)
                    if ref and qty > 0:
                        short = ref.split(' - ', 1)[1] if ' - ' in ref else ref
                        tracker_counts[short] = tracker_counts.get(short, 0) + qty
            project_info['tracker_summary'] = list(tracker_counts.items())

            # --- Per-module-type string/module totals ---
            mod_type_data = {}
            for group in self.groups:
                for seg in group.get('segments', []):
                    ref = seg.get('template_ref')
                    qty = seg.get('quantity', 0)
                    spt = seg.get('strings_per_tracker', 1)
                    if ref and ref in self.enabled_templates and qty > 0:
                        tdata = self.enabled_templates[ref]
                        mps = tdata.get('modules_per_string', 28)
                        mod = tdata.get('module_spec', {})
                        lbl = f"{mod.get('manufacturer', '?')}-{mod.get('wattage', '?')}W"
                        if lbl not in mod_type_data:
                            mod_type_data[lbl] = {'strings': 0, 'modules': 0}
                        seg_str = qty * int(spt)
                        mod_type_data[lbl]['strings'] += seg_str
                        mod_type_data[lbl]['modules'] += seg_str * mps
            if len(mod_type_data) > 1:
                project_info['module_type_totals'] = [
                    (lbl, d['strings'], d['modules'])
                    for lbl, d in mod_type_data.items()
                ]

            # --- Row spacing (consolidated by unique spacing value) ---
            import re as _re2
            def _rl(names):
                if len(names) == 1:
                    return names[0]
                nums = [_re2.search(r'\d+$', n) for n in names]
                if all(nums):
                    return f"Groups {nums[0].group()}-{nums[-1].group()}"
                return ', '.join(names)

            _rs_map = {}
            for grp in self.groups:
                _sp = round(grp.get('row_spacing_ft', 20.0), 2)
                _rs_map.setdefault(_sp, []).append(grp.get('name', 'Group'))
            if len(_rs_map) == 1:
                project_info['row_spacing_rows'] = [
                    ('Row Spacing', f"{list(_rs_map.keys())[0]:.2f} ft")
                ]
            else:
                project_info['row_spacing_rows'] = [
                    (f"Row Spacing ({_rl(names)})", f"{sp:.2f} ft")
                    for sp, names in _rs_map.items()
                ]

            # --- Build wiring specs for unique tracker templates ---
            wiring_specs = self._gather_wiring_specs() if include_wiring else None

            result = generate_site_pdf(
                filepath=filepath,
                group_layout=group_layout,
                device_positions=device_positions,
                pads=self.pads,
                colors=self.INVERTER_COLORS,
                topology=topology,
                device_label=device_label,
                project_info=project_info,
                show_routes=True,
                align_on_motor=True,
                wiring_specs=wiring_specs,
            )
            return result

        finally:
            temp_preview.destroy()

    def _gather_wiring_specs(self):
        """Gather unique tracker wiring specifications for DC cabling diagrams.
        
        Returns a list of dicts, one per unique (non-split) tracker template,
        each containing the info needed to draw a wiring diagram.
        """
        seen_refs = set()
        specs = []
        
        polarity = 'Negative Always South'  # default matches the QE combobox default
        if hasattr(self, 'polarity_convention_var'):
            polarity = self.polarity_convention_var.get()
        
        for group in self.groups:
            for seg in group.get('segments', []):
                ref = seg.get('template_ref')
                if not ref or ref in seen_refs:
                    continue
                if ref not in self.enabled_templates:
                    continue
                
                tdata = self.enabled_templates[ref]
                spt = tdata.get('strings_per_tracker', 1)
                
                # Skip half-string (non-integer SPT) trackers
                if spt != int(spt):
                    continue
                
                spt = int(spt)
                seen_refs.add(ref)
                
                # Extract short template name (strip manufacturer prefix)
                if ' - ' in ref:
                    template_short_name = ref.split(' - ', 1)[1]
                else:
                    template_short_name = ref
                
                mps = tdata.get('modules_per_string', 28)
                mod_spec = tdata.get('module_spec', {})
                orientation = tdata.get('module_orientation', 'Portrait')
                has_motor = tdata.get('has_motor', True)
                
                # Motor position data
                motor_placement_type = 'between_strings'
                motor_position_after = max(1, spt // 2)  # sensible default for between_strings
                motor_string_idx = 1
                m_split_north = mps // 2
                m_split_south = mps - mps // 2
                if has_motor:
                    motor_placement_type = tdata.get('motor_placement_type', 'between_strings')
                    if motor_placement_type == 'between_strings':
                        pos_after = tdata.get('motor_position_after_string', None)
                        str_idx = tdata.get('motor_string_index', None)
                        if pos_after is not None and int(pos_after) >= 0:
                            motor_position_after = int(pos_after)
                        elif str_idx is not None and int(str_idx) >= 0:
                            motor_position_after = int(str_idx)
                    elif motor_placement_type == 'middle_of_string':
                        str_idx = tdata.get('motor_string_index', None)
                        if str_idx is not None and int(str_idx) > 0:
                            motor_string_idx = int(str_idx)
                        m_split_north = tdata.get('motor_split_north', mps // 2)
                        m_split_south = tdata.get('motor_split_south', mps - mps // 2)
                
                # Harness config
                harness_config = self._get_effective_harness_config(seg)
                harness_sizes = self.parse_harness_config(harness_config)
                if not harness_sizes:
                    harness_sizes = [spt]
                
                # Wire gauges
                string_gauge = self.get_wire_size_for('string', 1)
                harness_gauge_map = {}
                whip_gauge_map = {}
                for h_size in set(harness_sizes):
                    harness_gauge_map[h_size] = self.get_wire_size_for('harness', h_size)
                    whip_gauge_map[h_size] = self.get_wire_size_for('whip', h_size)
                
                # Determine predominant device position from groups
                device_pos = 'south'  # default
                if hasattr(self, 'groups') and self.groups:
                    pos_counts = {}
                    for grp in self.groups:
                        dp = grp.get('device_position', 'south')
                        pos_counts[dp] = pos_counts.get(dp, 0) + 1
                    if pos_counts:
                        device_pos = max(pos_counts, key=pos_counts.get)

                specs.append({
                    'strings_per_tracker': spt,
                    'modules_per_string': mps,
                    'template_name': template_short_name,
                    'module_width_mm': mod_spec.get('width_mm', 1134),
                    'module_length_mm': mod_spec.get('length_mm', 2384),
                    'module_orientation': orientation,
                    'harness_sizes': harness_sizes,
                    'has_motor': has_motor,
                    'motor_placement_type': motor_placement_type,
                    'motor_position_after_string': motor_position_after,
                    'motor_string_index': motor_string_idx,
                    'motor_split_north': m_split_north,
                    'motor_split_south': m_split_south,
                    'polarity_convention': polarity,
                    'device_position': device_pos,
                    'wire_gauges': {
                        'string': string_gauge,
                        'harness': harness_gauge_map,
                        'whip': whip_gauge_map,
                    },
                })
        
        return specs

    def export_pdf_only(self):
        """Export just the site PDF (no Excel BOM) for quick testing."""
        self.calculate_estimate()

        unalloc = self._count_unallocated_strings()
        if unalloc > 0:
            messagebox.showerror(
                "Unallocated Strings",
                f"{unalloc} string(s) are not yet assigned to a device.\n\n"
                "Open Edit Devices in the Site Preview and drag all unallocated strings "
                "to a device before exporting.",
                parent=self,
            )
            return

        if not self.selected_module:
            messagebox.showwarning("No Module", "Please select a module before exporting.")
            return

        inv_summary = getattr(self, 'last_totals', {}).get('inverter_summary', {})
        if not inv_summary or not inv_summary.get('allocation_result'):
            messagebox.showinfo("No Data", "Run Calculate Estimate first to generate preview data.")
            return

        def clean_fn(s):
            return "".join(c for c in s if c.isalnum() or c in (' ', '-', '_')).strip()

        client = "Unknown_Client"
        project_name = "Unknown_Project"
        estimate_name = "Estimate"

        if self.current_project and self.current_project.metadata:
            client = clean_fn(self.current_project.metadata.client or "Unknown_Client")
            project_name = clean_fn(self.current_project.metadata.name or "Unknown_Project")
        if self.estimate_id and self.current_project:
            est_data = self.current_project.quick_estimates.get(self.estimate_id, {})
            estimate_name = clean_fn(est_data.get('name', 'Estimate'))

        suggested_name = f"{client}_{project_name}_String Allocation_{estimate_name}.pdf"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Export Site PDF",
            initialfile=suggested_name,
        )
        if not filepath:
            return

        success = self._generate_site_pdf(filepath, include_wiring=False)
        if success:
            import os
            try:
                os.startfile(filepath)
            except Exception:
                pass
            messagebox.showinfo("Success", f"PDF exported to:\n{filepath}")
        else:
            messagebox.showerror("Error", "Failed to generate PDF.")
    

    def _run_diagnostics(self):
        """Run all diagnostic checks and display results in a dialog."""
        from src.utils.diagnostics import (
            run_all_diagnostics, format_diagnostic_report
        )

        result = run_all_diagnostics(self, verbose=True)
        report = format_diagnostic_report(result)

        # Show in a scrollable dialog
        diag_win = tk.Toplevel(self)
        diag_win.title("Quick Estimate Diagnostics")
        diag_win.geometry("700x500")
        diag_win.transient(self.winfo_toplevel())

        text_frame = ttk.Frame(diag_win)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=('Consolas', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL,
                                   command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Color tags
        text_widget.tag_configure('pass', foreground='green')
        text_widget.tag_configure('fail', foreground='red')
        text_widget.tag_configure('issue', foreground='#CC6600')
        text_widget.tag_configure('header', font=('Consolas', 11, 'bold'))

        for line in report.split('\n'):
            if line.startswith('[PASS]'):
                text_widget.insert(tk.END, line + '\n', 'pass')
            elif line.startswith('[FAIL]'):
                text_widget.insert(tk.END, line + '\n', 'fail')
            elif line.strip().startswith('EXT_') or line.strip().startswith('SPLIT_') or \
                 line.strip().startswith('WHIP_') or line.strip().startswith('HARNESS_') or \
                 line.strip().startswith('DUPLICATE') or line.strip().startswith('MISSING') or \
                 line.strip().startswith('ORDER') or line.strip().startswith('CONTIGUITY') or \
                 line.strip().startswith('NO_DATA'):
                text_widget.insert(tk.END, line + '\n', 'issue')
            elif '===' in line or '---' in line:
                text_widget.insert(tk.END, line + '\n', 'header')
            else:
                text_widget.insert(tk.END, line + '\n')

        text_widget.configure(state=tk.DISABLED)

        # Close button
        ttk.Button(diag_win, text="Close",
                   command=diag_win.destroy).pack(pady=(0, 10))